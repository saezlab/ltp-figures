#!/usr/bin/env python2
# -*- coding: utf-8 -*-

#
#  Copyright (c) 2015-2016 - EMBL-EBI
#
#  File author(s): Dénes Türei (denes@ebi.ac.uk)
#
#  This code is not for public use.
#  Please do not redistribute.
#  For permission please contact me.
#
#  Website: http://www.ebi.ac.uk/~denes
#

#
    # filtering features accross samples and fractions
    # to find those relevant ones 
    # belonging to LTP bound lipids
#

import os
import sys
import copy
import time
import datetime

import collections
try:
    import cPickle as pickle
except:
    import pickle

import warnings
import timeit

import re

import xlrd
import openpyxl
import lxml.etree

import zlib
import base64
import struct

import numpy as np
import scipy as sp
from scipy import stats
import scipy.cluster.hierarchy as hc
import sklearn.decomposition
import fastcluster
import matplotlib as mpl
import matplotlib.backends.backend_pdf
import matplotlib.pyplot as plt
import seaborn as sns
import rpy2.robjects.packages as rpackages
rvcd = rpackages.importr('vcdExtra')
rbase = rpackages.importr('base')
rutils = rpackages.importr('utils')
rstats = rpackages.importr('stats')
rococo = rpackages.importr('rococo')

import mass
import progress
import _curl
from common import *

import rlcompleter, readline
readline.parse_and_bind('tab:complete')

warnings.filterwarnings('error')
warnings.filterwarnings('default')

class MolWeight(object):
    
    #
    # Thanks for
    # https://github.com/bsimas/molecular-weight/blob/master/chemweight.py
    #
    
    def __init__(self, formula = None, charge = 0, isotope = 0, **kwargs):
        '''
            **kwargs: elements & counts, e.g. c = 6, h = 12, o = 6...
        '''
        if not hasattr(mass, 'massFirstIso'):
            mass.getMassFirstIso()
        self.mass = mass.massFirstIso
        self.charge = charge
        self.isotope = isotope
        self.reform = re.compile(r'([A-Za-z][a-z]*)([0-9]*)')
        if formula is None:
            formula = ''.join('%s%u'%(elem.capitalize(), num) \
                for elem, num in kwargs.iteritems())
        self.formula = formula
        self.calc_weight()
    
    def __neg__(self):
        return -1 * self.weight
    
    def __add__(self, other):
        return float(other) + self.weight
    
    def __radd__(self, other):
        return self.__add__(other)
    
    def __iadd__(self, other):
        self.weight += float(other)
    
    def __sub__(self, other):
        return self.weight - float(other)
    
    def __rsub__(self, other):
        return float(other) - self.weight
    
    def __isub__(self, other):
        self.weight += float(other)
    
    def __truediv__(self, other):
        return self.weight / float(other)
    
    def __rtruediv__(self, other):
        return float(other) / self.weight
    
    def __itruediv__(self, other):
        self.weight /= float(other)
    
    def __mul__(self, other):
        return self.weight * float(other)
    
    def __rmul__(self, other):
        return self.__mul__(other)
    
    def __imul__(self, other):
        self.weight *= float(other)
    
    def __float__(self):
        return self.weight
    
    def __eq__(self, other):
        return abs(self.weight - float(other)) <= 0.01
    
    def calc_weight(self):
        atoms = self.reform.findall(self.formula)
        w = 0.0
        for element, count in atoms:
            count = int(count or '1')
            w += self.mass[element] * count
        w -= self.charge * mass.mass['electron']
        w += self.isotope * mass.mass['neutron']
        self.weight = w
    
    def reload(self):
        modname = self.__class__.__module__
        mod = __import__(modname, fromlist = [modname.split('.')[0]])
        reload(mod)
        new = getattr(mod, self.__class__.__name__)
        setattr(self, '__class__', new)

class AdductCalculator(object):
    
    def __init__(self):
        self.reform = re.compile(r'([A-Za-z][a-z]*)([0-9]*)')
        self.init_counts()
    
    def init_counts(self):
        self.counts = self.counts if hasattr(self, 'counts') else {}
    
    def formula2counts(self, formula):
        for elem, num in self.reform.findall(formula):
            yield elem, int(num or '1')
    
    def remove(self, formula):
        for elem, num in self.formula2counts(formula):
            self.add_atoms(elem, -num)
    
    def add(self, formula):
        for elem, num in self.formula2counts(formula):
            self.add_atoms(elem, num)
    
    def add_atoms(self, elem, num):
        self.counts[elem] = num if elem not in self.counts \
            else self.counts[elem] + num

class FattyFragment(MolWeight, AdductCalculator):
    
    def __init__(self, charge, c = 3, unsat = 0,
        minus = [], plus = [], isotope = 0, name = None, hg = []):
        self.c = c
        self.unsat = unsat
        self.minus = minus
        self.plus = plus
        self.hg = hg
        self.init_counts()
        self.add_atoms('C', self.c)
        self.add_atoms('H', self.c * 2)
        self.add_atoms('O', 2)
        self.add_atoms('H', self.unsat * - 2 )
        AdductCalculator.__init__(self)
        for formula in self.minus:
            self.remove(formula)
        for formula in self.plus:
            self.add(formula)
        MolWeight.__init__(self, charge = charge,
            isotope = isotope, **self.counts)
        self.set_name(name)
    
    def set_name(self, name):
        if name is None: name = ''
        self.name = '[%s(C%u:%u)%s%s]%s' % (
            name,
            self.c,
            self.unsat,
            self.adduct_str(),
            '' if self.isotope == 0 else 'i%u' % self.isotope,
            self.charge_str()
        )
    
    def charge_str(self):
        return '%s%s' % (
                ('%u' % abs(self.charge)) if abs(self.charge) > 1 else '',
                '-' if self.charge < 0 else '+'
            ) if self.charge != 0 else 'NL'
    
    def adduct_str(self):
        return '%s%s' % (
                ('-' if self.charge < 0 else '') \
                    if not self.minus  else \
                        '-%s' % ('-'.join(self.minus)),
                ('+' if self.charge > 0 else '') \
                    if not self.plus  else \
                        '+%s' % ('+'.join(self.plus))
            )
    
    def get_fragline(self):
        return [self.weight, self.name,
            '[M%s]%s' % (self.adduct_str(), self.charge_str()),
            ';'.join(self.hg)]

class LysoPEAlkenyl(FattyFragment):
    
    # from massbank.jp:
    # [lyso PE(alkenyl-18:0,-)]- 464.3140997565 -417 C23H47NO6P-
    
    def __init__(self, c, unsat = 0,
        minus = [], plus = [], isotope = 0,
        charge = -1):
        self.counts = {
            'C': 5,
            'H': 11,
            'O': 4,
            'N': 1,
            'P': 1
        }
        super(LysoPEAlkenyl, self).__init__(
            charge = charge,
            c = c,
            unsat = unsat,
            minus = minus,
            plus = plus,
            isotope = isotope,
            name = 'Lyso-PE-alkenyl',
            hg = ['PE']
        )

class LysoPE(FattyFragment):
    
    # from massbank.jp:
    # [lyso PE(18:0,-)]- 480.3090143786 -476 C23H47NO7P-
    
    def __init__(self, c, unsat = 0,
        minus = [], plus = [], isotope = 0,
        charge = -1):
        self.counts = {
            'C': 5,
            'H': 11,
            'O': 5,
            'N': 1,
            'P': 1
        }
        super(LysoPE, self).__init__(
            charge = charge,
            c = c,
            unsat = unsat,
            minus = minus,
            plus = plus,
            isotope = isotope,
            name = 'Lyso-PE',
            hg = ['PE']
        )

class LysoPEAlkyl(FattyFragment):
    
    def __init__(self, c, unsat = 0,
        minus = [], plus = [], isotope = 0,
        charge = -1):
        self.counts = {
            'C': 5,
            'H': 13,
            'O': 4,
            'N': 1,
            'P': 1
        }
        super(LysoPEAlkyl, self).__init__(
            charge = charge,
            c = c,
            unsat = unsat,
            minus = minus,
            plus = plus,
            isotope = isotope,
            name = 'Lyso-PE-alkyl',
            hg = ['PE']
        )

class LysoPCAlkenyl(FattyFragment):
    
    # from massbank.jp:
    # [lyso PC(alkenyl-18:0,-)]- 492.3453998849 -436 C25H51NO6P-
    
    def __init__(self, c, unsat = 0,
        minus = [], plus = [], isotope = 0,
        charge = -1):
        self.counts = {
            'C': 7,
            'H': 15,
            'O': 4,
            'N': 1,
            'P': 1
        }
        super(LysoPCAlkenyl, self).__init__(
            charge = charge,
            c = c,
            unsat = unsat,
            minus = minus,
            plus = plus,
            isotope = isotope,
            name = 'Lyso-PC-alkenyl',
            hg = ['PC']
        )

class LysoPCAlkyl(FattyFragment):
    
    # from massbank.jp:
    # [lyso PC(alkyl-18:0,-)]- 494.3610499491 -143 C25H53NO6P-
    
    def __init__(self, c, unsat = 0,
        minus = [], plus = [], isotope = 0,
        charge = -1):
        self.counts = {
            'C': 7,
            'H': 17,
            'O': 4,
            'N': 1,
            'P': 1
        }
        super(LysoPCAlkyl, self).__init__(
            charge = charge,
            c = c,
            unsat = unsat,
            minus = minus,
            plus = plus,
            isotope = isotope,
            name = 'Lyso-PC-alkyl',
            hg = ['PC']
        )

class LysoPC(FattyFragment):
    
    # from massbank.jp:
    # [lyso PC(18:0,-)]- 508.340314507 -373 C25H51NO7P-
    
    def __init__(self, c, unsat = 0,
        minus = [], plus = [], isotope = 0,
        charge = -1):
        self.counts = {
            'C': 7,
            'H': 15,
            'O': 5,
            'N': 1,
            'P': 1
        }
        super(LysoPC, self).__init__(
            charge = charge,
            c = c,
            unsat = unsat,
            minus = minus,
            plus = plus,
            isotope = isotope,
            name = 'Lyso-PC',
            hg = ['PC']
        )

class LysoPS(FattyFragment):
    
    # from massbank.jp:
    # [lyso PS(18:0,-)]- 437.2668152129 -358 C21H42O7P-
    
    def __init__(self, c, unsat = 0,
        minus = [], plus = [], isotope = 0,
        charge = -1):
        self.counts = {
            'C': 3,
            'H': 6,
            'O': 5,
            'P': 1
        }
        super(LysoPS, self).__init__(
            charge = charge,
            c = c,
            unsat = unsat,
            minus = minus,
            plus = plus,
            isotope = isotope,
            name = 'Lyso-PS',
            hg = ['PS']
        )

class LysoPSAlkenyl(FattyFragment):
    
    def __init__(self, c, unsat = 0,
        minus = [], plus = [], isotope = 0,
        charge = -1):
        self.counts = {
            'C': 3,
            'H': 6,
            'O': 4,
            'P': 1
        }
        super(LysoPSAlkenyl, self).__init__(
            charge = charge,
            c = c,
            unsat = unsat,
            minus = minus,
            plus = plus,
            isotope = isotope,
            name = 'Lyso-PS-alkenyl',
            hg = ['PS']
        )

class LysoPSAlkyl(FattyFragment):
    
    def __init__(self, c, unsat = 0,
        minus = [], plus = [], isotope = 0,
        charge = -1):
        self.counts = {
            'C': 3,
            'H': 8,
            'O': 4,
            'P': 1
        }
        super(LysoPSAlkyl, self).__init__(
            charge = charge,
            c = c,
            unsat = unsat,
            minus = minus,
            plus = plus,
            isotope = isotope,
            name = 'Lyso-PS-alkyl',
            hg = ['PS']
        )

class LysoPI(FattyFragment):
    
    # from massbank:
    # [lyso PI(-,18:0)]- 599.3196386444 -432 C27H52O12P-
    
    def __init__(self, c, unsat = 0,
        minus = [], plus = [], isotope = 0,
        charge = -1):
        self.counts = {
            'C': 9,
            'H': 16,
            'O': 10,
            'P': 1
        }
        super(LysoPI, self).__init__(
            charge = charge,
            c = c,
            unsat = unsat,
            minus = minus,
            plus = plus,
            isotope = isotope,
            name = 'Lyso-PI',
            hg = ['PI']
        )

class LysoPIAlkyl(FattyFragment):
    
    # from massbank.jp:
    # [lyso PI(alkyl-16:1,-)-H2O]- 537.2828592076 -228 C25H46O10P-
    # from this, derived 18:0-:
    # [lyso PI(alkyl-18:0,-)]- 585.3403740858 -228 C27H54O11P-
    
    def __init__(self, c, unsat = 0,
        minus = [], plus = [], isotope = 0,
        charge = -1):
        self.counts = {
            'C': 9,
            'H': 18,
            'O': 9,
            'P': 1
        }
        super(LysoPIAlkyl, self).__init__(
            charge = charge,
            c = c,
            unsat = unsat,
            minus = minus,
            plus = plus,
            isotope = isotope,
            name = 'Lyso-PI-alkyl',
            hg = ['PI']
        )

class LysoPG(FattyFragment):
    
    # from massbank:
    # [lyso PG(18:0,-)]- 511.3035946497 -495 C24H48O9P-
    
    def __init__(self, c, unsat = 0,
        minus = [], plus = [], isotope = 0,
        charge = -1):
        self.counts = {
            'C': 6,
            'H': 12,
            'O': 7,
            'P': 1
        }
        super(LysoPG, self).__init__(
            charge = charge,
            c = c,
            unsat = unsat,
            minus = minus,
            plus = plus,
            isotope = isotope,
            name = 'Lyso-PG',
            hg = ['PG']
        )

class LysoPA(FattyFragment):
    
    # from Characterization of Phospholipid 
    # Molecular Species by Means of HPLC-Tandem Mass Spectrometry:
    # [lyso PA(18:1,-)]- 435.2 C21H40O7P-
    
    def __init__(self, c, unsat = 0,
        minus = [], plus = [], isotope = 0,
        charge = -1):
        self.counts = {
            'C': 3,
            'H': 6,
            'O': 5,
            'P': 1
        }
        super(LysoPA, self).__init__(
            charge = charge,
            c = c,
            unsat = unsat,
            minus = minus,
            plus = plus,
            isotope = isotope,
            name = 'Lyso-PA',
            hg = ['PA']
        )

class FAminusH(FattyFragment):
    
    def __init__(self, c, unsat = 0, isotope = 0, **kwargs):
        super(FAminusH, self).__init__(charge = -1, c = c, unsat = unsat,
            minus = ['H'], isotope = isotope, name = 'FA')

class FAAlkenylminusH(FattyFragment):
    
    def __init__(self, c, unsat = 0, isotope = 0, **kwargs):
        self.counts = {
            'O': -1
        }
        super(FAAlkenylminusH, self).__init__(charge = -1, c = c, unsat = unsat,
            minus = ['H'], isotope = isotope, name = 'FA-alkenyl')

class NLFAminusH2O(FattyFragment):
    
    def __init__(self, c, unsat = 0, isotope = 0, **kwargs):
        super(NLFAminusH2O, self).__init__(charge = 0, c = c, unsat = unsat,
            minus = ['H2O'], isotope = isotope, name = 'NL FA-H2O')

class FAplusGlycerol(FattyFragment):
    
    def __init__(self, c, unsat = 0, isotope = 0,
        minus = [], plus = [], **kwargs):
        self.counts = {
            'C': 3,
            'H': 5,
            'O': 1
        }
        super(FAplusGlycerol, self).__init__(charge = 1, c = c, unsat = unsat,
            minus = minus, plus = plus,
            isotope = isotope, name = 'FA+G', hg = ['PG', 'BMP'])

class SphingosineBase(FattyFragment):
    
    def __init__(self, c, unsat = 0, isotope = 0,
        minus = [], plus = [], **kwargs):
        self.counts = {
            'N': 1,
            'H': 4
        }
        super(SphingosineBase, self).__init__(charge = 1, c = c, unsat = unsat,
            isotope = isotope, name = 'Sphingosine',
            minus = minus, plus = plus,
            hg = ['Sph', 'SM', 'Cer', 'HexCer'])

class FAFragSeries(object):
    
    def __init__(self, typ, charge, cmin = 2, unsatmin = 0,
        cmax = 36, unsatmax = 6,
        minus = [], plus = [], **kwargs):
        for attr, val in locals().iteritems():
            setattr(self, attr, val)
        self.fragments = []
        if self.unsatmax is None: self.unsatmax = self.unsatmin
        for unsat in xrange(self.unsatmin, self.unsatmax):
            this_cmin = max(self.cmin, unsat * 2 + 1)
            if this_cmin <= cmax:
                for cnum in xrange(this_cmin, cmax + 1):
                    self.fragments.append(
                        self.typ(charge = charge, c = cnum, unsat = unsat,
                            minus = self.minus, plus = self.plus, **kwargs)
                    )
    
    def __iter__(self):
        for fr in self.fragments:
            yield fr
    
    def iterweight(self):
        for fr in self.fragments:
            yield fr.weight
    
    def iterfraglines(self):
        for fr in self.fragments:
            yield fr.get_fragline()

class Mz():
    
    def __init__(self, mz, z = 1, sign = None, tolerance = 0.01):
        self.mz = mz
        self.z = z
        self.sign = sign
        self.tol = tolerance
    
    def __eq__(self, other):
        return self.z == other.z and \
            self.mz > other.mz - self.tol and \
            self.mz < other.mz + self.tol
    
    def __str__(self):
        return 'm/z = %f' % self.mz
    
    def adduct(self, m):
        return (self.mz * self.z + float(m)) / abs(self.z)
    
    def weight(self):
        return self.mz * self.z
    
    def remove_h(self):
        return self.adduct(-mass.proton)
    
    def remove_ac(self):
        m = MolWeight('H3C2O2')
        return self.adduct(-m - mass.electron)
    
    def remove_fo(self):
        m = MolWeight('HCO2')
        return self.adduct(-m - mass.electron)
    
    def remove_nh4(self):
        m = MolWeight('NH4')
        return self.adduct(-m + mass.electron)
    
    def remove_oh(self):
        m = MolWeight('OH')
        return self.adduct(-m - mass.electron)
    
    def add_h(self):
        return self.adduct(mass.proton)
    
    def add_2h(self):
        return self.adduct(2 * mass.proton)
    
    def add_3h(self):
        return self.adduct(3 * mass.proton)
    
    def add_oh(self):
        m = MolWeight('OH')
        return self.adduct(m + mass.electron)
    
    def add_fo(self):
        m = MolWeight('HCO2')
        return self.adduct(m + mass.electron)
    
    def add_ac(self):
        m = MolWeight('H3C2O2')
        return self.adduct(m + mass.electron)
    
    def add_nh4(self):
        m = MolWeight('NH4')
        return self.adduct(m - mass.electron)
    
    def reload(self):
        modname = self.__class__.__module__
        mod = __import__(modname, fromlist = [modname.split('.')[0]])
        reload(mod)
        new = getattr(mod, self.__class__.__name__)
        setattr(self, '__class__', new)

# ##

class Feature(object):
    '''
    Provides additional, more sophisticated methods
    for identification of a single feature.
    
    In the original concept all methods for identification
    based on MS1 and MS2 took place in class LTP(),
    as those could simply iterate through the arrays.
    
    Later more complex methods became necessary, so
    I created this class to group them.
    '''
    
    def __init__(self, main, protein, mode, oi, log = True):
        '''
        @main : ltp.LTP() instance
            One LTP() instance with MS1 and MS2 processing already done.
        
        @protein : str
            Protein name
        
        @mode : str
            MS mode (`pos` or `neg`)
        
        @oi : int
            Original index of one feature.
        
        @log : bool
            Whether output verbose messages to logfile.
        '''
        self.main = main
        self.log = log
        self.protein = protein
        self.mode = mode
        self.oi = oi
        self.tbl = self.main.valids[self.protein][self.mode]
        self.ms2 = self.tbl['ms2'][self.oi]
        self.i = self.main.oi2i(self.protein, self.mode, self.oi)
        self.fa = {}
        self.scans_fractions = map(
            lambda tpl: tuple(map(int, tpl)),
            uniqList(
                map(
                    tuple,
                    self.ms2[:,[12,14]]
                )
            )
        )
        self.classes = ['PA', 'PC', 'PE', 'PG', 'PS']
        self.identities = set([])
        # get carbon counts from MS1
        self.ms1fa = self.tbl['ms1fa'][oi]
        # sorting by fractions/scans
        self.scans = dict(
            map(
                lambda (sc, fr):
                    (
                        (sc, fr),
                        self.ms2[
                            np.where(
                                np.logical_and(
                                    self.ms2[:,12] == sc,
                                    self.ms2[:,14] == fr
                                )
                            )
                        ]
                    ),
                self.scans_fractions
            )
        )
        # sorting by intensity desc
        self.scans = dict(
            map(
                lambda (k, v):
                    (
                        k,
                        v[v[:,2].argsort()[::-1],:]
                    ),
                self.scans.iteritems()
            )
        )
        self._scans = dict(
            map(
                lambda (k, v):
                    (
                        k,
                        MS2Scan(v, self)
                    ),
                self.scans.iteritems()
            )
        )
        self.maxins = dict(
            map(
                lambda (k, v):
                    (
                        k,
                        v[0,2]
                    ),
                self.scans.iteritems()
            )
        )
        self.medins = dict(
            map(
                lambda (k, v):
                    (
                        k,
                        np.median(v[:,2])
                    ),
                self.scans.iteritems()
            )
        )
        self.msg('\n::: Analysing feature: %s :: %s :: index = %u ::'\
                ' m/z = %.03f :: number of MS2 scans: %u\n' % \
            (self.protein, self.mode, self.oi, self.tbl['mz'][self.i],
                len(self._scans))
        )
        self.msg('\n::: Database lookup resulted '\
            'the following species: %s\n' % self.print_db_species())
        self.msg('\n::: Intensities:\n%s%s\n' % \
            (' ' * 24, '          '.join(['A09', 'A10', 'A11', 'A12', 'B01'])))
        self.msg('%s%s' % (' ' * 16, '=' * 63))
        self.msg('\n    - absolute:  %s' % '   '.join(
            map(lambda x: '%10.01f' % x, self.tbl['fe'][self.i,:]))
        )
        self.msg('\n    - relative: %s\n' % \
            '  '.join(
                map(
                    lambda xx:
                        '%10.02f%%' % (xx * 100.0),
                    map(
                        lambda x:
                            x / np.nanmax(self.tbl['fe'][self.i,:]),
                        self.tbl['fe'][self.i,:]
                    )
                )
            )
        )
        self.msg('\n::: MS2 scans available (%u):\n\n' % len(self.scans))
        
        for sc in self.scans:
            self.print_scan(sc)
    
    def print_db_species(self):
        return ', '.join(
            map(
                lambda hg:
                    '%s' % (
                        hg \
                            if hg not in self.tbl['ms1fa'][self.oi] \
                            or not len(self.tbl['ms1fa'][self.oi][hg]) \
                            else \
                        ', '.join(
                            map(
                                lambda fa:
                                    '%s(%s)' % (hg, fa),
                                self.tbl['ms1fa'][self.oi][hg]
                            )
                        )
                    ),
                self.tbl['ms1hg'][self.oi]
            )
        ) \
        if len(self.tbl['ms1hg'][self.oi]) \
        else 'none'
    
    def reload(self):
        modname = self.__class__.__module__
        mod = __import__(modname, fromlist = [modname.split('.')[0]])
        reload(mod)
        new = getattr(mod, self.__class__.__name__)
        setattr(self, '__class__', new)
    
    def __str__(self):
        return ', '.join(
            map(
                lambda (hg, fas):
                    ', '.join(
                        map(
                            lambda fa:
                                '%s(%s)' % (hg, fa),
                            fas
                        )
                    ),
                self.fa.iteritems()
            )
        )
    
    def print_scan(self, scan):
        ms1mz = self.tbl['mz'][self.i]
        header = '\tFrag. m/z\tIntensity\tIdentity%sNL mass\n'\
            '\t%s\n' % (' ' * 26, '=' * 73)
        table = '\n\t'.join(
            map(
                lambda sc:
                    '%9.4f\t%10.2f\t%s%s%9.4f' % \
                        tuple(list(sc[[1, 2, 7]]) + \
                            [' ' * (32 - len(sc[7])), ms1mz - sc[1]]),
                self.scans[scan]
            )
        )
        fri = scan[1] - 9 if scan[1] != 1 else 4
        self.msg('\tScan %u (fraction %s%u; %s %s; intensity = %.01f (%.02f%%)):\n\n%s\t%s\n\n' % \
            (scan[0],
             'A' if 8 < scan[1] < 13 else 'B',
             scan[1],
             'contains' \
                if self.main.samples_upper[self.protein][fri + 1] == 1 \
                else 'does not contain',
             self.protein,
             self.tbl['fe'][self.i, fri] \
                 if fri < self.tbl['fe'].shape[1] else np.nan,
             (self.tbl['fe'][self.i, fri] \
                 if fri < self.tbl['fe'].shape[1] else np.nan) / \
                 np.nanmax(self.tbl['fe'][self.i, :]) * 100.0,
             header,
             table)
        )
    
    def msg(self, text):
        if self.log:
            with open(self.main.ms2log, 'a') as f:
                f.write(text)
    
    def _any_scan(self, method, **kwargs):
        for i, sc in self._scans.iteritems():
            self.msg('\t\t:: Calling method %s() on scan #%u\n' % (method, i[0]))
            if getattr(sc, method)(**kwargs):
                return True
        return False
    
    def identify(self):
        for hg in self.classes:
            self.msg('\t>>> Attempting to identify %s in all scans\n' % (hg))
            if self._any_scan('is_%s' % hg.lower()):
                self.identities.add(hg)
                self.msg('\t<<< Result: identified as %s\n' % hg)
            else:
                self.msg('\t<<< Result: not %s\n' % hg)

class MS2Scan(object):
    
    def __init__(self, scan, feature):
        self.scan = scan
        self.feature = feature
        self.recc = re.compile(r'.*[^0-9]([0-9]{1,2}):([0-9]).*')
        self.fa = {}
        self.fa1 = {}
    
    def reload(self):
        modname = self.__class__.__module__
        mod = __import__(modname, fromlist = [modname.split('.')[0]])
        reload(mod)
        new = getattr(mod, self.__class__.__name__)
        setattr(self, '__class__', new)
    
    def most_abundant_mz(self):
        result = self.scan[0,1]
        self.feature.msg('\t\t  -- Most abundant m/z is %.03f\n' % result)
        return result
    
    def mz_match(self, mz_detected, mz):
        return abs(mz_detected - mz) <= self.feature.main.ms2_tlr
    
    def mz_lookup(self, mz):
        '''
        Returns the index of the closest m/z value
        detected in the scan if it is within the
        range of tolerance, otherwise None.
        '''
        self.scan = self.scan[self.scan[:,1].argsort(),:]
        ui = self.scan[:,1].searchsorted(mz)
        if ui < self.scan.shape[0]:
            du = mz - self.scan[ui,1]
        if ui > 0:
            dl = self.scan[ui - 1,1] - mz
        i = ui if du < dl else ui - 1
        i = i if self.mz_match(self.scan[i,1], mz) else None
        sort = self.scan[:,2].argsort()[::-1]
        if i is not None:
            i = np.where(sort == i)[0][0]
        self.scan = self.scan[sort,:]
        return i
    
    def has_mz(self, mz):
        result = self.mz_lookup(mz) is not None
        self.feature.msg('\t\t  -- m/z %.03f occures in this scan? -- %s\n' % \
            (mz, str(result)))
        return result
    
    def most_abundant_mz_is(self, mz):
        result = self.mz_match(self.most_abundant_mz(), mz)
        self.feature.msg('\t\t  -- m/z %.03f is the most abundant? -- %s\n' % \
            (mz, str(result)))
        return result
    
    def mz_among_most_abundant(self, mz, n = 2):
        result = False
        for i in xrange(min(n, self.scan.shape[0])):
            if self.mz_match(self.scan[i,1], mz):
                result = True
                break
        self.feature.msg('\t\t  -- m/z %.03f is among the %u most abundant? -- '\
            '%s\n' % (mz, n, str(result)))
        return result
    
    def mz_percent_of_most_abundant(self, mz, percent = 80.0):
        insmax = self.scan[0,2]
        result = False
        for frag in self.scan:
            if self.mz_match(frag[1], mz):
                result = True
                break
            if frag[2] < insmax * 100.0 / percent:
                result = False
                break
        self.feature.msg('\t\t  -- m/z %.03f has abundance at least %.01f %% of'\
            ' the highest abundance? -- %s\n' % \
            (mz, percent, str(result)))
        return result
    
    def fa_type_is(self, i, fa_type):
        result = fa_type in self.scan[i,8] or fa_type in self.scan[i,7]
        self.feature.msg('\t\t  -- Fragment #%u (%s, %s): fatty acid type '\
            'is %s?  -- %s\n' % \
                (i, self.scan[i,7], self.scan[i,8], fa_type, str(result)))
        return result
    
    def is_fa(self, i):
        result = 'FA' in self.scan[i,7] or self.scan[i,7][:7] == 'Lyso-PA'
        self.feature.msg('\t\t  -- Fragment #%u (%s): is fatty acid? '\
            '-- %s\n' % (i, self.scan[i,7], str(result)))
        return result
    
    def most_abundant_fa(self, fa_type, head = 1):
        result = False
        for i in xrange(self.scan.shape[0]):
            if i == head:
                break
            if self.is_fa(i):
                result = self.fa_type_is(i, fa_type)
        self.feature.msg('\t\t  -- Having fatty acid %s among %u most abundant '\
            'features? -- %s\n' % (fa_type, head, str(result)))
        return result
    
    def fa_among_most_abundant(self, fa_type, n = 2, min_mass = None):
        fa_frags = 0
        for i in xrange(self.scan.shape[0]):
            if self.is_fa(i) and self.scan[i,1] >= min_mass:
                fa_frags += 1
                if min_mass is not None:
                    self.feature.msg('\t\t\t-- Fragment #%u having mass larger '\
                        'than %.01f\n' % (i, min_mass))
                if self.fa_type_is(i, fa_type):
                    result = True
                if fa_frags == n:
                    break
            elif min_mass is not None:
                self.feature.msg('\t\t\t-- Fragment #%u having mass lower '\
                        'than %.01f\n' % (i, min_mass))
        self.feature.msg('\t\t  -- Having fatty acid fragment %s among %u most '\
            'abundant -- %s\n' % (fa_type, n, str(result)))
        return result
    
    def fa_percent_of_most_abundant(self, fa_type, percent = 80.0):
        insmax = self.scan[0,2]
        for i in xrange(self.scan.shape[0]):
            if self.is_fa(i):
                if self.fa_type_is(i, fa_type):
                    return True
            if self.scan[i,2] < insmax * 100.0 / percent:
                return False
        return False
    
    def mz_most_abundant_fold(self, mz, fold):
        result = False
        if self.most_abundant_mz_is(mz):
            result = self.scan.shape[0] == 1 or \
                self.scan[1,2] * fold <= self.scan[0,2]
        self.feature.msg('\t\t  -- m/z %.03f is at least %u times higher than '\
            'any other? -- %s\n' % (mz, fold, str(result)))
        return result
    
    def get_cc(self, fa):
        m = self.recc.match(fa)
        if m is not None:
            return tuple(map(int, m.groups()))
        return (None, None)
    
    def most_abundant_fa_cc(self, fa_type = None, head = 2):
        fa_cc = []
        for i, frag in enumerate(self.scan):
            if i == head:
                break
            if self.is_fa(i) and (fa_type is None or self.fa_type_is(i, fa_type)):
                cc = self.get_cc(frag[7])
                if cc[0] is not None:
                    fa_cc.append((cc, frag[2]))
        return fa_cc
    
    def cc2str(self, cc):
        return '%u:%u' % cc
    
    def sum_cc2str(self, ccs):
        return self.cc2str(
            tuple(
                reduce(
                    lambda ((c1, uns1), ins1), ((c2, uns2), ins2):
                        (c1 + c2, uns1 + uns2),
                    ccs
                )
            )
        )
    
    def add_fa1(self, fa, hg):
        if hg not in self.fa1:
            self.fa1[hg] = set([])
            self.fa1[hg].add(
                tuple(
                    map(
                        lambda (_fa, ins):
                            _fa,
                        fa
                    )
                )
            )
            fastr = ', '.join(
                    map(
                        lambda (_fa, ins):
                            self.cc2str(_fa),
                        fa
                    )
                )
            self.feature.msg('\t\t  -- Adding fatty acids %s at headgroup '\
                '%s\n' % (fastr, hg))
    
    def fa_ccs_agree_ms1(self, hg, fa_type = None, head = 2):
        fa_cc = self.most_abundant_fa_cc(fa_type = fa_type, head = head)
        if len(fa_cc) > 0:
            cc = self.sum_cc2str([fa_cc[0]] * 2)
            agr = self.fa_cc_agrees_ms1(cc, hg)
            if agr:
                self.add_fa1(fa_cc[:1], hg)
            if len(fa_cc) > 1:
                cc = self.sum_cc2str(fa_cc[:2])
                agr = self.fa_cc_agrees_ms1(cc, hg)
                if agr:
                    self.add_fa1(fa_cc[:2], hg)
        return hg in self.fa
    
    def fa_cc_agrees_ms1(self, cc, hg):
        result = False
        if hg in self.feature.ms1fa and cc in self.feature.ms1fa[hg]:
            if hg not in self.feature.fa:
                self.feature.fa[hg] = set([])
            if hg not in self.fa:
                self.fa[hg] = set([])
            self.feature.fa[hg].add(cc)
            self.fa[hg].add(cc)
            result = True
        self.feature.msg('\t\t  -- Carbon count from MS2: %s; from databases '\
            'lookup: %s -- Any of these matches: %s\n' % \
                (
                    cc,
                    str(self.feature.ms1fa[hg]) \
                        if hg in self.feature.ms1fa else '-',
                    str(result))
                )
        return result
    
    def is_something(self, hg):
        fa_cc = self.most_abundant_fa_cc(self, )
    
    def is_pe(self):
        if self.feature.mode == 'pos':
            return self.pa_pe_ps_pg_pos('PE')
        else:
            return self.pe_pc_pg_neg('PE')
    
    def is_pc(self):
        if self.feature.mode == 'pos':
            return self.pc_pos('PC')
        else:
            return self.pe_pc_pg_neg('PC')
    
    def is_pa(self):
        if self.feature.mode == 'pos':
            return self.pa_pe_ps_pg_pos('PA')
        else:
            return self.pa_ps_neg('PA')
    
    def is_ps(self):
        if self.feature.mode == 'pos':
            return self.pa_pe_ps_pg_pos('PS')
        else:
            return self.pa_ps_neg('PS')
    
    def is_pg(self):
        if self.feature.mode == 'pos':
            return self.pa_pe_ps_pg_pos('PG')
        else:
            return self.pe_pc_pg_neg('PG')
    
    def pa_pe_ps_pg_pos(self, hg):
        return self.mz_among_most_abundant(141.0191) \
            and self.fa_among_most_abundant('-O]+', min_mass = 140.0) \
            and self.fa_ccs_agree_ms1(hg, '-O]+')
    
    def pa_ps_neg(self, hg):
        return self.has_mz(152.9958366) and self.has_mz(78.95905658) \
            and self.most_abundant_fa('-H]-') \
            and self.fa_ccs_agree_ms1(hg, '-H]-')
    
    def pe_pc_pg_neg(self, hg):
        return self.most_abundant_fa('-H]-') \
            and self.fa_ccs_agree_ms1(hg, '-H]-')
    
    def pc_pos(self, hg):
        return self.mz_most_abundant_fold(184.0733, 3) \
            and self.fa_ccs_agree_ms1(hg, head = 4)

# ##

class LTP(object):
    
    def __init__(self, **kwargs):
        self.defaults = {
            'path_root': '/',
            'basedir': ['home', 'denes', 'Documents' , 'ltp'],
            'data_basedir': None,
            'ltpdirs': [['share'], ['share', '2015_06_Popeye']],
            'samplesf': 'control_sample.csv',
            'ppfracf': 'fractions.csv',
            'ppsecdir': 'SEC_profiles',
            'stddir': 'Standards_mzML format',
            'seqfile': 'Sequence_list_LTP_screen_2015.csv',
            'pptablef': 'proteins_by_fraction.csv',
            'lipnamesf': 'lipid_names_v2.csv',
            'bindpropf': 'binding_properties.csv',
            'metabsf': 'Metabolites.xlsx',
            'swisslipids_url': 'http://www.swisslipids.org/php/'\
                'export.php?action=get&file=lipids.csv',
            'lipidmaps_url': 'http://www.lipidmaps.org/resources/downloads/'\
                'LMSDFDownload28Jun15.tar.gz',
            'lipidmaps_fname': 'LMSDFDownload28Jun15/'\
                'LMSDFDownload28Jun15FinalAll.sdf',
            'pfragmentsfile': 'lipid_fragments_positive_mode_v3.txt',
            'nfragmentsfile': 'lipid_fragments_negative_mode_v3.txt',
            'featurescache': 'features.pickle',
            'pprofcache': 'pprofiles_raw.pickle',
            'auxcache': 'save.pickle',
            'stdcachefile': 'calibrations.pickle',
            'validscache': 'valids.pickle',
            'ms2log': 'ms2identities.log',
            'swl_levels': ['Species'],
            'ms1_tolerance': 0.01,
            'ms2_tolerance': 0.02,
            'std_tolerance': 0.02,
            'aadducts': {
                1: {
                    'pos': {
                        '[M+H]+': 'remove_h',
                        '[M+NH4]+': 'remove_nh4'
                    },
                    'neg': {
                        '[M-H]-': 'add_h',
                        '[M+Fo]-': 'remove_fo'
                    }
                },
                2: {
                    'pos': {},
                    'neg': {
                        '[M-2H]2-': 'add_2h'
                    }
                },
                3: {
                    'pos': {},
                    'neg': {
                        '[M-3H]3-': 'add_3h'
                    }
                }
            },
            'metrics': [
                ('Kendall\'s tau', 'ktv', False),
                ('Spearman corr.', 'spv', False),
                ('Pearson corr.', 'pev', False),
                ('Euclidean dist.', 'euv', True),
                ('Robust corr.', 'rcv', False),
                ('Goodman-Kruskal\'s gamma', 'gkv', False),
                ('Difference', 'dfv', True)
            ]
        }
        
        self.in_basedir = ['samplesf', 'ppfracf', 'seqfile',
            'pptablef', 'lipnamesf', 'bindpropf', 'metabsf',
            'pfragmentsfile', 'nfragmentsfile', 'featurescache',
            'auxcache', 'stdcachefile', 'validscache']
        
        for attr, val in self.defaults.iteritems():
            if attr in kwargs:
                setattr(self, attr, kwargs[attr])
            else:
                setattr(self, attr, val)
        
        self.basedir = os.path.join(*([self.path_root] + self.basedir)) \
            if type(self.basedir) is list \
            else self.basedir
        
        for name in self.in_basedir:
            if hasattr(self, name):
                path = getattr(self, name)
                path = path if type(path) in charTypes else os.path.join(*path)
                setattr(self, name, os.path.join(self.basedir, path))
        
        self.data_basedir = self.basedir \
            if self.data_basedir is None else self.data_basedir
        
        self.ltpdirs = map(lambda p:
            os.path.join(*([self.data_basedir] + p)) if type(p) is list else p,
            self.ltpdirs
        )
        
        self.stddir = os.path.join(self.ltpdirs[0], self.stddir)
        
        self.paths_exist()
        
        if type(self.swl_levels) in charTypes:
            self.swl_levels = set([self.swl_levels])
        elif type(self.swl_levels) is list:
            self.swl_levels = set(self.swl_levels)
        
        self.ms1_tlr = self.ms1_tolerance
        self.ms2_tlr = self.ms2_tolerance
        self.std_tlr = self.std_tolerance
        
        self.nAdducts = None
        self.pAdducts = None
        self.exacts = None
        self.ltps_drifts = None
        
        self.html_table_template = '''<!DOCTYPE html>
            <html lang="en">
            <head>
                <meta charset="utf-8">
                <title>%s</title>
                <style type="text/css">
                    table {
                        border
                    }
                    th, .rowname {
                        font-weight: bold;
                    }
                    .positive {
                        background-color: #D04870;
                        color: #FFFFFF;
                    }
                    .negative {
                        background-color: #3A7AB3;
                        color: #FFFFFF;
                    }
                    .both {
                        background: linear-gradient(120deg, #D04870, #3A7AB3);
                        color: #FFFFFF;
                    }
                    .matching {
                        background-color: #03928C;
                        color: #FFFFFF;
                    }
                    .nothing {
                        background-color: #FFFFFF;
                        color: #000000;
                        font-weight: normal;
                    }
                    .clickable {
                        cursor: pointer;
                    }
                    td, th {
                        border: 1px solid #CCCCCC;
                    }
                </style>
                <script type="text/javascript">
                    function showTooltip(e) {
                        e = e || window.event;
                        var targ = e.target || e.srcElement;
                        if (targ.nodeType == 3) targ = targ.parentNode;
                        text = targ.getAttribute('title')
                        if (text !== undefined) alert(text);
                    }
                </script>
            </head>
            <body>
                <h1>%s</h1>
                <table border="0">
                    %s
                </table>
            </body>
            </html>'''
        
        self.colors = [
            '#B81466', # blue maguerite
            '#C441B3', # slate blue
            '#D6708B', # pale violet red
            '#69B3D6', # viking
            '#3A7AB3', # chetwode blue
            '#608784', # gothic
            '#03928C', # java
            '#CF5836', # mandy
            '#B3443D', # blush
            '#65B9B9', # seagull
            '#009BB9', # pelorous
            '#D88776', # my pink
            '#755987', # kimberly
            '#6F3940', # sanguine brown
            '#A09255', # teak
            '#7C997B', # avocado
            '#582E5E', # minsk
            '#224A7A', # fun blue
            '#FE2E08', # orange red
            '#2D16E2', # medium blue
            '#CBE216', # fuego
            '#14B866', # medium sea green
            '#987B99'  # london hue
        ]
    
    def paths_exist(self):
        paths = map(lambda name:
            getattr(self, name),
            self.in_basedir
        ) + self.ltpdirs
        for path in paths:
            if not os.path.exists(path) and path[-6:] != 'pickle':
                sys.stdout.write('\t:: Missing input file/path: %s\n' % path)
        sys.stdout.flush()
    
    def reload(self):
        modname = self.__class__.__module__
        mod = __import__(modname, fromlist = [modname.split('.')[0]])
        reload(mod)
        new = getattr(mod, self.__class__.__name__)
        setattr(self, '__class__', new)
    
    #
    # read standards data
    #

    def which_day(self, ltp, mode = 'pos'):
        '''
        Which day the samples for a given LTP have been run?
        '''
        if not hasattr(self, 'seq'):
            self.read_seq()
        return map(lambda (date, samples):
            date,
            filter(lambda (date, samples):
                (ltp, mode) in map(lambda sample:
                    (sample[0], sample[1]),
                    samples
                ),
                self.seq.iteritems()
            )
        )

    def standards_theoretic_masses(self):
        '''
        Reads the exact masses and most abundant ion masses of
        the lipid standards.
        '''
        result = {}
        tbl = self.read_xls(self.metabsf, 9)
        hdr = tbl[0]
        for pn in [('pos', 1, 6), ('neg', 8, 16)]:
            hdr = tbl[pn[1] - 1]
            result[pn[0]] = dict(
                map(lambda line:
                    (
                        line[0],
                        {
                            'exact': float(line[1]),
                            'ion': float(line[2]),
                            ''.join(filter(lambda c: c.isdigit(), hdr[3])):
                                self.to_float(line[3]),
                            ''.join(filter(lambda c: c.isdigit(), hdr[4])):
                                self.to_float(line[4]),
                            ''.join(filter(lambda c: c.isdigit(), hdr[5])):
                                self.to_float(line[5])
                        }
                    ),
                    tbl[pn[1]:pn[2]]
                )
            )
        for pn, dd in result.iteritems():
            for lstd, d in dd.iteritems():
                for k in d.keys():
                    if k.isdigit():
                        d['diff_%s'%k] = d['ion'] - d[k] \
                            if type(d[k]) is float else np.nan
        self.stdmasses = result

    #
    # obtaining lipid data from SwissLipids
    #

    def get_id(self, dct, value):
        if value not in dct:
            i = 0 if len(dct) == 0 else max(dct.values()) + 1
            dct[value] = i
        return dct[value]

    def get_lipidmaps(self):
        lipidmaps = []
        fields = ['lm_id', 'systematic_name', 'synonyms', 'category',
            'main_class', 'exact_mass', 'formula', 'inchi_key', 'pubchem_sid',
            'pubchem_cid', 'common_name']
        record = {}
        expect = False
        lmapsf = _curl.curl(self.lipidmaps_url, large = True,
            silent = False, files_needed = [self.lipidmaps_fname])
        for l in lmapsf.values()[0]:
            if expect:
                record[expect] = l.strip() \
                    if expect != 'exact_mass' else float(l.strip())
                expect = False
            elif l[0] == '>':
                label = l.strip()[3:-1].lower()
                if label in fields:
                    expect = label
            elif l[0] == '$':
                lipidmaps.append([record[label] if label in record else None \
                    for label in fields])
                record = {}
        return lipidmaps

    def lipidmaps_adducts(self):
        _nAdducts = []
        _pAdducts = []
        if not hasattr(self, 'lipidmaps'):
            self.get_lipidmaps()
        for l in self.lipidmaps:
            if l[5] is not None:
                _pAdducts.append([l[0], 'Species', \
                    '|'.join([str(l[10]), str(l[1]), str(l[2])]), l[6],
                    '[M+H]+', Mz(l[5]).add_h()])
                _pAdducts.append([l[0], 'Species', \
                    '|'.join([str(l[10]), str(l[1]), str(l[2])]), l[6],
                    '[M+NH4]+', Mz(l[5]).add_nh4()])
                _nAdducts.append([l[0], 'Species', \
                    '|'.join([str(l[10]), str(l[1]), str(l[2])]), l[6],
                    '[M-H]-', Mz(l[5]).remove_h()])
                _nAdducts.append([l[0], 'Species', \
                    '|'.join([str(l[10]), str(l[1]), str(l[2])]), l[6],
                    '[M+Fo]-', Mz(l[5]).add_fo()])
        _pAdducts = np.array(sorted(_pAdducts, key = lambda x: x[-1]),
            dtype = np.object)
        _nAdducts = np.array(sorted(_nAdducts, key = lambda x: x[-1]),
            dtype = np.object)
        self.pAdducts = _pAdducts \
            if self.pAdducts is None \
            else np.vstack((pAdducts, _pAdducts))
        self.pAdducts = self.pAdducts[self.pAdducts[:,-1].argsort()]
        self.nAdducts = _nAdducts \
            if self.nAdducts is None \
            else np.vstack((nAdducts, _nAdducts))
        self.nAdducts = self.nAdducts[self.nAdducts[:,-1].argsort()]

    def lipidmaps_exact(self):
        _exacts = []
        lipidmaps = self.get_lipidmaps()
        for l in lipidmaps:
            if l[5] is not None:
                _exacts.append([l[0], 'Species', \
                    '|'.join([str(l[10]), str(l[1]),
                        str(l[2])]), l[6], '', l[5]])
        _exactd = np.array(
            sorted(_exacts, key = lambda x: x[-1]),
            dtype = np.object)
        self.exacts = _exacts \
            if self.exacts is None \
            else np.vstack((self.exacts, _exacts))
        self.exacts = self.exacts[self.exacts[:,-1].argsort()]

    def get_swisslipids(self, adducts = None,
        exact_mass = False):
        '''
        Downloads the total SwissLipids lipids dataset.
        Returns numpy array with all queried adducts and optionally exact
        masses, calculates the formate adduct m/z's on demand.
        Please note, for a number of lipids some adduct m/z's are missing
        from SL, while the exact masses are given for all the lipids. Also
        storing masses of many adducts of the same lipid in an array results
        extensive memory usage (couple of hundreds MB).
        '''
        adduct_method = {
            '[M+OAc]-': 'add_ac',
            '[M+H]+': 'add_h',
            '[M-H]-': 'remove_h',
            '[M+Fo]-': 'add_fo',
            '[M+NH4]+': 'add_nh4'
        }
        if type(adducts) is list:
            adducts = set(adducts)
        readd = re.compile(r'.*(\[.*)')
        swl = _curl.curl(self.swisslipids_url, silent = False,
            compr = 'gz', large = True)
        swl.fileobj.seek(-4, 2)
        ucsize = struct.unpack('I', swl.fileobj.read(4))[0]
        swl.fileobj.seek(0)
        hdr = swl.readline().split('\t')
        positives = []
        negatives = []
        exact_masses = []
        prg = progress.Progress(ucsize, 'Processing SwissLipids', 101)
        for l in swl:
            prg.step(len(l))
            l = l.split('\t')
            if len(l) > 22:
                for i in xrange(13, 23):
                    if len(l[i]) > 0:
                        add = '' if i == 13 else readd.findall(hdr[i])[0]
                        if not adducts or add in adducts or i == 13 and \
                            exact_mass or formiate and i == 13:
                            mz = to_float(l[i])
                            if i == 13:
                                if formiate:
                                    mzfo = Mz(mz).add_fo()
                                    foline = [l[0], l[1], l[2],
                                        l[10], '[M+Fo]-', mzfo]
                                    negatives.append(foline)
                                if exact_mass:
                                    exact_masses.append([l[0], l[1], l[2],
                                        l[10], add, mz])
                            elif hdr[i].endswith('+'):
                                positives.append([l[0], l[1], l[2],
                                    l[10], add, mz])
                            elif hdr[i].endswith('-'):
                                negatives.append([l[0], l[1], l[2],
                                    l[10], add, mz])
        prg.terminate()
        swl.close()
        positives = sorted(positives, key = lambda x: x[-1])
        positives = np.array(positives, dtype = np.object)
        negatives = sorted(negatives, key = lambda x: x[-1])
        negatives = np.array(negatives, dtype = np.object)
        return positives, negatives

    def get_swisslipids_exact(self):
        '''
        Downloads the total SwissLipids lipids dataset.
        Returns numpy array with only exact mass values.
        '''
        readd = re.compile(r'.*(\[.*)')
        swl = _curl.curl(self.swisslipids_url, silent = False,
            compr = 'gz', large = True)
        swl.fileobj.seek(-4, 2)
        ucsize = struct.unpack('I', swl.fileobj.read(4))[0]
        swl.fileobj.seek(0)
        hdr = swl.readline().split('\t')
        _exacts = []
        prg = progress.Progress(ucsize, 'Processing SwissLipids', 101)
        for l in swl:
            prg.step(len(l))
            l = l.split('\t')
            if len(l) > 22:
                mz = self.to_float(l[14])
                add = ''
                _exacts.append([l[0], l[1], l[2], l[10], add, mz])
        prg.terminate()
        swl.close()
        _exacts = sorted(_exacts, key = lambda x: x[-1])
        _exacts = np.array(_exacts, dtype = np.object)
        self.exacts = _exacts \
            if self.exacts is None \
            else np.vstack((self.exacts, _exacts))
        self.exacts = self.exacts[self.exacts[:,-1].argsort()]

    def database_set(self, exacts, names, levels = ['Species']):
        levels = [levels] if type(levels) in [str, unicode] else levels
        levels = set([levels]) if type(levels) is list else levels
        names = [names] if type(names) in [str, unicode] else names
        idx = []
        for i, e in enumerate(exacts):
            if e[1] in levels:
                for n in names:
                    if n in e[2]:
                        idx.append(i)
                        break
        return exacts[idx,:]
    
    def process_db_keywords(self, kwdstr):
        return \
            map(
                lambda kwdset:
                    {
                        'neg':
                            map(
                                lambda kwd:
                                    kwd[1:],
                                filter(
                                    lambda kwd:
                                        len(kwd) and kwd[0] == '!',
                                    kwdset.split(';')
                                )
                            ),
                        'pos':
                            filter(
                                lambda kwd:
                                    len(kwd) and kwd[0] != '!',
                                kwdset.split(';')
                            )
                    },
                kwdstr.split('|')
            )
    
    def read_lipid_names(self):
        result = {}
        with open(self.lipnamesf, 'r') as f:
            nul = f.readline()
            for l in f:
                l = l.strip().split('\t')
                result[l[0]] = {
                    'full_name': l[1],
                    'swl': self.process_db_keywords(l[2]),
                    'lmp': self.process_db_keywords(l[3]),
                    'pos_adduct': l[4] if l[4] != 'ND' else None,
                    'neg_adduct': l[5] if l[5] != 'ND' else None
                }
        self.lipnames = result

    def read_binding_properties(self):
        result = {}
        with open(self.bindpropf, 'r') as f:
            data = map(lambda l:
                l.strip('\n').split('\t'),
                filter(lambda l:
                    len(l),
                    f
                )[1:]
            )
        for l in data:
            if l[2] not in result:
                result[l[2]] = set([])
            try:
                for lip in l[6].split(';'):
                    if lip != 'ND' and lip != '':
                        result[l[2]].add(lip)
            except IndexError:
                print l
        self.bindprop = result
    
    def binders_of(self, hg):
        return map(
            lambda (k, v):
                k,
            filter(
                lambda (k, v):
                    hg in v and k in self.samples_upper,
                self.bindprop.iteritems()
            )
        )

    #
    # reading the SEC absorption values and calculating the protein 
    # profiles in the fractions
    #
    
    def _GLTPD1_profile_correction_eq(self):
        eq = (self.pprofs['GLTPD1']['a11'] + self.pprofs['GLTPD1']['a12']) / 2.0
        self.pprofs['GLTPD1']['a11'] = eq
        self.pprofs['GLTPD1']['a12'] = eq
    
    def _GLTP1_profile_correction_inv(self):
        a11 = self.pprofs['GLTPD1']['a11']
        a12 = self.pprofs['GLTPD1']['a12']
        self.pprofs['GLTPD1']['a11'] = a12
        self.pprofs['GLTPD1']['a12'] = a11
    
    def protein_profiles(self, cache = True, correct_GLTPD1 = True, GLTPD_correction = 'eq'):
        '''
        For each protein, for each fraction, calculates the mean of 
        absorptions of all the measurements belonging to one fraction.
        '''
        if cache and os.path.exists(self.pprofcache):
            self.pprofs = pickle.load(open(self.pprofcache, 'rb'))
            if correct_GLTPD1:
                getattr(self, '_GLTPD1_profile_correction_%s' % GLTPD_correction)()
            return None
        reltp = re.compile(r'.*[\s_-]([A-Za-z0-9]{3,})\.xls')
        result = {}
        fnames = os.listdir(self.basedir)
        with open(self.ppfracf, 'r') as f:
            frac = [(self.to_float(i[0]), self.to_float(i[1]), i[2]) for i in \
                [l.split(';') for l in f.read().split('\n')]]
        for fname in fnames:
            ltpname = reltp.findall(fname)[0]
            frac_abs = dict((i[2], []) for i in frac)
            gfrac = (i for i in frac)
            fr = gfrac.next()
            try:
                tbl = self.read_xls(os.path.join(self.basedir, fname))[3:]
            except xlrd.biffh.XLRDError:
                sys.stdout.write('Error reading XLS:\n\t%s\n' % \
                    os.path.join(self.basedir, fname))
            minabs = min(0.0, min(self.to_float(l[5]) for l in tbl))
            for l in tbl:
                # l[4]: volume (ml)
                ml = to_float(l[4])
                if ml < fr[0]:
                    continue
                if ml >= fr[1]:
                    try:
                        fr = gfrac.next()
                    except StopIteration:
                        break
                # l[5] mAU UV3 215nm
                frac_abs[fr[2]].append(to_float(l[5]) - minabs)
            result[ltpname] = dict((fnum, np.mean(a)) \
                for fnum, a in frac_abs.iteritems())
        pickle.dump(result, open(self.pprofcache, 'wb'))
        self.pprofs = result
        if correct_GLTPD1:
            getattr(self, '_GLTPD1_profile_correction_%s' % GLTPD_correction)()

    def zero_controls(self):
        self.pprofs_original = copy.deepcopy(self.pprofs)
        fracs = ['a9', 'a10', 'a11', 'a12', 'b1']
        for ltpname, sample in self.samples.iteritems():
            for i, fr in enumerate(fracs):
                if sample[i + 1] == 0:
                    self.pprofs[ltpname.upper()][fr] = 0.0

    def one_sample(self):
        self.singles = [k.upper() for k, v in self.samples.iteritems() \
            if sum((i for i in v if i is not None)) == 1]

    def write_pptable(self):
        '''
        Writes protein profiles in a table, so we don't need to read
        all the 60 XLS files every time.
        '''
        with open(self.pptablef, 'w') as f:
            f.write('\t%s%s' % ('\t'.join(
                sorted(self.pprofs.values()[0].keys(),
                key = lambda x: (x[0], int(x[1:])))), '\n'))
            f.write('\n'.join('%s\t%s' % (
                    ltp,
                    '\t'.join('%.20f' % d[fr] for fr in sorted(d.keys(),
                        key = lambda x: (x[0], int(x[1:]))))) \
                for ltp, d in self.pprofs.iteritems()
            ))

    def read_pptable(self):
        '''
        Reads protein profiles from table.
        '''
        with open(self.pptablef, 'r') as f:
            header = f.readline().strip().split('\t')
            return dict((ll[0], dict(zip(header, self.float_lst(ll[1:])))) \
                for ll in (l.split('\t') for l in f.read().split('\n')))

    def read_xls(self, xls_file, sheet = 0, csv_file = None,
        return_table = True):
        '''
        Generic function to read MS Excel XLS file, and convert one sheet
        to CSV, or return as a list of lists
        '''
        try:
            book = xlrd.open_workbook(xls_file, on_demand = True)
            try:
                if type(sheet) is int:
                    sheet = book.sheet_by_index(sheet)
                else:
                    sheet = book.sheet_by_name(sheet)
            except xlrd.biffh.XLRDError:
                sheet = book.sheet_by_index(0)
            table = [[str(c.value) \
                for c in sheet.row(i)] \
                for i in xrange(sheet.nrows)]
        except IOError:
            sys.stdout.write('No such file: %s\n' % xls_file)
            sys.stdout.flush()
        except:
            book = openpyxl.load_workbook(filename = xls_file,
                read_only = True)
            try:
                if type(sheet) is int:
                    sheet = book.worksheets[sheet]
                else:
                    sheet = book.get_sheet_by_name(sheet)
            except:
                sheet = book.worksheets[0]
            cells = sheet.get_squared_range(1, 1,
                sheet.max_column, sheet.max_row)
            table = map(lambda row:
                map(lambda c:
                    str(c.value),
                    row
                ),
                cells
            )
        if csv_file:
            with open(csv_file, 'w') as csv:
                csv.write('\n'.join(['\t'.join(r) for r in table]))
        if not return_table:
            table = None
        book.release_resources()
        return table

    #
    # reading in a small table for keeping track which fractions are samples 
    # and which ones are controls
    #

    def read_samples(self):
        '''
        Reads from file sample/control annotations 
        for each proteins.
        '''
        data = {}
        with open(self.samplesf, 'r') as f:
            null = f.readline()
            for l in f:
                l = l.split(',')
                data[l[0].replace('"', '')] = \
                    np.array([self.to_int(x) \
                        if x != '' else None for x in l[1:]])
        self.samples = data

    def read_seq(self):
        refrac = re.compile(r'[AB][9120]{1,2}')
        result = {}
        with open(self.seqfile, 'r') as f:
            for l in f:
                l = l.replace('"', '').replace('2606', '0626')\
                    .replace('2627', '0627')
                if len(l) and l[:6] != 'Bracke' and l[:6] != 'Sample':
                    l = l.split(',')[1]
                    mode = 'neg' if 'neg' in l \
                        else 'pos' if 'pos' in l \
                        else None
                    seq = l.split(mode)[-1] if mode is not None else None
                    if seq is not None and len(seq) and seq[0] == '_':
                        seq = seq[1:]
                    l = l.split('_')
                    date = l[0]
                    if 'extracted' in l:
                        if 'SEC' in l:
                            protein = '#BUF'
                        elif 'Std1' in l:
                            protein = '#STD'
                    elif refrac.match(l[-1]) or refrac.match(l[-2]):
                        protein = l[4]
                        if protein == 'ORP9STARD15':
                            protein = protein[4:]
                    else:
                        protein = None
                    if date not in result:
                        result[date] = []
                    result[date].append((protein, mode, seq))
        del result['150330']
        del result['150331']
        self.seq = result
    
    def recalibration(self):
        self.standards_filenames()
        self.read_seq()
        self.standards_theoretic_masses()
        self.read_standards()
        self.drifts_by_standard()
        self.drifts_by_date()
        self.drifts2ltps()
        self.recalibrate()
    
    def standards_filenames(self):
        fnames = os.listdir(self.stddir)
        result = {}
        for fname in fnames:
            if 'Seq' not in fname:
                _fname = os.path.join(self.stddir, fname)
                lFname = fname[:-5].split('_')
                date = lFname[0]
                mode = 'pos' if 'pos' in lFname else 'neg'
                seq = fname.split(mode)[-1][:-5]
                if len(seq) and seq[0] == '_': seq = seq[1:]
                if date not in result:
                    result[date] = {}
                result[date][('#STD', mode, seq)] = _fname
        self.stdfiles = result

    def is_recalibrated(self):
        return np.any(['recalibrated' in tbl and tbl['recalibrated']\
            for d in self.valids.values() for tbl in d.values()])

    def recalibrate(self, missing = False):
        if self.is_recalibrated():
            if not missing:
                sys.stdout.write('\t:: Looks recalibration '\
                    'already has been done.\n'\
                    '\t   Set `recalibrated` '\
                    'values to `False` and call this method\n'\
                    '\t   again if you really want to repeat it.\n\n')
                sys.stdout.flush()
                return None
            else:
                sys.stdout.write('\t:: Recalibration already done.\n'\
                    '\t   Looking for missing LTP/mode cases, and '\
                    'doing only those.\n'\
                    '\t   If you want to recalibrate everything '\
                    'again,\n'\
                    '\t   reload the original data.\n\n')
                sys.stdout.flush()
        for ltp, d in self.valids.iteritems():
            for mode, tbl in d.iteritems():
                if 'recalibrated' not in tbl or \
                (not tbl['recalibrated'] and missing):
                    if ltp in self.ltps_drifts:
                        if mode in self.ltps_drifts[ltp]:
                            ppm = np.median(
                                self.ltps_drifts[ltp][mode].values())
                            ratio = self.ppm2ratio(ppm)
                            tbl['mz'] = tbl['mz'] * ratio
                            tbl['recalibrated'] = True
                        else:
                            tbl['recalibrated'] = False
                            sys.stdout.write('\t:: No drift value '\
                                'for %s in mode %s :(\n' % (ltp, mode))
                    else:
                        tbl['recalibrated'] = False
                        sys.stdout.write('\t:: No drift values'\
                            ' for %s-%s :(\n' % (ltp, mode))

    def read_standards(self, accuracy = 5, cache = True):
        if os.path.exists(self.stdcachefile) and cache:
            self.std_measured = pickle.load(open(self.stdcachefile, 'rb'))
            return None
        result = dict(map(lambda date:
            (date, {}),
            self.stdfiles.keys()
        ))
        for date, samples in self.stdfiles.iteritems():
            for sample, fname in samples.iteritems():
                if sample[0] == '#STD':
                    time, scans, centroids = self.read_mzml(fname)
                    peaks = self.find_peaks(scans, centroids, accuracy)
                    peaks = self.filter_peaks(peaks)
                    stdmeasured = self.standards_lookup(peaks, sample[1])
                    result[date][sample] = stdmeasured
        sys.stdout.write('\n\t:: Results saved to file %s\n' % \
            self.stdcachefile)
        sys.stdout.flush()
        pickle.dump(result, open(self.stdcachefile, 'wb'))
        self.std_measured = result

    def drifts_by_standard(self, write_table = 'drifts_ppm.tab'):
        drifts = dict(map(lambda (date, samples):
            (date, dict(map(lambda (sample, lipids):
                (sample, dict(map(lambda lipid:
                    (lipid, None),
                    lipids.keys()))
                ),
                samples.iteritems()))
            ),
            self.std_measured.iteritems())
        )
        hdr = ['date', 'run', 'mode', 'lipid', 'theoretical',
            'measured', 'ratio', 'ppm']
        tab = [hdr]
        for date in sorted(self.std_measured.keys()):
            samples = self.std_measured[date]
            for sample in sorted(samples.keys()):
                lipids = samples[sample]
                for lipid in sorted(lipids.keys()):
                    mmz = lipids[lipid]
                    mode = sample[1]
                    tmz = self.stdmasses[mode][lipid]['ion']
                    ratio = None if mmz is None else tmz / mmz
                    ppm = None if ratio is None else self.ratio2ppm(ratio)
                    drifts[date][sample][lipid] = ppm
                    tab.append([
                        '20%s-%s-%s' % (date[:2], date[2:4], date[4:6]),
                        sample[2],
                        'positive' if sample[1] == 'pos' else 'negative',
                        lipid,
                        '%.06f' % tmz,
                        '%.06f' % mmz if mmz is not None else 'n/a',
                        '%.09f' % ratio if ratio is not None else 'n/a',
                        '%.03f' % ppm if ppm is not None else 'n/a'
                    ])
        if write_table and type(write_table) in charTypes:
            tab = '\n'.join(map(lambda line:
                '\t'.join(line),
                tab
            ))
            with open(write_table, 'w') as f:
                f.write(tab)
        self.drifts = drifts

    def drifts_table(self, outfile = 'drifts_ppm.tab'):
        hdr = ['date', 'run', 'mode', 'lipid', 'ppm']
        tab = [hdr]
        for date, samples in self.drifts.iteritems():
            for sample, lipids in samples.iteritems():
                for lipid, ppm in lipids.iteritems():
                    tab.append([
                        '20%s-%s-%s' % (date[:2], date[2:4], date[4:6]),
                        sample[2],
                        'positive' if sample[1] == 'pos' else 'negative',
                        lipid,
                        '%.03f' % ppm if ppm is not None else 'n/a'
                    ])
        tab = '\n'.join(map(lambda line:
            '\t'.join(line),
            tab
        ))
        with open(outfile, 'w') as f:
            f.write(tab)

    def drifts_by_date(self):
        drifts2 = dict(map(lambda (date, samples):
            (date, dict(map(lambda sample:
                (sample, None),
                samples.keys()))
            ),
            self.drifts.iteritems())
        )
        for date, samples in self.drifts.iteritems():
            for sample, lipids in samples.iteritems():
                drifts2[date][sample] = \
                np.nanmedian(
                    self.remove_outliers(
                        np.array(
                            filter(lambda x:
                                x is not None,
                                lipids.values()
                            )
                        )
                    )
                )
                if np.isnan(drifts2[date][sample]):
                    sys.stdout.write('\t:: No values in %s %s\n' % \
                        (date, str(sample)))
        self.drifts2 = drifts2

    def drifts2ltps(self, write_table = 'LTPs_drifts.tab'):
        # before doing anything, fix some inconsistencies:
        stard10fractions = ['A09', 'A10', 'A11', 'A12']
        for i, fr in zip(
                map(lambda (i, sample):
                    i,
                    filter(lambda (i, sample):
                        sample[0] == 'STARD10',
                        enumerate(self.seq['150310'])
                    )
                ),
                stard10fractions * 2
            ):
            self.seq['150310'][i] = \
                (self.seq['150310'][i][0], self.seq['150310'][i][1], fr)
        #
        ltps_drifts = {}
        hdr = ['LTP', 'mode', 'fraction', 'ratio', 'ppm']
        tab = [hdr]
        def standard_indices(mode, se):
            return np.array(
                map(lambda (i, s):
                    i,
                    filter(lambda (i, s):
                        s[0] == '#STD' and s[1] == mode,
                        enumerate(se)
                    )
                )
            )
        def add_row(tab, ltp, mode, fr, d):
            tab.append([
                ltp,
                'positive' if mode == 'pos' else 'negative',
                fr,
                '%.09f' % self.ppm2ratio(d),
                '%.06f' % d
            ])
            return tab
        for date in sorted(self.seq.keys()):
            se = self.seq[date]
            stdi = {}
            lastbuf = {'pos': None, 'neg': None}
            prev = None
            stdi['neg'] = standard_indices('neg', se)
            stdi['pos'] = standard_indices('pos', se)
            for i, sample in enumerate(se):
                typ = sample[0]
                if typ is not None and typ != '#STD':
                    mode = sample[1]
                    ui = stdi[mode].searchsorted(i)
                    if ui == 0:
                        std1i = stdi[mode][0]
                    else:
                        std1i = stdi[mode][ui - 1]
                    if ui >= len(stdi[mode]):
                        std2i = stdi[mode][-1]
                    else:
                        std2i = stdi[mode][ui]
                    if np.isnan(self.drifts2[date][se[std1i]]):
                        std1i = std2i
                    if std1i == std2i or \
                        np.isnan(self.drifts2[date][se[std2i]]):
                        std2i = None
                    if std2i is None:
                        d = self.drifts2[date][se[std1i]]
                        sys.stdout.write('\t:: For %s-%s-%s only one standard'\
                            ' available\n' % (typ, mode, sample[2]))
                    elif std1i is not None and \
                        np.isnan(self.drifts2[date][se[std1i]]) and std2i is not None:
                        d = self.drifts2[date][se[std2i]]
                        sys.stdout.write('\t:: For %s-%s-%s only one standard'\
                            ' available\n' % (typ, mode, sample[2]))
                    elif std2i is not None and \
                        np.isnan(self.drifts2[date][se[std2i]]) and std1i is not None:
                        d = self.drifts2[date][se[std1i]]
                        sys.stdout.write('\t:: For %s-%s-%s only one standard'\
                            ' available\n' % (typ, mode, sample[2]))
                    else:
                        o1 = i - std1i
                        o2 = std2i - i
                        d = (self.drifts2[date][se[std1i]] * o2 + \
                            self.drifts2[date][se[std2i]] * o1) / (o1 + o2)
                    if typ == '#BUF':
                        lastbuf[mode] = d
                    else:
                        if typ not in ltps_drifts:
                            ltps_drifts[typ] = {}
                        if mode not in ltps_drifts[typ]:
                            ltps_drifts[typ][mode] = {}
                        if prev != typ:
                            ltps_drifts[typ][mode]['ctrl'] = lastbuf[mode]
                            tab = add_row(tab, typ, mode, 'CTL', lastbuf[mode])
                        ltps_drifts[typ][mode][sample[2]] = d
                        tab = add_row(tab, typ, mode, sample[2], d)
                    prev = typ
        if write_table and type(write_table) in charTypes:
            with open(write_table, 'w') as f:
                f.write('\n'.join(
                    map(lambda line:
                        '\t'.join(line),
                        tab)
                    )
                )
            sys.stdout.write('\n\t:: Table written to file %s\n\n'%write_table)
        self.ltps_drifts =  ltps_drifts

    def remove_outliers(self, data, m = 2):
        return data[np.where(abs(data - np.nanmean(data)) < \
            m * np.nanstd(data))]

    def ratio2ppm(self, ratio):
        return (1.0 - ratio) * 1.0e06

    def ppm2ratio(self, ppm):
        return 1.0 - ppm / 1.0e06

    def standards_lookup(self, peaks, mode):
        measured = {}
        for lipid, values in self.stdmasses[mode].iteritems():
            _mz = values['ion']
            ui = peaks[:,0].searchsorted(_mz)
            ud = 999.0 if ui >= peaks.shape[0] else peaks[ui, 0] - _mz
            ld = 999.0 if ui == 0 else _mz - peaks[ui - 1, 0]
            ci = ui if ud < ld else ui - 1
            mz = peaks[ci,0] if abs(peaks[ci,0] - _mz) <= self.std_tlr else None
            measured[lipid] = mz
        return measured

    def _process_binary_array(self, binaryarray, length = None):
        prefix = '{http://psi.hupo.org/ms/mzml}'
        cvparam = '%scvParam' % prefix
        _binary = '%sbinary' % prefix
        for par in binaryarray.findall(cvparam):
            if par.attrib['name'][-5:] == 'float':
                typ = par.attrib['name'][:2]
            if par.attrib['name'][-5:] == 'ssion':
                comp = par.attrib['name'][:4] == 'zlib'
            if par.attrib['name'][-5:] == 'array':
                name = par.attrib['name'].split()[0]
        binary = binaryarray.find(_binary)
        decoded = base64.decodestring(binary.text)
        if comp:
            decoded = zlib.decompress(decoded)
        # length is stated in the spectrum tag, but knowing the data type
        # we can also calculate:
        length = len(decoded) * 8 / int(typ) if length is None else length
        # float or double:
        _typ = 'd' if typ == '64' else 'f'
        # mzml binary data must be always little endian:
        dt = np.dtype('<%s'%_typ)
        arr = np.frombuffer(decoded, dtype = dt)
        # this does the same:
        # arr = struct.unpack('<%u%s' % (length, _typ), decoded)
        # arr = np.array(arr, dtype = np.float32)
        return name, arr

    def _centroid(self, raw):
        '''
        Perform a Gauss fit to centroid the peaks for the property
        Code from http://pymzml.github.io/_modules/spec.html
        '''
        peaks = []
        if 'm/z' in raw and 'intensity' in raw:
            intensity = raw['intensity']
            mz = raw['m/z']
            for i, ins in enumerate(intensity[:-1]):
                if i < 2:
                    continue
                if 0.0 < intensity[i - 1] < ins > intensity[i + 1] > 0.0:
                    x1 = mz[i - 1]
                    x2 = mz[i]
                    x3 = mz[i + 1]
                    y1 = intensity[i - 1]
                    y2 = intensity[i]
                    y3 = intensity[i + 1]
                    
                    if x2 - x1 > (x3 - x2) * 10 or (x2 - x1) * 10 < x3 - x2:
                        continue
                    if y3 == y1:
                        lower = 2
                        upper = 2
                        while (not 0 < y1 < y2 > y3 > 0) and \
                            y1 == y3 and upper < 10:
                            if i - lower < 0:
                                i_lower = 0
                            else:
                                i_lower = i - lower
                            if i + upper >= len(mz):
                                i_upper = len(mz) - 1
                            else:
                                i_upper = i + upper
                            x1 = mz[i_lower]
                            x3 = mz[i_upper]
                            y1 = intensity[i_lower]
                            y3 = intensity[i_upper]
                            if lower % 2 == 0:
                                upper += 1
                            else:
                                lower += 1
                    if not (0 < y1 < y2 > y3 > 0) or y1 == y3:
                        continue
                    try:
                        logratio = np.log(y2 / y1) / np.log(y3 / y1)
                        mu = (logratio * (x1**2 - x3**2) - x1**2 + x2**2) / \
                            (2 * (x2 - x1) - 2 * logratio* (x3 - x1))
                        csquared = (x2**2 - x1**2 - 2 * x2 * mu + \
                                2 * x1 * mu) / \
                            (2 * np.log(y1 / y2))
                        a = y1 * np.exp((x1 - mu)**2 / (2 * csquared))
                    except:
                        continue
                    peaks.append([mu, a])
        return np.array(peaks, dtype = np.float32)

    def find_peaks(self, scans, centroids, accuracy = 5, dope = False):
        '''
        Finds peaks detected in consecutive scans.
        Returns list of dicts with centroid m/z,
        cumulative intensity and RT range for
        each peak.
        Returns array with 
        m/z, rt_min, rt_max, area, scan count
        as coulumns.
        
        accuracy : int, float
            Instrument accuracy in ppm.
            The difference over we consider 2
            m/z values to be distinct.
        
        # values order: (rt, mz, int)
        '''
        accuracy = 1000000.0 / accuracy
        for scan, c in centroids.iteritems():
            centroids[scan]['centroids'] = \
                c['centroids'][c['centroids'][:,0].argsort(),:]
        peaks = []
        consecutive = {}
        def get_id():
            i = 0
            while True:
                yield i
                i += 1
        _peakid = get_id()
        prg = progress.Progress(len(scans), 'Processing scans', 1)
        for scan in scans:
            prg.step()
            if scan in centroids:
                _scan = centroids[scan]['centroids']
                rt = centroids[scan]['rt']
                added = set([])
                to_remove = set([])
                for peakid, cons in consecutive.iteritems():
                    ld = ud = 999.0
                    cons_mz = np.array(map(lambda x: x[1], cons))
                    cons_in = np.array(map(lambda x: x[2], cons))
                    cons_centroid_mz = np.sum(cons_mz * cons_in) / \
                        np.sum(cons_in)
                    ui = _scan[:,0].searchsorted(cons_centroid_mz)
                    if ui < _scan.shape[0]:
                        ud = _scan[ui, 0] - cons_centroid_mz
                    if ui > 0:
                        ld = cons_centroid_mz - _scan[ui - 1, 0]
                    ci = ui if ud < ld else ui - 1
                    if dope and _dope(cons_centroid_mz):
                        print '\n\t:: This mz looks DOPE: %f, closest in'\
                            ' this scan (%u): %f; accuracy: %f' % \
                            (cons_centroid_mz, scan, _scan[ci, 0], 
                                cons_centroid_mz / accuracy)
                    if abs(_scan[ci, 0] - cons_centroid_mz) <= \
                        cons_centroid_mz / accuracy:
                        # same m/z detected in another consecutive scan:
                        consecutive[peakid].append(
                            (rt, _scan[ci, 0], _scan[ci, 1]))
                        if dope and _dope(_scan[ci, 0]):
                            print '\n\t:: DOPE found in next scan, '\
                                'appended to consecutive series: %u, %f' % \
                                    (scan, _scan[ci,0])
                        # this index won't initiate a new consecutive series:
                        added.add(ci)
                    else:
                        # this consecutive series interrupted here,
                        # so move it to the peaks stack:
                        cons_rt = np.array(map(lambda x: x[0], cons))
                        peaks.append([
                            cons_centroid_mz,
                            min(cons_rt),
                            max(cons_rt),
                            cons_in.sum(),
                            len(cons)
                        ])
                        if dope and _dope(cons_centroid_mz):
                            print '\n\t:: DOPE series interrupted, moved to '\
                                'peaks stack: %f, %f' % \
                                (cons_centroid_mz, cons_in.sum())
                        to_remove.add(peakid)
                for peakid in to_remove & set(consecutive.keys()):
                    del consecutive[peakid]
                for i, (mz, ins) in enumerate(_scan):
                    if i not in added:
                        peakid = _peakid.next()
                        consecutive[peakid] = [(rt, mz, ins)]
                        if dope and _dope(mz):
                            print '\n\t:: New DOPE found, '\
                                'new consecutive series started'\
                                ': scan %u, m/z = %f, peakid = %u' % \
                                (scan, mz, peakid)
        prg.terminate()
        peaks = np.array(peaks, dtype = np.float32)
        peaks = peaks[peaks[:,0].argsort(),:]
        return peaks

    def _dope(self, mz):
        return mz < 742.5331 and mz > 742.530

    def filter_peaks(self, peaks, min_scans = 3, min_area = 10000):
        peaks = peaks[np.where(
            np.logical_and(
                peaks[:,4] >= min_scans,
                peaks[:,3] >= min_area
            )
        )]
        peaks = peaks[peaks[:, 0].argsort(),:]
        return peaks

    def read_mzml(self, fname):
        prefix = '{http://psi.hupo.org/ms/mzml}'
        run = '%srun' % prefix
        spectrum = '%sspectrum' % prefix
        cvparam = '%scvParam' % prefix
        binarraylist = '%sbinaryDataArrayList' % prefix
        binarray = '%sbinaryDataArray' % prefix
        scanlist = '%sscanList' % prefix
        scan = '%sscan' % prefix
        mslevel = 'ms level'
        basepeakmz = 'base peak m/z'
        basepeakin = 'base peak intensity'
        starttime = 'scan start time'
        scans = []
        centroids = {}
        time = datetime.datetime.utcfromtimestamp(0)
        with open(fname, 'r') as f:
            sys.stdout.write('\t:: Reading file `%s`\n'%fname)
            mzml = lxml.etree.iterparse(f, events = ('end',))
            ms1 = False
            try:
                for ev, elem in mzml:
                    if elem.tag == run:
                        time = elem.attrib['startTimeStamp']
                        time = datetime.datetime.strptime(time,
                            '%Y-%m-%dT%H:%M:%SZ')
                    if elem.tag == spectrum:
                        ms1 = False
                        length = elem.attrib['defaultArrayLength']
                        scanid = int(elem.attrib['id'].split('=')[-1])
                        for cvp in elem.findall(cvparam):
                            if cvp.attrib['name'] == mslevel:
                                level = cvp.attrib['value']
                                if level != '1':
                                    break
                                else:
                                    ms1 = True
                            if cvp.attrib['name'] == basepeakmz:
                                mz = float(cvp.attrib['value'])
                            if cvp.attrib['name'] == basepeakin:
                                intensity = float(cvp.attrib['value'])
                        if ms1:
                            _scan = elem.find(scanlist).find(scan)
                            for cvp in _scan.findall(cvparam):
                                if cvp.attrib['name'] == starttime:
                                    rt = float(cvp.attrib['value'])
                            raw = {}
                            for bda in elem.find(binarraylist).\
                                findall(binarray):
                                name, arr = self._process_binary_array(bda, length)
                                raw[name] = arr
                            _centroids = self._centroid(raw)
                            scans.append(scanid)
                            centroids[scanid] = \
                                {'rt': rt, 'centroids': _centroids}
            except lxml.etree.XMLSyntaxError as error:
                sys.stdout.write('\n\tWARNING: XML processing error: %s\n' % \
                    str(error))
                sys.stdout.flush()
        return time, scans, centroids

    def dope(self, c):
        '''
        Tells whether if a series of m/z's might be DOPE.
        '''
        return np.any(np.logical_and(c[:,0] > 742.52, c[:,0] < 742.54))

    #
    # scanning the directory tree and collecting the MS data csv files
    #

    def get_filenames(self):
        '''
        Files are placed in 2 root directories, in specific subdirectories,
        positive and negative MS mode in separate files.
        This function collects all the file paths and returns them 
        in a dict of dicts. Keys are LTP names, and 'pos'/'neg'.
        '''
        redirname = re.compile(r'(^[0-9a-zA-Z]+)[ _]')
        fnames = {}
        ltpdlsts = \
            map(lambda d:
                filter(lambda dd:
                    os.path.isdir(os.path.join(d, dd)),
                    os.listdir(d)
                ),
                self.ltpdirs
            )
        if sum(map(len, ltpdlsts)) == 0:
            sys.stdout.write('\t:: Please mount the shared folder!\n')
            return fnames
        for path, ltpdl in zip(self.ltpdirs, ltpdlsts):
            for ltpd in ltpdl:
                ltpname = redirname.findall(ltpd)
                pos = 'pos' if 'pos' in ltpd \
                    else 'neg' if 'neg' in ltpd else None
                if pos is not None and len(ltpname) > 0:
                    ltpname = ltpname[0]
                    fpath = [ltpd, 'features']
                    for f in os.listdir(os.path.join(path, *fpath)):
                        if 'LABELFREE' in f:
                            fpath.append(f)
                            break
                    for f in os.listdir(os.path.join(path, *fpath)):
                        if f.endswith('.csv'):
                            fpath.append(f)
                            break
                    if ltpname not in fnames:
                        fnames[ltpname] = {}
                    if pos not in fnames[ltpname] or ltpd.endswith('update'):
                        fnames[ltpname][pos] = os.path.join(path, *fpath)
        self.datafiles = fnames

    def read_file_np(self, fname, read_vars = ['Normalized Area']):
        '''
        Reads one MS file, returns numpy masked array, 
        with void mask.
        Column order:
        quality, m/z, rt-min, rt-max, charge, control, a9, a10, a11, a12, b1
        '''
        # typos in the headers what need to be fixed
        typos = {
            'Sample 5': '',
            'Sample 6': '',
            '_ ': '_'
        }
        rehdr = re.compile(r'([_0-9a-zA-Z]+)[_\s]([/0-9a-zA-Z\s]+)')
        refra = re.compile(r'.*_[ab]{1,2}([0-9]{,2}).*')
        retyp = re.compile(r'(' + '|'.join(typos.keys()) + r')')
        # order of samples (fractions)
        sname = [0, 9, 10, 11, 12, 1]
        # variable names
        vname = {
            'm/z': 0,
            'RT mean': 1,
            'Normalized Area': 2
        }
        data = []
        # col nums for each variable, for each fraction
        scols = dict([(var, dict([(i, None) for i in sname])) \
            for var in vname.keys()])
        with open(fname, 'r') as f:
            hdr = retyp.sub(lambda x:
                    typos[x.group()],
                f.readline()
            ).split(',')[1:]
            cols = [tuple([i] + [x.strip() \
                    for x in list(rehdr.match(h).groups(0))]) \
                for i, h in enumerate(hdr[6:-7])]
            # testing if order of columns is always 
            # m/z, RT, norm area; all files passed
            for c in cols:
                if c[0] % 3 != vname[c[2]]:
                    sys.stdout.write('erroneous column order: col %u '\
                        'with header %s\n\tin file %s\n' % (c[0], c[2], fname))
                fnum = 0 if 'ctrl' in c[1] or 'ctlr' in c[1] \
                        else int(refra.match(c[1]).groups(0)[0])
                # assigning columns to variable->fraction slots
                # col offsets from 6
                scols[c[2]][fnum] = c
            for l in f:
                l = [i.strip() for i in l.split(',')][1:]
                vals = [self.to_float(l[0]), self.to_float(l[2])] + \
                    [self.to_float(i.strip()) for i in l[3].split('-')] + \
                    [self.to_float(l[4])]
                for var, scol in scols.iteritems():
                    # number of columns depends on which variables we need
                    if var in read_vars:
                        for s in sname:
                            if scol[s] is None:
                                vals.append(None)
                            else:
                                # col offset is 6 !
                                vals.append(self.to_float(l[scol[s][0] + 6]))
                data.append(vals)
        data.sort(key = lambda x: x[1])
        return np.ma.masked_array(data, dtype = 'float64')

    def read_data(self):
        '''
        Iterates through dict of dicts with file paths, reads
        each file, and makes array view and masks to make it easy
        to handle missing data, and access m/z values of samples 
        and controls, and other values.
        '''
        data = dict((ltp, {}) for ltp in self.datafiles.keys())
        prg = progress.Progress(len(self.datafiles) * 2, 'Reading files', 1)
        for ltp, pos_neg in self.datafiles.iteritems():
            for p, fname in pos_neg.iteritems():
                prg.step()
                try:
                    data[ltp][p] = {}
                    # this returns a masked array:
                    data[ltp][p]['raw'] = self.read_file_np(fname)
                except:
                    sys.stdout.write('\nerror reading file: %s\n' % fname)
                    sys.stdout.flush()
                # making a view with the intensities:
                data[ltp][p]['int'] = data[ltp][p]['raw'][:, 5:]
                # mask non measured:
                data[ltp][p]['mes'] = data[ltp][p]['int'].view()
                data[ltp][p]['mes'].mask = \
                    np.array([[x is None for x in self.samples[ltp]] * \
                        ((data[ltp][p]['int'].shape[1] - 5) / 6)] * \
                    data[ltp][p]['int'].shape[0])
                data[ltp][p]['mes'] = np.ma.masked_invalid(data[ltp][p]['mes'])
                # controls:
                data[ltp][p]['ctr'] = data[ltp][p]['mes'][:,
                    np.array([x == 0 for x in self.samples[ltp]] * \
                        (data[ltp][p]['mes'].shape[1] / 6))]
                # samples with lipids:
                data[ltp][p]['lip'] = data[ltp][p]['mes'][:,
                    np.array([x == 1 for x in self.samples[ltp]] * \
                        (data[ltp][p]['mes'].shape[1] / 6))]
                # all samples except blank control:
                data[ltp][p]['smp'] = data[ltp][p]['mes'][:,
                    np.array([False] + [x is not None \
                        for x in self.samples[ltp][1:]] * \
                        (data[ltp][p]['mes'].shape[1] / 6))]
        prg.terminate()
        self.data = data

    #
    # Generic helper functions
    #

    def to_float(self, num):
        '''
        Extracts float from string, or returns None.
        '''
        renum = re.compile(r'([-]?[0-9]*[\.]?[0-9]+[eE]?[-]?[0-9]*)')
        num = renum.match(num.strip())
        if num:
            return float(num.groups(0)[0])
        else:
            return num

    def to_int(self, num):
        '''
        Extracts int from string or returns None.
        '''
        renum = re.compile(r'([-]?[0-9]+[\.]?[0-9]*)')
        num = renum.match(num.strip())
        if num:
            return int(num.groups(0)[0])
        else:
            return num

    def float_lst(self, l):
        '''
        Converts elements of a list to floats.
        '''
        return [self.to_float(x) for x in l]

    '''
    Filtering functions
    '''

    def quality_filter(self, threshold = 0.2):
        for ltp, d in self.data.iteritems():
            for pn, tbl in d.iteritems():
                tbl['qly'] = np.array(tbl['raw'][:,0] >= threshold)

    def charge_filter(self, charge = 1):
        for ltp, d in self.data.iteritems():
            for pn, tbl in d.iteritems():
                tbl['crg'] = np.array(tbl['raw'][:,4] == charge)

    def area_filter(self, area = 10000.0):
        for ltp, d in self.data.iteritems():
            for pn, tbl in d.iteritems():
                tbl['are'] = np.nanmax(tbl['lip'], 1) >= area

    def peaksize_filter(self, peakmin = 2.0, peakmax = 5.0, area = 10000):
        # minimum and maximum of all intensities over all proteins:
        mini = min(
            map(lambda tb:
                np.nanmin(np.nanmax(tb, 1)),
                (tbl['int'][np.nanmax(tbl['lip'], 1) >= area,:] \
                    for d in self.data.values() for tbl in d.values())
            )
        )
        maxi = max(
            map(
                np.nanmax,
                (tbl['int'] for d in self.data.values() for tbl in d.values())
            )
        )
        prg = progress.Progress(len(self.data) * 2, 'Peak size filter', 1)
        for ltp, d in self.data.iteritems():
            for pn, tbl in d.iteritems():
                prg.step()
                tbl['pks'] = (
                    # the peaksize:
                    np.nanmax(tbl['lip'], 1) / \
                        (np.nanmax(tbl['ctr'], 1) + 0.001)
                    # must be larger:
                    >
                    # than the min peaksize as a function of intensity:
                    (((np.nanmax(tbl['lip'], 1) - mini) / (maxi - mini)) *
                        (peakmax - peakmin) + peakmin)
                )
        prg.terminate()

    def cprofile_filter(self):
        profile_filter(prfx = 'c')

    def profile_filter(self, prfx = ''):
        frs = ['c0', 'a9', 'a10', 'a11', 'a12', 'b1']
        notf = []
        cols = 'smp' if prfx == 'c' else 'lip'
        prg = progress.Progress(len(data) * 2, 'Profile filter', 1,
            percent = False)
        for ltp, d in self.data.iteritems():
            for pn, tbl in d.iteritems():
                prg.step()
                if ltp.upper() not in self.pprofs:
                    notf.append(ltp)
                    continue
                # i != 0 : we always drop the blank control
                ppr = np.array([self.pprofs[ltp.upper()][frs[i]] \
                    for i, fr in enumerate(self.samples[ltp]) \
                        if fr == 1 and i != 0])
                ppr = self.norm_profile(ppr).astype(np.float64)
                prr = stats.rankdata(ppr)
                pranks = sorted((i for i, x in enumerate(prr)),
                    key = lambda x: x, reverse = True)
                if tbl[cols][0,:].count() > 1:
                    flatp = ppr[pranks[0]] - ppr[pranks[1]] < \
                        ppr[pranks[0]] * 0.1
                # 0 if have only one fraction in sample
                tbl['%sprf'%prfx] = np.apply_along_axis(
                    lambda x: self.diff_profiles(ppr, norm_profile(x)),
                        axis = 1, arr = tbl[cols])
                # all True if have only one fraction in sample
                tbl['%srpr'%prfx] = np.array([True] * tbl[cols].shape[0]) \
                    if tbl[cols].shape[1] == 1 else \
                    np.apply_along_axis(
                    lambda x: self.comp_profiles(ppr, 
                        self.norm_profile(x), prr, flatp),
                        axis = 1, arr = tbl[cols])
        prg.terminate()
        sys.stdout.write('No protein profiles found for %s\n\n' % ', '.join(notf))
        sys.stdout.flush()

    def diff_profiles(self, p1, p2):
        # profiles are numpy arrays
        # of equal length
        if len(p1) > 1:
            return np.nansum(np.abs(p1 - p2))
        else:
            return 1.0 * len(p2)

    def _comp_profiles(self, p1, p2):
        p1r = stats.rankdata(p1)
        pranks = sorted((i for i, x in enumerate(p1r)), 
            key = lambda x: x, reverse = True)
        flatp = p1r[pranks[0]] - p1r[pranks[1]] < \
            p1r[pranks[0]] * 0.1
        return self.comp_profiles(p1, p2, p1r, flatp), 0.0

    def comp_profiles(self, p1, p2, p1r = False, flatp = False):
        highest = False
        flat = False
        hollow = False
        p2r = stats.rankdata(p2)
        if type(p1r) is not np.ndarray:
            p1r = stats.rankdata(p1)
        highest = np.any(np.logical_and(p1r == 1, p2r == 1))
        p2ranks = sorted((i for i, x in enumerate(p2r)), 
            key = lambda x: x, reverse = True)
        if len(p1r) == 3:
            hollow = p1r[1] == 2 or p2r[1] == 2
        if not highest:
            flat = flatp and sum(p2) > 0.0 and \
                p2[p2ranks[0]] - p2[p2ranks[1]] < \
                    p2[p2ranks][0] * 0.1
        return not hollow and (highest or flat)

    def norm_profile(self, profile):
        return (profile - np.nanmin(profile.astype(np.float64))) / \
            (np.nanmax(profile) - np.nanmin(profile))

    def norm_profiles(self, tbl):
        if type(tbl) == np.ma.core.MaskedArray:
            tbl = tbl.data
        return (tbl - np.nanmin(tbl, axis = 1, keepdims = True)) / \
            (np.nanmax(tbl, axis = 1, keepdims = True) - \
                np.nanmin(tbl, axis = 1, keepdims = True))

    def ubiquity_filter_old(self, only_valid = True):
        prg = progress.Progress(len(data)**2, 'Ubiquity filter', 1)
        ltps = sorted(self.data.keys())
        for pn in ['pos', 'neg']:
            for i, ltp1 in enumerate(ltps):
                for ltp2 in ltps[i+1:]:
                    ltp1s = set([])
                    ltp2s = set([])
                    prg.step()
                    if pn in data[ltp1] and pn in data[ltp2]:
                        ltp1t = data[ltp1][pn]
                        ltp2t = data[ltp2][pn]
                        if 'ubi' not in ltp1t:
                            ltp1t['ubi'] = \
                                np.zeros([ltp1t['raw'].shape[0], 1], \
                                dtype = np.int8)
                        if 'ubi' not in ltp2t:
                            ltp2t['ubi'] = \
                                np.zeros([ltp2t['raw'].shape[0], 1], \
                                dtype = np.int8)
                        for i1, mz1 in np.ndenumerate(ltp1t['raw'][:,1]):
                            i2u = ltp2t['raw'][:,1].searchsorted(mz1)
                            if i2u < ltp2t['raw'].shape[0] and \
                                ltp2t['raw'][i2u,1] - mz1 <= self.ms1_tlr:
                                if i1 not in ltp1s:
                                    ltp1t['ubi'][i1] += 1
                                    ltp1s.add(i1)
                                if i2u not in ltp2s:
                                    ltp2t['ubi'][i2u] += 1
                                    ltp2s.add(i2u)
                            if mz1 - ltp2t['raw'][i2u - 1,1] <= self.ms1_tlr:
                                if i1 not in ltp1s:
                                    ltp1t['ubi'][i1] += 1
                                    ltp1s.add(i1)
                                if i2u - 1 not in ltp2s:
                                    ltp2t['ubi'][i2u - 1] += 1
                                    ltp2s.add(i2u - 1)
        prg.terminate()

    def ubiquity_filter(self, only_valid = True):
        prg = progress.Progress(len(self.valids)**2, 'Ubiquity filter', 1)
        ltps = sorted(self.valids.keys())
        for pn in ['pos', 'neg']:
            for i, ltp1 in enumerate(ltps):
                for ltp2 in ltps[i+1:]:
                    ltp1s = set([])
                    ltp2s = set([])
                    prg.step()
                    if pn in self.valids[ltp1] and pn in self.valids[ltp2]:
                        ltp1t = self.valids[ltp1][pn]
                        ltp2t = self.valids[ltp2][pn]
                        if 'ubi' not in ltp1t:
                            ltp1t['ubi'] = np.zeros(ltp1t['mz'].shape[0],
                                dtype = np.int8)
                        if 'ubi' not in ltp2t:
                            ltp2t['ubi'] = np.zeros(ltp2t['mz'].shape[0],
                                dtype = np.int8)
                        for i1, mz1 in np.ndenumerate(ltp1t['mz']):
                            if ltp1t['cpv'][i1] or not only_valid:
                                i2u = ltp2t['mz'].searchsorted(mz1)
                                u = 0
                                while True:
                                    if i2u + u < ltp2t['mz'].shape[0] and \
                                        ltp2t['mz'][i2u + u] - mz1 <= \
                                        self.ms1_tlr:
                                        if ltp2t['cpv'][i2u + u] or \
                                            not only_valid:
                                            if i1 not in ltp1s:
                                                ltp1t['ubi'][i1] += 1
                                                ltp1s.add(i1)
                                            if i2u + u not in ltp2s:
                                                ltp2t['ubi'][i2u + u] += 1
                                                ltp2s.add(i2u + u)
                                        u += 1
                                    else:
                                        break
                                l = 1
                                while True:
                                    if i2u - l >= 0 and \
                                        mz1 - ltp2t['mz'][i2u - l] <= \
                                        self.ms1_tlr:
                                        if ltp2t['cpv'][i2u - l] or \
                                            not only_valid:
                                            if i1 not in ltp1s:
                                                ltp1t['ubi'][i1] += 1
                                                ltp1s.add(i1)
                                            if i2u - l not in ltp2s:
                                                ltp2t['ubi'][i2u - l] += 1
                                                ltp2s.add(i2u - l)
                                        l += 1
                                    else:
                                        break
        prg.terminate()
    
    def remove_filter(self, attr_name):
        '''
        Deletes the result of a previously done filtering.
        '''
        for ltp, d in self.data.iteritems():
            for pn, tbl in d.iteritems():
                if attr_name in tbl:
                    del tbl[attr_name]
    
    def rprofile_filter(self):
        '''
        Dummy function for eval_filter()
        '''
        pass
    
    class get_hits(object):
        
        def __init__(self, fun):
            self.fun = fun
        
        def __call__(self, data, **kwargs):
            result = empty_dict(data)
            for ltp in data.keys():
                print ltp
                for pn, tbl in data[ltp].iteritems():
                    result[ltp][pn] = self.fun(tbl, **kwargs)
                    if result[ltp][pn] is None:
                        sys.stdout.write('No profile info for %s, %s, %s()\n' \
                            % (ltp, pn, self.fun.__name__))
            return result
    
    class combine_filters(object):
        
        def __init__(self, fun):
            self.fun = fun
        
        def __call__(self, **kwargs):
            for ltp in self.obj.data.keys():
                for pn, tbl in self.obj.data[ltp].iteritems():
                    try:
                        self.fun(self.obj, tbl, **kwargs)
                    except:
                        pass
        
        def __get__(self, instance, owner):
            self.obj = instance
            self.cls = owner
            
            return self.__call__
    
    class count_hits(object):
        
        def __init__(self, fun):
            self.fun = fun
        
        def __call__(self, **kwargs):
            hits = self.obj.empty_dict(self.obj.data)
            phits = self.obj.empty_dict(self.obj.data)
            for ltp in self.obj.data.keys():
                for pn, tbl in self.obj.data[ltp].iteritems():
                    try:
                        hits[ltp][pn] = self.fun(tbl, **kwargs)
                        phits[ltp][pn] = self.fun(tbl, **kwargs) / \
                            float(len(tbl[kwargs['name']])) * 100
                    except:
                        pass
            return hits, phits
        
        def __get__(self, instance, owner):
            self.obj = instance
            self.cls = owner
            
            return self.__call__
    
    @get_hits
    def val_ubi_prf_rpr_hits(self, tbl, ubiquity = 7, treshold = 0.15,
        tresholdB = 0.25, profile_best = False):
        '''
        Column order:
        quality, m/z, rt_min, rt_max, charge, control, a9, a10, a11, a12, b1,
        profile_score, control_profile_score, rank_profile_boolean,
        ubiquity_score, ubiquity_score,
        original_index
        '''
        prf = 'cprf' if tbl['lip'][1,:].count() == 1 else 'prf'
        _treshold = (treshold, tresholdB)
        if prf not in tbl:
            return None
        # dict of profile matching score values and their indices
        selected = np.logical_and(
            np.logical_and(
                np.logical_and(
                    np.logical_and(
                        np.logical_and(
                            np.logical_and(
                                tbl['qly'], tbl['crg']
                            ),
                            tbl['are']
                        ),
                        tbl['pks']
                    ),
                    tbl['rpr']
                ),
                tbl['ubi'] <= ubiquity
            ),
            np.logical_or(
                np.logical_not(
                    np.isnan(tbl['prf'])
                ),
                np.logical_not(
                    np.isnan(tbl['cprf'])
                )
            )
        )
        prf_values = dict(zip(np.where(selected)[0],
            zip(tbl['prf'][selected], tbl['cprf'][selected])))
        if profile_best:
            # select the best scoring records
            prf_values = sorted(prf_values.items(), key = lambda x: x[1])
            try:
                score_treshold = prf_values[min(profile_best - 1,
                    len(prf_values) - 1)][1]
            except IndexError:
                print prf_values
                return None
            # comments?
            indices = np.array(sorted(map(lambda x: x[0], \
                filter(lambda x: x[1] <= score_treshold, prf_values))))
        else:
            # or select the records with profile
            # match less then or equal to treshold
            indices = np.array(sorted((i for i, v in prf_values.iteritems() \
                if v <= _treshold)))
        return (tbl['raw'][indices,:], tbl['prf'][indices],
            tbl['cprf'][indices],
            tbl['rpr'][indices], tbl['ubi'][indices], tbl['uby'][indices] \
                if 'uby' in tbl else np.zeros(len(indices)), indices)

    @get_hits
    def pass_through(self, tbl, ubiquity = 7, treshold = 0.15,
        tresholdB = 0.25, profile_best = False):
        '''
        Column order:
        quality, m/z, rt_min, rt_max, charge, control, a9, a10, a11, a12, b1,
        profile_score, control_profile_score, rank_profile_boolean,
        ubiquity_score, ubiquity_score
        '''
        prf = 'cprf' if tbl['lip'][1,:].count() == 1 else 'prf'
        _treshold = (treshold, tresholdB)
        if prf not in tbl:
            return None
        # dict of profile matching score values and their indices
        selected = np.logical_or(
            np.logical_not(
                np.isnan(tbl['prf'])
            ),
            np.logical_not(
                np.isnan(tbl['cprf'])
            )
        )
        prf_values = dict(zip(np.where(selected)[0],
            zip(tbl['prf'][selected], tbl['cprf'][selected])))
        if profile_best:
            # select the best scoring records
            prf_values = sorted(prf_values.items(), key = lambda x: x[1])
            try:
                score_treshold = prf_values[min(profile_best - 1,
                    len(prf_values) - 1)][1]
            except IndexError:
                print prf_values
                return None
            # comments?
            indices = np.array(sorted(map(lambda x: x[0], \
                filter(lambda x: x[1] <= score_treshold, prf_values))))
        else:
            # or select the records with profile
            # match less then or equal to treshold
            indices = np.array(sorted((i for i, v in prf_values.iteritems() \
                if v <= _treshold)))
        print 'Selected: ', np.nansum(tbl['cprf'][indices])
        print 'Total: ', np.nansum(tbl['cprf'])
        return (tbl['raw'][indices,:], tbl['prf'][indices],
            tbl['cprf'][indices],
            tbl['rpr'][indices], tbl['ubi'][indices], tbl['uby'][indices] \
                if 'uby' in tbl else np.zeros(len(indices)))
    
    @combine_filters
    def validity_filter(self, tbl):
        tbl['vld'] = np.logical_and(
            np.logical_and(
                np.logical_and(
                    tbl['qly'],
                    tbl['crg']
                ),
                tbl['are']
            ),
            tbl['pks']
        )

    @combine_filters
    def val_ubi_filter(self, tbl, ubiquity = 7):
        tbl['vub'] = np.logical_and(
            np.logical_and(
                np.logical_and(
                    np.logical_and(
                        tbl['qly'],
                        tbl['crg']
                    ),
                    tbl['are']
                ),
                tbl['pks']
            ),
            tbl['ubi'][:,0] <= ubiquity
        )
    
    @combine_filters
    def val_prf_filter(self, tbl, treshold = 0.15):
        tbl['vpr'] = np.logical_and(
            np.logical_and(
                np.logical_and(
                    np.logical_and(
                        tbl['qly'],
                        tbl['crg']
                    ),
                    tbl['are']
                ),
                tbl['pks']
            ),
            tbl['prf'] <= treshold
        )
    
    @combine_filters
    def val_rpr_filter(self, tbl):
        tbl['vrp'] = np.logical_and(
            np.logical_and(
                np.logical_and(
                    np.logical_and(
                        tbl['qly'],
                        tbl['crg']
                    ),
                    tbl['are']
                ),
                tbl['pks']
            ),
            tbl['rpr']
        )
    
    @combine_filters
    def val_ubi_prf_filter(self, tbl, ubiquity = 7, treshold = 0.15):
        tbl['vup'] = np.logical_and(
            np.logical_and(
                np.logical_and(
                    np.logical_and(
                        np.logical_and(
                            tbl['qly'],
                            tbl['crg']
                        ),
                        tbl['are']
                    ),
                    tbl['pks']
                ),
                tbl['prf'] <= treshold
            ), tbl['ubi'][:,0] <= ubiquity
        )

    @combine_filters
    def val_ubi_prf_rprf_filter(self, tbl, ubiquity = 7, treshold = 0.15):
        tbl['vur'] = np.logical_and(
            np.logical_and(
                np.logical_and(
                    np.logical_and(
                        np.logical_and(
                            np.logical_and(
                                tbl['qly'],
                                tbl['crg']
                            ),
                            tbl['are']
                        ),
                        tbl['pks']
                    ),
                    tbl['rpr']
                ),
                tbl['prf'] <= treshold
            ),
            tbl['ubi'][:,0] <= ubiquity
        )


    def apply_filters(self,
        filters = ['quality', 'charge', 'area', 'peaksize'],
        param = {}):
        for f in filters:
            p = param[f] if f in param else {}
            fun = getattr(self, '%s_filter' % f)
            fun(**p)

    def empty_dict(self):
        return dict((ltp, {'pos': None, 'neg': None}) \
            for ltp in self.data.keys())

    def eval_filter(self, filtr, param = {}, runtime = True,
        repeat = 3, number = 10, hit = lambda x: x):
        t = None
        fun = globals()['%s_filter' % filtr]
        filters = {
                'quality': 'qly',
                'charge': 'crg',
                'peaksize': 'pks',
                'ubiquity': 'ubi',
                'area': 'are',
                'profile': 'prf',
                'rprofile': 'rpr',
                'cprofile': 'cprf',
                'validity': 'vld',
                'val_ubi': 'vub',
                'val_prf': 'vpr',
                'val_ubi_prf': 'vup',
                'val_ubi_prf_rprf': 'vur',
                'val_rpr': 'vrp'
            }
        name = filters[filtr]
        if runtime:
            t = min(timeit.Timer(lambda: fun(self.data, **param)).\
                repeat(repeat = repeat, number = number)) / float(number)
        @count_hits
        def counter(self, tbl, name = None):
            return np.nansum(hit(tbl[name]))
        hits, phits = counter(self.data, name = name)
        nums = np.array(list(j for i in  hits.values() for j in i.values()),
            dtype = np.float)
        pcts = np.array(list(j for i in  phits.values() for j in i.values()),
            dtype = np.float)
        return t, hits, phits, (np.nanmin(nums), np.nanmax(nums)), \
            (np.nanmin(pcts), np.nanmax(pcts))

    def combined_hits(self, profile = 0.15,
        ubiquity = 20, verbose = True):
        @count_hits
        def counter(self, tbl):
            return np.logical_and(
                np.logical_and(
                    np.logical_and(
                        np.logical_and(
                            np.logical_and(
                                tbl['qly'],
                                tbl['crg']
                            ),
                            tbl['are']
                        ),
                        tbl['pks']
                    ),
                    tbl['prf'] <= profile
                ),
                tbl['ubi'][:,0] <= ubiquity
            ).sum()
        hits = counter(self.data)
        if verbose:
            sys.stdout.write('\n\tLTP\t\t+\t-\n\t%s\n'%('='*30))
            for ltp in sorted(hits.keys()):
                sys.stdout.write('\t%s\t'%ltp)
                for pn in ['pos', 'neg']:
                    num = hits[ltp][pn]
                    sys.stdout.write('\t%s' % \
                        (str(num) if num is not None else 'n/a'))
                sys.stdout.write('\n')
        return hits

    '''
    Save and reload functions
    '''

    def save_data(self, fname = None):
        fname = os.path.join(self.basedir, self.featurescache) \
            if fname is None else fname
        sys.stdout.write('\t:: Saving data to %s ...\n' % fname)
        sys.stdout.flush()
        pickle.dump(self.data, open(fname, 'wb'))
        sys.stdout.write('\t:: Data has been saved to %s.\n' % fname)
        sys.stdout.flush()

    def load_data(self, fname = None):
        fname = os.path.join(self.basedir, self.featurescache) \
            if fname is None else fname
        sys.stdout.write('\t:: Loading data from data to %s ...\n' % fname)
        sys.stdout.flush()
        self.data = pickle.load(open(fname, 'rb'))

    def save(self, fname = 'save.pickle'):
        fname = os.path.join(self.basedir, self.auxcache) \
            if fname is None else fname
        sys.stdout.write('\t:: Saving auxiliary data to %s ...\n' % fname)
        sys.stdout.flush()
        pickle.dump((self.datafiles, self.samples, self.pprofs),
            open(fname, 'wb'))
        sys.stdout.write('\t:: Data has been saved to %s.\n' % fname)
        sys.stdout.flush()

    def load(self, fname = None):
        fname = os.path.join(self.basedir, self.auxcache) \
            if fname is None else fname
        sys.stdout.write('\t:: Loading auxiliary data '\
            'from %s ...\n' % fname)
        sys.stdout.flush()
        self.datafiles, self.samples, self.pprofs = \
            pickle.load(open(fname, 'rb'))

    '''
    END: save & reload
    '''

    '''
    Lipid databases lookup functions
    '''

    def find_lipids(self, hits, lipnames):
        '''
        Column order:
        
        in:
        [0] quality, m/z, rt_min, rt_max, charge, control, a9, a10, a11, a12, b1
        [1] profile_score
        [2] control_profile_score
        [3] rank_profile_boolean
        [4] ubiquity_score
        [5] ubiquity_score
        
        out:
        ltp_name, m/z, 
        profile_score, control_profile_score,
        rank_profile_boolean, ubiquity_score,
        ubiquity_score, swisslipids_ac, level,
        lipid_name, lipid_formula, adduct, 
        adduct_m/z
        '''
        # levels: 'Structural subspecies', 'Isomeric subspecies',
        # 'Species', 'Molecular subspecies'
        lipids = dict((ltp.upper(), {}) for ltp in hits.keys())
        for ltp, d in hits.iteritems():
            for pn, tbl in d.iteritems():
                adducts = self.pAdducts if pn == 'pos' else self.nAdducts
                result = []
                if tbl[0] is not None:
                    for i in xrange(tbl[0].shape[0]):
                        swlipids = self.adduct_lookup(tbl[0][i,1], adducts)
                        if swlipids is not None:
                            for lip in swlipids:
                                hg, p_add, n_add = \
                                    self.headgroup_from_lipid_name(lip)
                                fa = self.fattyacid_from_lipid_name(lip)
                                result.append(np.concatenate(
                                    # tbl[1] and tbl[2] are the profile 
                                    # and cprofile scores
                                    (np.array([ltp.upper(), tbl[0][i,1],
                                        tbl[1][i], tbl[2][i],
                                        tbl[3][i], tbl[4][i],
                                        tbl[5][i]],
                                        dtype = np.object),
                                    lip,
                                    np.array(hg, fa, dtype = np.object)),
                                    axis = 0))
                    lipids[ltp.upper()][pn] = np.vstack(
                        sorted(result, key = lambda x: x[1]))
        self.lipids = lipids

    def headgroup_from_lipid_name(self, lip):
        '''
        For one database record attempts to identify the lipid class
        by looking up keywords.
        '''
        db = 'lmp' if lip[0][0] == 'L' else 'swl'
        for shortname, spec in self.lipnames.iteritems():
            for kwset in spec[db]:
                matched = [kw in lip[2] for kw in kwset['pos']]
                if sum(matched) == len(kwset['pos']) and \
                    sum(matched) > 0:
                    matched = [kw in lip[2] for kw in kwset['neg']]
                    if sum(matched) == 0:
                        return shortname, spec['pos_adduct'], spec['neg_adduct']
        return None, None, None

    def fattyacid_from_lipid_name(self, lip, _sum = True):
        refa = re.compile(r'([dl]?)([0-9]+:[0-9])\(?([,0-9EZ]+)?\)?')
        _fa = None
        if lip[0][0] == 'L':
            names = lip[2].split('|')
            for i in [0, 2, 1]:
                fa = refa.findall(names[i])
                if len(fa) > 0:
                    _fa = fa
                    break
        elif lip[0][0] == 'S':
            fa = refa.findall(lip[2])
            if len(fa) > 0:
                _fa = fa
        if _fa is not None and _sum:
            _fa = map(sum, zip(*[[int(i) for i in f[1].split(':')] \
                for f in _fa]))
        return _fa

    def find_lipids_exact(self, verbose = False,
        outfile = None, ltps = None, charge = 1):
        '''
        Looks up lipids by m/z among database entries in
        `exacts`, and stores the result in dict under key
        `lip`, where keys are the original indices (`i`).
        '''
        
        if verbose:
            outfile = sys.stdout if outfile is None else open(outfile, 'w')
        for ltp, d in self.valids.iteritems():
            if ltps is None or ltp in ltps:
                for pn, tbl in d.iteritems():
                    tbl['lip'] = {}
                    for i in xrange(tbl['mz'].shape[0]):
                        tbl['lip'][tbl['i'][i]] = self.adduct_lookup_exact(
                            tbl['mz'][i], pn,
                            verbose, outfile, charge = 1
                        )
        if type(outfile) is file and outfile != sys.stdout:
            outfile.close()

    def adduct_lookup(self, mz, adducts):
        '''
        Looks up m/z values in the table containing the reference database
        already converted to a pool of all possible adducts.
        (Does not convert the m/z to other adducts.)
        '''
        result = []
        iu = adducts[:,-1].searchsorted(mz)
        if adducts.shape[0] > iu:
            u = 0
            while True:
                if adducts[iu + u,-1] - mz <= self.ms1_tlr:
                    if adducts[iu + u,1] in self.swl_levels:
                        result.append(adducts[iu + u,:])
                    u += 1
                else:
                    break
        if iu > 0:
            l = 1
            while True:
                if iu - l >= 0 and mz - adducts[iu - l,-1] <= self.ms1_tlr:
                    if adducts[iu - l,1] in self.swl_levels:
                        result.append(adducts[iu - l,:])
                    l += 1
                else:
                    break
        return None if len(result) == 0 else np.vstack(result)

    def adduct_lookup_exact(self, mz, mode, verbose = False,
        outfile = None, charge = 1):
        '''
        Looks up m/z values in the table containing the reference database
        casting the m/z to specific adducts.
        '''
        if verbose and outfile is None:
            outfile = sys.stdout
        result = []
        adducts = self.aadducts[charge][mode]
        for addName, addFun in adducts.iteritems():
            addMz = getattr(Mz(mz), addFun)()
            if verbose:
                outfile.write('\t:: Searching for %s adducts.\n\t '\
                    '-- Adduct mass (measured): '\
                    '%.08f, exact mass (calculated): %.08f.\n' % \
                    (addName, mz, addMz))
            iu = self.exacts[:,-1].searchsorted(addMz)
            if self.exacts.shape[0] > iu:
                u = 0
                while True:
                    if self.exacts[iu + u,-1] - addMz <= self.ms1_tlr:
                        if self.exacts[iu + u,1] in self.swl_levels:
                            lip = self.exacts[iu + u,:]
                            hg, p_add, n_add = \
                                self.headgroup_from_lipid_name(lip)
                            fa = self.fattyacid_from_lipid_name(lip)
                            lip = np.concatenate((lip,
                                np.array([addMz, hg, fa], dtype = np.object)),
                                axis = 0)
                            lip[4] = addName
                            if verbose:
                                outfile.write('\t -- Found: %s\n' % \
                                    str(list(lip)))
                            result.append(lip)
                        u += 1
                    else:
                        break
            if iu > 0:
                l = 1
                while True:
                    if iu - l >= 0 and addMz - self.exacts[iu - l,-1] <= \
                        self.ms1_tlr:
                        if self.exacts[iu - l,1] in self.swl_levels:
                            lip = self.exacts[iu - l,:]
                            # headgroup and fatty acid guess from lipid name
                            # happens here!
                            hg, p_add, n_add = \
                                self.headgroup_from_lipid_name(lip)
                            fa = self.fattyacid_from_lipid_name(lip)
                            lip = np.concatenate((lip, 
                                np.array([addMz, hg, fa], dtype = np.object)),
                                axis = 0)
                            lip[4] = addName
                            if verbose:
                                outfile.write('\t -- Found: %s\n' % \
                                    str(list(lip)))
                            result.append(lip)
                        l += 1
                    else:
                        break
        return None if len(result) == 0 else np.vstack(result)

    def negative_positive(self, add_col = 12, mz_col = 1, swl_col = 8):
        '''
        Column order:
        
        in:
        ltp_name, m/z,
        profile_score, control_profile_score,
        rank_profile_boolean, ubiquity_score, ubiquity_score,
        original_index,
        swisslipids_ac, level, lipid_name, lipid_formula, adduct, adduct_m/z
        
        out:
        [0] pos_m/z, pos_profile_score,
        [2] pos_control_profile_score, pos_rank_profile_boolean,
        [4] pos_ubiquity_score, pos_ubiquity_score,
        [6] pos_original_index, pos_swisslipids_ac, pos_level,
        [9] pos_lipid_name, pos_lipid_formula, pos_adduct, pos_adduct_m/z
        [13] neg_m/z, neg_profile_score, neg_control_profile_score,
        [16] neg_rank_profile_boolean,
        [17] neg_ubiquity_score, neg_ubiquity_score, neg_original_index,
        [18] neg_swisslipids_ac, neg_level,
        [22] neg_lipid_name, neg_lipid_formula, neg_adduct, neg_adduct_m/z
        '''
        result = dict((ltp.upper(), []) for ltp in self.lipids.keys())
        prg = progress.Progress(len(result), 'Matching positive & negative',
            1, percent = False)
        for ltp, tbl in self.lipids.iteritems():
            prg.step()
            if 'neg' in tbl and 'pos' in tbl:
                for neg in tbl['neg']:
                    adds = [neg[add_col]] \
                        if neg[add_col] != 'Unknown' else ['[M-H]-', '[M+Fo]-']
                    for add in adds:
                        if add == '[M-H]-':
                            poshmz = Mz(Mz(neg[mz_col]).add_h()).add_h()
                            posnh3mz = Mz(Mz(neg[mz_col]).add_h()).add_nh4()
                        elif add == '[M+Fo]-':
                            poshmz = Mz(Mz(neg[mz_col]).remove_fo()).add_h()
                            posnh3mz = Mz(Mz(neg[mz_col]).\
                                remove_fo()).add_nh4()
                        else:
                            continue
                        iu = tbl['pos'][:,mz_col].searchsorted(poshmz)
                        u = 0
                        if iu < len(tbl['pos']):
                            while iu + u < len(tbl['pos']):
                                if tbl['pos'][iu + u,mz_col] - poshmz <= \
                                    self.ms1_tlr:
                                    if tbl['pos'][iu + u,swl_col] \
                                            == neg[swl_col] and \
                                        tbl['pos'][iu + u,add_col] \
                                            == '[M+H]+' or \
                                        tbl['pos'][iu + u,add_col] \
                                            == 'Unknown':
                                        result[ltp].append(np.concatenate(
                                            (tbl['pos'][iu + u,1:], neg[1:]),
                                                axis = 0
                                        ))
                                    u += 1
                                else:
                                    break
                        if iu > 0:
                            l = 1
                            while iu >= l:
                                if poshmz - tbl['pos'][iu - l,mz_col] <= \
                                    self.ms1_tlr:
                                    if tbl['pos'][iu - l,swl_col] \
                                            == neg[swl_col] and \
                                        tbl['pos'][iu - l,add_col] \
                                            == '[M+H]+' or \
                                        tbl['pos'][iu - l,add_col] \
                                            == 'Unknown':
                                        result[ltp].append(np.concatenate(
                                            (tbl['pos'][iu - l,1:], neg[1:]),
                                            axis = 0
                                        ))
                                    l += 1
                                else:
                                    break
                        iu = tbl['pos'][:,1].searchsorted(posnh3mz)
                        u = 0
                        if iu < len(tbl['pos']):
                            while iu + u < len(tbl['pos']):
                                if tbl['pos'][iu + u,mz_col] - \
                                    posnh3mz <= self.ms1_tlr:
                                    if tbl['pos'][iu + u,swl_col] \
                                            == neg[swl_col] and \
                                        tbl['pos'][iu + u,add_col] \
                                            == '[M+NH4]+' or \
                                        tbl['pos'][iu + u,add_col] \
                                            == 'Unknown':
                                        result[ltp].append(np.concatenate(
                                            (tbl['pos'][iu + u,1:], neg[1:]),
                                            axis = 0
                                        ))
                                    u += 1
                                else:
                                    break
                        if iu > 0:
                            l = 1
                            while iu >= l:
                                if posnh3mz - tbl['pos'][iu - l,mz_col] \
                                    <= self.ms1_tlr:
                                    if tbl['pos'][iu - l,swl_col] \
                                            == neg[swl_col] and \
                                        tbl['pos'][iu - l,add_col] \
                                            == '[M+NH4]+' or \
                                        tbl['pos'][iu - l,add_col] \
                                            == 'Unknown':
                                        result[ltp].append(np.concatenate(
                                            (tbl['pos'][iu - l,1:], neg[1:]),
                                            axis = 0
                                        ))
                                    l += 1
                                else:
                                    break
                if len(result[ltp]) > 0:
                    result[ltp] = np.vstack(result[ltp])
        prg.terminate()
        return result

    def negative_positive2(self):
        '''
        Results in dicts ['pos']['neg'] and ['neg']['pos'].
        Values in each array: assumed positive adduct, assumed negative adduct, 
            measured positive m/z, measured negative m/z
        '''
        adds = self.aadducts[1]
        self.sort_alll('mz')
        for ltp, tbl in self.valids.iteritems():
            tbl['pos']['neg'] = dict((i, {}) for i in tbl['pos']['i'])
            tbl['pos']['neg_lip'] = dict((i, {}) for i in tbl['pos']['i'])
            tbl['neg']['pos'] = dict((i, {}) for i in tbl['neg']['i'])
            tbl['neg']['pos_lip'] = dict((i, {}) for i in tbl['neg']['i'])
        prg = progress.Progress(len(self.valids),
            'Matching positive & negative',
            1, percent = False)
        for ltp, tbl in self.valids.iteritems():
            prg.step()
            for pi, poi in enumerate(tbl['pos']['i']):
                measured_pos_mz = tbl['pos']['mz'][pi]
                for pos_add, pos_add_fun in adds['pos'].iteritems():
                    calculated_exact = \
                        getattr(Mz(measured_pos_mz), pos_add_fun)()
                    for neg_add, neg_add_fun in adds['neg'].iteritems():
                        calculated_neg_mz = getattr(Mz(calculated_exact),
                            neg_add_fun)()
                        iu = tbl['neg']['mz'].searchsorted(calculated_neg_mz)
                        u = 0
                        if iu < len(tbl['neg']['i']):
                            while iu + u < len(tbl['neg']['i']):
                                if tbl['neg']['mz'][iu + u] - \
                                    calculated_neg_mz <= self.ms1_tlr:
                                    noi = tbl['neg']['i'][iu + u]
                                    match = np.array([pos_add, neg_add,
                                        measured_pos_mz,
                                        tbl['neg']['mz'][iu + u]],
                                        dtype = np.object)
                                    tbl['pos']['neg'][poi][noi] = match
                                    tbl['neg']['pos'][noi][poi] = match
                                    self.negative_positive_lipids(tbl, poi,
                                        tbl['neg']['i'][iu + u],
                                        pos_add, neg_add)
                                    u += 1
                                else:
                                    break
                        if iu > 0:
                            l = 1
                            while iu >= l:
                                if calculated_neg_mz - \
                                    tbl['neg']['mz'][iu - l] <= self.ms1_tlr:
                                    noi = tbl['neg']['i'][iu - l]
                                    match = np.array([pos_add, neg_add,
                                        measured_pos_mz,
                                        tbl['neg']['mz'][iu - l]],
                                        dtype = np.object)
                                    tbl['pos']['neg'][poi][noi] = match
                                    tbl['neg']['pos'][noi][poi] = match
                                    self.negative_positive_lipids(tbl, poi,
                                        tbl['neg']['i'][iu - l],
                                        pos_add, neg_add)
                                    l += 1
                                else:
                                    break
        prg.terminate()

    def negative_positive_lipids(self, tbl, poi, noi, pos_add, neg_add):
        '''
        Result columns:
        database id positive, database id negative,
        lipid name positive, lipid name negative,
        adduct positive, adduct negative,
        database m/z positive, database m/z negative,
        measured m/z positive, measured m/z negative,
        headgroup positive, headgroup negative,
        fatty acids positive, fatty acids negative,
        dominant adduct positive, dominant adduct negative
        '''
        result = []
        if tbl['pos']['lip'][poi] is not None and \
            tbl['neg']['lip'][noi] is not None:
            for plip in tbl['pos']['lip'][poi]:
                if plip[4] == pos_add:
                    for nlip in tbl['neg']['lip'][noi]:
                        if plip[7] is not None and nlip[7] is not None and \
                            nlip[4] == neg_add and plip[7] == nlip[7] and \
                            plip[8] == nlip[8]:
                                result.append(np.array(
                                [plip[0], nlip[0], plip[2], nlip[2], plip[4],
                                nlip[4], plip[5], nlip[5], plip[6], nlip[6],
                                plip[7], nlip[7], plip[8], nlip[8],
                                self.lipnames[plip[7]]['pos_adduct'],
                                self.lipnames[nlip[7]]['neg_adduct']],
                                dtype = np.object))
        result = np.vstack(result) if len(result) > 0 else None
        tbl['pos']['neg_lip'][poi][noi] = result
        tbl['neg']['pos_lip'][noi][poi] = result
    
    def ms1(self):
        self.lipid_lookup_exact()
        self.ms1_headgroups()
        self.negative_positive2()
        self.headgroups_negative_positive('ms1')
    
    '''
    END: lipid databases lookup
    '''

    '''
    Functions for MS2
    '''
    def ms2(self):
        self.ms2_metabolites()
        self.ms2_filenames()
        self.ms2_map()
        self.ms2_main()
        self.ms2_headgroups()
        self.headgroups_by_fattya()
        self.identity_combined()
    
    def identify(self):
        self.headgroups_by_fattya()
        self.headgroups_negative_positive('ms2')
        self.identity_combined()
        self.identity_combined_ms2()
        self.feature_identity_table()
    
    def ms2_metabolites(self):
        self.pFragments, self.pHgfrags, self.pHeadgroups = \
            self.ms2_read_metabolites(self.pfragmentsfile)
        self.nFragments, self.nHgfrags, self.nHeadgroups = \
            self.ms2_read_metabolites(self.nfragmentsfile)
    
    def export_auto_fraglist(self, infile,
        outfile = 'lipid_fragments_negative_mode_ext.txt'):
        lst = self.ms2_read_metabolites(infile, extra_fragments = True,
            return_fraglines = True)
        with open(outfile, 'w') as out:
            out.write(
                '\n'.join(
                    map(
                        lambda l:
                            '%.07f\t%s\t%s%s' %\
                                tuple(l[:3] +
                                    ['\t%s'%l[3] if len(l) >= 4 else '']),
                        lst
                    )
                )
            )
    
    def ms2_read_metabolites(self, fname, extra_fragments = True, return_fraglines = False):
        '''
        In part from Toby Hodges.
        Reads metabolite fragments data.
        '''
        Metabolites = []
        Hgroupfrags = {}
        rehg = re.compile(r'.*\(([\+;A-Z]+)\).*')
        rehgsep = re.compile(r'[;\+]')
        with open(fname, 'r') as Handle:
            lst = \
            map(
                lambda l:
                    map(
                        lambda (i, v):
                            v if i > 0 else self.to_float(v),
                        enumerate(
                            l.strip().split('\t')
                        )
                    ),
                Handle.readlines()
            )
        if extra_fragments:
            if 'negative' in fname:
                # fatty acid -CO2- fragments:
                lst += self.auto_fragment_list(
                    FattyFragment, -1, minus = ['H', 'CO2'], name = 'FA')
                lst += self.auto_fragment_list(FAAlkenylminusH, -1)
                for hg in ('PE', 'PC', 'PS', 'PI'):
                    lst += self.auto_fragment_list(
                        globals()['Lyso%s' % hg], -1
                    )
                    lst += self.auto_fragment_list(
                        globals()['Lyso%s' % hg], -1, minus = ['H2O']
                    )
                    lst += self.auto_fragment_list(
                        globals()['Lyso%sAlkyl' % hg], -1
                    )
                    lst += self.auto_fragment_list(
                        globals()['Lyso%sAlkyl' % hg], -1, minus = ['H2O']
                    )
                    if hg != 'PI':
                        lst += self.auto_fragment_list(
                            globals()['Lyso%sAlkenyl' % hg], -1, cmin = 4
                        )
                        lst += self.auto_fragment_list(
                            globals()['Lyso%sAlkenyl' % hg], -1, cmin = 4, minus = ['H2O']
                        )
                    if hg == 'PI':
                        lst += self.auto_fragment_list(
                            LysoPI, -1, minus = ['H', 'H2O', 'C6H10O5']
                        )
                        lst += self.auto_fragment_list(
                            LysoPI, -1, minus = ['CO2']
                        )
                    if hg == 'PE':
                        lst += self.auto_fragment_list(
                            LysoPE, -1, minus = ['CO2']
                        )
                    if hg == 'PC':
                        lst += self.auto_fragment_list(
                            LysoPC, -1, minus = ['CH3']
                        )
                        lst += self.auto_fragment_list(
                            LysoPC, -1, minus = ['CH3', 'H2O']
                        )
                lst += self.auto_fragment_list(LysoPG, -1)
                lst += self.auto_fragment_list(LysoPG, -1, minus = ['H2O'])
                lst += self.auto_fragment_list(LysoPA, -1, minus = ['H2O'])
            if 'positive' in fname:
                lst += self.auto_fragment_list(NLFAminusH2O, 0)
                lst += self.auto_fragment_list(FattyFragment, 0, name = 'FA')
                lst += self.auto_fragment_list(FAplusGlycerol, 1)
                lst += self.auto_fragment_list(SphingosineBase, 1, cmin = 14, unsatmin = 0, cmax = 19, unsatmax = 3, minus = ['H2O'])
                lst += self.auto_fragment_list(SphingosineBase, 1, cmin = 14, unsatmin = 0, cmax = 19, unsatmax = 3, minus = ['H2O', 'H2O'])
                lst += self.auto_fragment_list(SphingosineBase, 1, cmin = 14, unsatmin = 0, cmax = 19, unsatmax = 3, minus = ['C', 'H2O', 'H2O'])
                lst += self.auto_fragment_list(SphingosineBase, 1, cmin = 14, unsatmin = 0, cmax = 19, unsatmax = 3, minus = ['H5'])
            if return_fraglines:
                return lst
            
        for Line in lst:
            try:
                MetabMass, MetabType, MetabCharge = Line[:3]
            except:
                print Line
            hgm = rehg.match(Line[1])
            if len(Line) == 4 or hgm:
                Hgroupfrags[MetabType] = set([])
                if len(Line) == 4:
                    Hgroupfrags[MetabType] = Hgroupfrags[MetabType] | \
                        set(rehgsep.split(Line[3]))
                if hgm:
                    Hgroupfrags[MetabType] = Hgroupfrags[MetabType] | \
                        set(rehgsep.split(hgm.groups()[0]))
            Metabolites.append([MetabMass, MetabType, MetabCharge])
            if '+' not in MetabCharge and '-' not in MetabCharge \
                and 'NL' not in MetabCharge:
                sys.stdout.write('WARNING: fragment %s has no '\
                    'valid charge information!\n' % \
                    MetabType)
        Headgroups = {}
        for frag, hgs in Hgroupfrags.iteritems():
            for hg in hgs:
                if len(hg):
                    if hg not in Headgroups:
                        Headgroups[hg] = set([])
                    Headgroups[hg].add(frag)
        return np.array(sorted(Metabolites, key = lambda x: x[0]),
            dtype = np.object), Hgroupfrags, Headgroups
    
    def auto_fragment_list(self, typ, charge, cmin = 2, unsatmin = 0,
        cmax = 36, unsatmax = 6, minus = [], plus = [], **kwargs):
        series = FAFragSeries(typ, charge, cmin = cmin, unsatmin = unsatmin,
            cmax = cmax, unsatmax = unsatmax, minus = minus, plus = plus,
            **kwargs)
        return(series.iterfraglines())

    def ms2_filenames(self):
        '''
        Files are placed in 2 root directories, in specific subdirectories,
        positive and negative MS mode in separate files.
        This function collects all the file paths and returns them 
        in a dict of dicts. Keys are LTP names, and 'pos'/'neg'.
        '''
        redirname = re.compile(r'(^[0-9a-zA-Z]+)[ _](pos|neg)')
        refractio = re.compile(r'.*_([AB][0-9]{1,2}).*')
        fnames = {}
        for d in self.ltpdirs:
            ltpdd = os.listdir(d)
            ltpdd = [i for i in ltpdd if os.path.isdir(os.path.join(d, i))]
            if len(ltpdd) == 0:
                sys.stdout.write('\t:: Please mount the shared folder!\n')
                return fnames
            for ltpd in ltpdd:
                try:
                    ltpname, pos = redirname.findall(ltpd)[0]
                    ltpname = ltpname.upper()
                    if ltpname not in fnames:
                        fnames[ltpname] = {}
                    if pos not in fnames[ltpname] or ltpd.endswith('update'):
                        fnames[ltpname][pos] = {}
                except:
                    # other dirs are irrelevant:
                    continue
                fpath = [ltpd, 'Results']
                if os.path.isdir(os.path.join(d, *fpath)):
                    for f in os.listdir(os.path.join(d, *fpath)):
                        if f.endswith('mgf') and 'buffer' not in f:
                            try:
                                fr = refractio.match(f).groups()[0]
                            except AttributeError:
                                sys.stdout.write(
                                    'could not determine '\
                                    'fraction number: %s' % f)
                            fr = 'B01' if fr == 'B1' \
                                else 'A09' if fr == 'A9' \
                                else fr
                            fnames[ltpname][pos][fr] = \
                                os.path.join(*([d] + fpath + [f]))
        self.ms2files = fnames

    def ms2_map(self, charge = 1):
        stRpos = 'pos'
        stRneg = 'neg'
        redgt = re.compile(r'[AB]([\d]+)$')
        result = dict((ltp.upper(), {'pos': None, 'neg': None, 
                'ms2files': {'pos': {}, 'neg': {}}}) \
            for ltp, d in self.ms2files.iteritems())
        prg = progress.Progress(len(self.ms2files) * 2, 'Indexing MS2 data', 1,
            percent = False)
        for ltp, d in self.ms2files.iteritems():
            pFeatures = []
            nFeatures = []
            ultp = ltp.upper()
            for pn, files in d.iteritems():
                prg.step()
                features = pFeatures if pn == stRpos else nFeatures
                #fractions = set([])
                for fr, fl in files.iteritems():
                    fr = int(redgt.match(fr).groups()[0])
                    #fractions.add(fr)
                    mm = self.ms2_index(fl, fr, charge = charge)
                    m = np.vstack(mm)
                    features.extend(mm)
                    result[ultp]['ms2files'][pn][int(fr)] = fl
            pFeatures = np.array(sorted(pFeatures, key = lambda x: x[0]),
                dtype = np.float64)
            nFeatures = np.array(sorted(nFeatures, key = lambda x: x[0]),
                dtype = np.float64)
            result[ultp]['pos'] = pFeatures
            result[ultp]['neg'] = nFeatures
        prg.terminate()
        self.ms2map = result

    def ms2_index(self, fl, fr, charge = 1):
        '''
        Looking up offsets in one MS2 mgf file.
        
        Columns:
            -- pepmass
            -- intensity
            -- retention time
            -- scan num
            -- offset in file
            -- fraction num
        '''
        stRrtinseconds = 'RTINSECONDS'
        stRtitle = 'TITLE'
        stRbe = 'BE'
        stRen = 'EN'
        stRch = 'CH'
        stRpepmass = 'PEPMASS'
        stRempty = ''
        reln = re.compile(r'^([A-Z]+).*=([\d\.]+)[\s]?([\d\.]*)["]?$')
        features = []
        offset = 0
        with open(fl, 'rb', 8192) as f:
            for l in f:
                if not l[0].isdigit() and \
                    not l[:2] == stRbe and not l[:2] == stRen:
                    if not l[:2] == stRch:
                        try:
                            m = reln.match(l).groups()
                        except:
                            print fl, l
                            continue
                        if m[0] == stRtitle:
                            scan = float(m[1])
                        if m[0] == stRrtinseconds:
                            rtime = float(m[1]) / 60.0
                        if m[0] == stRpepmass:
                            pepmass = float(m[1])
                            intensity = 0.0 if m[2] == stRempty else float(m[2])
                    if l[:2] == stRch:
                        _charge = int(l[7]) if len(l) >= 8 else None
                        if _charge == charge:
                            features.append([pepmass, intensity,
                                rtime, scan, offset + len(l), fr])
                            scan = None
                            rtime = None
                            intensity = None
                            pepmass = None
                            _charge = None
                offset += len(l)
        return features

    def ms2_main(self, verbose = False):
        '''
        For all LTPs and modes obtains the MS2 data from original files.
        '''
        # ms2map columns: pepmass, intensity, rtime, scan, offset, fraction
        prg = progress.Progress(len(self.valids) * 2,
            'Looking up MS2 fragments', 1,
            percent = False)
        for ltp, d in self.valids.iteritems():
            for pn, tbl in d.iteritems():
                prg.step()
                # we look up the real measured MS1 m/z's in MS2,
                # so will divide recalibrated values by the drift ratio
                drift = 1.0 \
                    if not hasattr(self, 'ltp_drifts') \
                    or 'recalibrated' not in tbl \
                    or not tbl['recalibrated'] \
                    else self.ppm2ratio(np.nanmedian(
                            np.array(self.ltp_drifts[ltp][pn].values())
                        ))
                ms2matches = self.ms2_match(tbl['mz'],
                    tbl['rt'], tbl['i'],
                    ltp, pn, drift = drift)
                # this already returns final result, from protein containing
                # fractions and with the relevant retention times
                tbl['ms2'] = self.ms2_lookup(ltp, pn, ms2matches)
                tbl['ms2r'] = self.ms2_result(tbl['ms2'])
                if verbose:
                    print '\n'
                    print 'number of positive mzs:', len(posMzs)
                    print 'negative matching:', len(pos_matches)
                    print 'number of negative mzs:', len(negMzs)
                    print 'negative matching:', len(neg_matches)
        prg.terminate()

    def ms2_result(self, ms2matches):
        '''
        Extracts the most relevant information from the MS2
        result arrays, throwing away the rest.
        
        Columns in output arrays (6):
            # MS2 fragment name, MS2 adduct name, (0-1)
            # MS2 fragment m/z, MS2 fragment intensity (2-3)
            # fraction number, scan number (4-5)
        '''
        result = dict((oi, []) for oi in ms2matches.keys())
        for oi, ms2s in ms2matches.iteritems():
            for ms2i in ms2s:
                result[oi].append(np.array([ms2i[7], ms2i[8],
                    ms2i[1], ms2i[2], ms2i[14], ms2i[12]], dtype = np.object))
        for oi, ms2s in result.iteritems():
            result[oi] = np.vstack(ms2s) if len(ms2s) > 0 else np.array([])
            result[oi].shape = (len(result[oi]), 6)
        return result

    def ms2_match(self, ms1Mzs, ms1Rts, ms1is, ltp, pos,
        verbose = False, outfile = None, drift = 1.0):
        '''
        Looks up matching pepmasses for a list of MS1 m/z's.
        '''
        outfile = sys.stdout if outfile is None else open(outfile, 'w')
        matches = []
        ms2tbl = self.ms2map[ltp][pos]
        # iterating over MS1 m/z, MS1 original index, and retention time
        for ms1Mz, ms1i, rt in zip(ms1Mzs, ms1is, ms1Rts):
            # drift is the ratio
            # we divide here to have the original measured MS1 m/z,
            # not the recalibrated one:
            ms1Mz = ms1Mz / drift
            # if error comes here, probably MS2 files are missing
            try:
                iu = ms2tbl[:,0].searchsorted(ms1Mz)
            except IndexError:
                sys.stdout.write('\nMissing MS2 files for %s-%s?\n' % \
                    (ltp, pos))
            if verbose:
                outfile.write('\t:: Looking up MS1 m/z %.08f. '\
                    'Closest values found: %.08f and %.08f\n' % (
                    ms1Mz,
                    ms2tbl[iu - 1, 0] if iu > 0 else 0.0,
                    ms2tbl[iu, 0] if iu < ms2tbl.shape[0] else 0.0))
            u = 0
            if iu < ms2tbl.shape[0]:
                while iu + u < ms2tbl.shape[0]:
                    if ms2tbl[iu + u,0] - ms1Mz <= self.ms1_tlr:
                        if verbose:
                            outfile.write('\t -- Next value within '\
                                'range of tolerance: %.08f\n' % \
                                    ms2tbl[iu + u, 0])
                            if ms2tbl[iu + u, 2] >= rt[0] \
                                and ms2tbl[iu + u, 2] <= rt[1]:
                                outfile.write('\t -- Retention time OK, '\
                                    'accept this match\n')
                            else:
                                outfile.write('\t -- Retention time is %.02f'\
                                    ', should be in range %.02f-%.02f to'\
                                    ' match.\n' % \
                                    (ms2tbl[iu + u, 2], rt[0], rt[1]))
                                outfile.write('\t -- Retention time not OK, '\
                                    'drop this match\n')
                        # checking retention time
                        if ms2tbl[iu + u, 2] >= rt[0] and \
                            ms2tbl[iu + u, 2] <= rt[1]:
                            matches.append((ms1Mz, iu + u, ms1i))
                        u += 1
                    else:
                        break
            l = 1
            if iu > 0:
                while iu >= l:
                    if ms1Mz - ms2tbl[iu - l,0] <= self.ms1_tlr:
                        # checking retention time
                        if verbose:
                            outfile.write('\t -- Next value within '\
                                'range of tolerance: %.08f\n' % \
                                    ms2tbl[iu - l, 0])
                            if ms2tbl[iu - l, 2] >= rt[0] \
                                and ms2tbl[iu - l, 2] <= rt[1]:
                                outfile.write('\t -- Retention time OK, '\
                                    'accept this match\n')
                            else:
                                outfile.write('\t -- Retention time is %.02f'\
                                    ', should be in range %.02f-%.02f to'\
                                    ' match.\n' % \
                                    (ms2tbl[iu - l, 2], rt[0], rt[1]))
                                outfile.write('\t -- Retention time not OK, '\
                                    'drop this match\n')
                        if ms2tbl[iu - l, 2] >= rt[0] and \
                            ms2tbl[iu - l, 2] <= rt[1]:
                            matches.append((ms1Mz, iu - l, ms1i))
                        l += 1
                    else:
                        break
        if type(outfile) is file and outfile != sys.stdout:
            outfile.close()
        return sorted(uniqList(matches), key = lambda x: x[0])

    def ms2_verbose(self, ltp, mode, outfile = None):
        tbl = valids[ltp][mode]
        ms2_match(tbl['mz'], tbl['rt'], tbl['i'],
                ltp, mode, verbose = True,
                outfile = outfile)

    def ms2_lookup(self, ltp, mode, ms1matches):
        '''
        For the matching MS2 m/z's given, reads and identifies
        the list of fragments.
        
        Columns in output arrays (15):
            # MS1 m/z, MS2 fragment m/z, MS2 fragment intensity, (0-2)
            # MS2 table index, direct/inverted match, fraction number, (3-5)
            # MS2 fragment m/z in annotation, MS2 fragment name, (6-7)
            # MS2 adduct name, (8)
            # MS1 pepmass, MS1 intensity, rtime, MS2 scan, (9-12)
            # MS2 file offset, fraction number (13-14)
        '''
        # indices of fraction numbers
        fragments = self.pFragments if mode == 'pos' else self.nFragments
        ms2map = self.ms2map[ltp][mode]
        ms2files = self.ms2map[ltp]['ms2files'][mode]
        samples = self.samples_upper[ltp]
        sample_i = {
            9: 1, 10: 2, 11: 3, 12: 4, 1: 5
        }
        stRcharge = 'CHARGE'
        # opening all files to have file pointers ready
        files = dict((fr, open(fname, 'r')) \
            for fr, fname in ms2files.iteritems())
        # Initializing dict with original indices
        ms2matches = dict((m[2], []) for m in ms1matches)
        # iterating over MS1 m/z, MS2 table index, MS1 original index
        for ms1mz, ms2i, ms1oi in ms1matches:
            # ms2map columns: pepmass, intensity, rtime, scan, offset, fraction
            ms2item = ms2map[ms2i,:]
            fr = int(ms2item[5])
            # only samples with the LTP
            if samples[sample_i[fr]] == 1 and fr in files:
                f = files[fr]
                # jumping to offset
                f.seek(int(ms2item[4]), 0)
                # zero means no clue about charge
                charge = 0
                for l in f:
                    if l[:6] == stRcharge:
                        # one chance to obtain the charge
                        charge = int(l.strip()[-2])
                        continue
                    if not l[0].isdigit():
                        # finish at next section
                        break
                    else:
                        # reading fragment masses
                        mi = l.strip().split()
                        try:
                            mass = float(mi[0])
                        except ValueError:
                            print '\n:::\n'
                            print l[0].isdigit()
                            print prevp
                            print prevl
                            print f.tell()
                            print l
                            f.seek(f.tell(), 0)
                            print f.read(10)
                            f.seek(prevp, 0)
                            print f.read(10)
                            print ':::\n'
                        intensity = float(mi[1]) if len(mi) > 1 else np.nan
                        # matching fragment --- direct
                        ms2hit1 = self.ms2_identify(mass, fragments,
                            compl = False)
                        # matching fragment --- inverted
                        ms2hit2 = self.ms2_identify(ms2item[0] - mass,
                            fragments, compl = True)
                        # matched fragment --- direct
                        # columns (14):
                        # MS1 m/z, MS2 fragment m/z, MS2 fragment intensity,
                        # MS2 table index, direct/inverted match, 
                        # fraction number,
                        # MS2 fragment m/z in annotation, MS2 fragment name,
                        # MS2 adduct name,
                        # MS1 pepmass, MS1 intensity, rtime, MS2 scan,
                        # MS2 file offset, fraction number
                        for frag in ms2hit1:
                            ms2matches[ms1oi].append(np.concatenate(
                                (np.array([ms1mz, mass,
                                    intensity, ms2i, 0, fr]),
                                frag, ms2item)
                            ))
                        # matched fragment --- inverted
                        for frag in ms2hit2:
                            ms2matches[ms1oi].append(np.concatenate((
                                np.array([ms1mz, mass,
                                    intensity, ms2i, 1, fr]),
                                frag, ms2item)
                            ))
                        # no fragment matched --- unknown fragment
                        if not len(ms2hit1) and not len(ms2hit2):
                            thisRow = np.concatenate((
                                np.array([ms1mz, mass,
                                    intensity, ms2i, 0, fr],
                                    dtype = np.object),
                                np.array([None, 'unknown', 'unknown'],
                                    dtype = np.object),
                                ms2item)
                            )
                            thisRow.shape = (1, 15)
                            ms2matches[ms1oi].append(thisRow)
                    prevl = l
                    prevp = f.tell()
        # removing file pointers
        for f in files.values():
            f.close()
        for oi, ms2match in ms2matches.iteritems():
            if len(ms2match) > 0:
                ms2matches[oi] = np.vstack(ms2match)
            else:
                ms2matches[oi] = np.array([[]])
                ms2matches[oi].shape = (0, 15)
        return ms2matches

    def ms2_identify(self, mass, fragments, compl):
        '''
        Looks up one MS2 m/z value in list of known fragments masses.
        Either with matching between MS2 m/z and fragment m/z within
        a given tolerance, or between the fragment mass and the
        residual mass after substracting MS2 m/z from MS1 m/z.
        Returns the fragment's mass, name and adduct type, or None
        in case of no match.
        '''
        result = []
        i = -1
        du = None
        dl = None
        iu = fragments[:,0].searchsorted(mass)
        if iu < len(fragments) and \
            (compl and 'NL' in fragments[iu,2] or \
            not compl and ('+' in fragments[iu,2] or \
                '-' in fragments[iu,2])):
            du = fragments[iu,0] - mass
        if iu > 0 and \
            (compl and 'NL' in fragments[iu - 1,2] or \
            not compl and ('+' in fragments[iu - 1,2] or \
                '-' in fragments[iu - 1,2])):
            dl = mass - fragments[iu - 1,0]
        if du is not None and (du < dl or dl is None) and du < self.ms2_tlr:
            i = iu
            st = 1
        if dl is not None and (dl < du or du is None) and dl < self.ms2_tlr:
            i = iu - 1
            st = -1
        val = fragments[i,0]
        while len(fragments) > i >= 0 and fragments[i,0] == val:
            result.append(fragments[i,:])
            i += st
        return result

    def ms2_collect(self, ms2matches, ms1mz, unknown = False):
        '''
        Deprecated.
        '''
        result = []
        fragments = ms2matches[ms2matches[:,0] == ms1mz,:]
        for frag in uniqList(fragments[:,7]):
            if frag != 'Unknown':
                thisFrag = fragments[fragments[:,7] == frag,:]
                maxInt = thisFrag[:,2].max()
                thisFragMass = thisFrag[0,1]
                result.append((frag, maxInt, thisFragMass))
        if unknown:
            unknowns = fragments[fragments[:,7] == 'Unknown',:]
            result += [('Unknown', l[2], l[1]) for l in unknowns]
        return result

    def ms2_headgroups(self):
        '''
        Creates dictionaries named ms2hg having the
        original IDs as keys and the sets of the
        identified possible headgroups as values.
        '''
        for ltp, d in self.valids.iteritems():
            for pn, tbl in d.iteritems():
                tbl['ms2hg'] = {}
                tbl['ms2fa'] = {}
                tbl['ms2fai'] = {}
                tbl['ms2fas'] = {}
                hgfrags = self.pHgfrags \
                    if pn == 'pos' else self.nHgfrags
                headgroups = self.pHeadgroups \
                    if pn == 'pos' else self.nHeadgroups
                result = {}
                for oi, ms2r in tbl['ms2r'].iteritems():
                    hgroups = self.ms2_headgroup2(ms2r, hgfrags, headgroups)
                    ms2fa, ms2fai, ms2fas = self.ms2_fattya(ms2r)
                    tbl['ms2hg'][oi] = hgroups
                    tbl['ms2fa'][oi] = ms2fa
                    tbl['ms2fai'][oi] = ms2fai
                    tbl['ms2fas'][oi] = ms2fas

    def ms2_headgroup(self, ms2r, hgfrags, headgroups):
        '''
        Identifies headgroups from MS2 results for one
        feature, based on dictionary of fragments and
        the characteristic combinations of fragments
        identifying headgroups.
        '''
        hgroups = None
        frags = set([])
        # collecting all possible headgroups for
        # each fragment into `hgroups` and all
        # fragments into `frags`
        for ms2item in ms2r:
            if ms2item[0] in hgfrags:
                hgroups = hgfrags[ms2item[0]] if hgroups is None \
                    else hgroups & hgfrags[ms2item[0]]
                frags.add(ms2item[0])
        # if any headgroup related fragment found
        if hgroups is not None and len(hgroups) > 0:
            missingFrags = set([])
            for hg in hgroups:
                if len(headgroups[hg] - frags) > 0:
                    missingFrags.add(hg)
            hgroups = hgroups - missingFrags if \
                len(hgroups - missingFrags) > 0 else hgroups
        return hgroups

    def ms2_headgroup2(self, ms2r, hgfrags, headgroups):
        '''
        Identifies headgroups from MS2 results for one
        feature, based on dictionary of fragments and
        the characteristic combinations of fragments
        identifying headgroups.
        '''
        _hgroups = []
        _hgroups2 = []
        hgroups = set([])
        frags = set([])
        # collecting all possible headgroups for
        # each fragment into `hgroups` and all
        # fragments into `frags`
        for ms2item in ms2r:
            if ms2item[0] in hgfrags:
                _hgroups.append(hgfrags[ms2item[0]])
                frags.add(ms2item[0])
        # if any headgroup related fragment found
        if len(_hgroups) > 0:
            for hgs in _hgroups:
                added = False
                for i, hgs2 in enumerate(_hgroups2):
                    if len(hgs & hgs2) > 0:
                        _hgroups2[i] = hgs & hgs2
                        added = True
                if not added:
                    _hgroups2.append(hgs)
        for hgs in _hgroups2:
            hgroups = hgroups | hgs
        return hgroups

    def headgroups_negative_positive(self, ms):
        '''
        Creates dictionaries named ms1hg_pos, ms1hg_neg,
        ms2hg_pos or ms2hg_neg with the original
        IDs of the given mode as keys, with dicts as
        values having the original IDs of the other mode
        as keys and the combined set of headgroups as 
        values. The combined set is the intersection of
        those detected in the 2 modes, or the union, if
        there is no intersection.
        '''
        for ltp, d in self.valids.iteritems():
            d['pos']['%shg_neg'%ms] = {}
            d['neg']['%shg_pos'%ms] = {}
            for poi, nois in d['pos']['neg'].iteritems():
                if poi in d['pos']['%shg'%ms]:
                    if d['pos']['%shg'%ms][poi] is not None:
                        for noi in nois.keys():
                            if noi in d['neg']['%shg'%ms]:
                                if d['neg']['%shg'%ms][noi] is not None:
                                    if poi not in d['pos']['%shg_neg'%ms]:
                                        d['pos']['%shg_neg'%ms][poi] = {}
                                    if noi not in d['neg']['%shg_pos'%ms]:
                                        d['neg']['%shg_pos'%ms][noi] = {}
                                    poshg = d['pos']['%shg'%ms][poi]
                                    neghg = d['neg']['%shg'%ms][noi]
                                    combined = poshg & neghg \
                                        if len(poshg & neghg) > 0 \
                                        else poshg | neghg
                                    d['pos']['%shg_neg'%ms][poi][noi] = \
                                        combined
                                    d['neg']['%shg_pos'%ms][noi][poi] = \
                                        combined

    def ms2_fattya(self, ms2r, highest = 2):
        '''
        Identifies the fatty acids with highest intensities.
        Returns only number of `highest` fatty acids.
        '''
        recnum = re.compile(r'.*[^0-9]([0-9]+:[0-9]+).*')
        # reverse sort by intensities
        ms2fa = set([])
        ms2fai = []
        ms2fas = None
        # if we have MS2 data at all:
        if ms2r.shape[0] > 0:
            ms2r = ms2r[ms2r[:,3].argsort()[::-1],:]
            fanum = 0
            for ms2f in ms2r:
                if 'FA' in ms2f[0] or 'Lyso' in ms2f[0]:
                    fa = recnum.match(ms2f[0])
                    fa = fa.groups()[0]
                    if fa not in ms2fa:
                        # a set with fatty std formulas
                        # [carbon count]:[unsaturated count]
                        ms2fa.add(fa)
                        # a list with intensities decreasing
                        ms2fai.append((fa, ms2f[3]))
                        fanum += 1
                if fanum == highest:
                    break
            # sum formula
            if len(ms2fa) == highest:
                carbs = 0
                unsats = 0
                for fa in ms2fa:
                    carb, unsat = map(int, fa.split(':'))
                    carbs += carb
                    unsats += unsat
                ms2fas = '%u:%u' % (carbs, unsats)
        return ms2fa, ms2fai, ms2fas
    
    '''
    END: MS2 functions
    '''
    
    '''
    BEGIN: New identification methods
    '''
    
    def ms2_scans_identify(self):
        prg = progress.Progress(len(self.valids) * 2,
            'Analysing MS2 scans and identifying features', 1, percent = False)
        for protein, d in self.valids.iteritems():
            for mode, tbl in d.iteritems():
                prg.step()
                tbl['ms2f'] = {}
                tbl['ms2i'] = {}
                for i, oi in enumerate(tbl['i']):
                    if oi in tbl['ms2']:
                        tbl['ms2f'][oi] = Feature(self, protein, mode, oi)
                        tbl['ms2f'][oi].identify()
                        tbl['ms2i'][oi] = tbl['ms2f'][oi].identities
        prg.terminate()
    
    '''
    END: New identification methods
    '''
    
    '''
    Pipeline elements
    '''
    
    def write_out(self, matches, fname):
        '''
        In:
        [0] pos_m/z, pos_profile_score, 
        [2] pos_control_profile_score, pos_rank_profile_boolean,
        [4] pos_ubiquity_score, pos_ubiquity_score,
        [6] pos_original_index, pos_swisslipids_ac, pos_level, 
        [9] pos_lipid_name, pos_lipid_formula, pos_adduct, pos_adduct_m/z
        [13] neg_m/z, neg_profile_score, neg_control_profile_score,
        [16] neg_rank_profile_boolean,
        [17] neg_ubiquity_score, neg_ubiquity_score, neg_original_index,
        [20] neg_swisslipids_ac, neg_level,
        [22] neg_lipid_name, neg_lipid_formula, neg_adduct, neg_adduct_m/z
        '''
        with open(fname, 'w') as f:
            hdr = ['LTP',
                'Positive_m/z', 'Positive_m/z_in_SwissLipids',
                'Positive_adduct',
                'Negative_m/z', 'Negative_m/z_in_SwissLipids', 
                'Negative_adduct',
                'SwissLipids_AC', 'SwissLipids_formula', 'SwissLipids_name']
            f.write('\t'.join(hdr) + '\n')
            for ltp, tbl in matches.iteritems():
                for l in tbl:
                    f.write('\t'.join([ltp, '%08f'%l[0],
                        '%08f'%l[12], l[11], '%08f'%l[13],
                        '%08f'%l[25], l[24], l[7], l[10], str(l[9])]) + '\n')

    def counts_redundancy_table(self, lipids, unknowns):
        with open('unique_features_counts.csv', 'w') as f:
            f.write('\t'.join(['LTP-mode', 'unknown_features', 
                'lipid_matching_features', 'lipids']) + '\n')
            f.write('\n'.join('%s\t%u\t%u\t%u'%i \
                for i in zip(
                ['%s-%s'%(b,a) for b in unknowns.keys() \
                    for a in unknowns[b].keys()],
                [len(uniqList(list(a[:,7]))) for b in lipids.values() \
                    for a in b.values()], 
                [len(uniqList(list(a[:,7]))) for b in unknowns.values() \
                    for a in b.values()],
                [len(list(a[:,7])) for b in lipids.values() \
                    for a in b.values()])))

    def samples_with_controls(self):
        self.csamples = dict((
                k,
                np.array([1 if x is not None else None for x in v])
            )
            for k, v in self.samples.iteritems())

    def upper_samples(self):
        '''
        Creates the dict `samples_upper` of samples with
        uppercase LTP names.
        '''
        self.samples_upper = \
            dict((l.upper(), s) for l, s in self.samples.iteritems())

    def init_from_scratch(self):
        '''
        Does all the initial preprocessing.
        Saves intermediate data, so it can be loaded faster 
        for next sessions.
        '''
        self.get_filenames()
        self.read_samples()
        # at first run, after reading from saved textfile
        self.protein_profiles()
        self.write_pptable()
        del self.datafiles['ctrl']
        self.small_inputs()
        self.save()
        self.read_data()
        self.save_data()

    def init_reinit(self, data = False):
        '''
        Initializing from preprocessed and dumped data.
        Pickle file has a 2.0GB size.
        '''
        self.load()
        if data:
            self.load_data()
        self.small_inputs()
        
    def small_inputs(self):
        self.samples_with_controls()
        self.upper_samples()
        self.one_sample()
        self.zero_controls()
        self.read_lipid_names()
        self.read_binding_properties()

    def basic_filters(self, profile_treshold = 0.25, ubiquity_treshold = 7):
        '''
        Deprecated with new data structure.
        '''
        self.apply_filters()
        self.validity_filter()
        self.profile_filter()
        self.profile_filter(prfx = 'c')
        self.ubiquity_filter()
        self.val_ubi_filter(ubiquity = ubiquity_treshold)
        self.rprofile_filter()
        self.val_prf_filter(treshold = profile_treshold)
        self.val_rpr_filter()
        self.val_ubi_prf_filter(treshold = profile_treshold,
            ubiquity = ubiquity_treshold)
        self.val_ubi_prf_rprf_filter(treshold = profile_treshold,
            ubiquity = ubiquity_treshold)

    def basic_filters_with_evaluation(self,
        profile_treshold = 0.25, ubiquity_treshold = 7):
        '''
        Deprecated with new data structure.
        '''
        filtr_results = {}
        for f in ['quality', 'charge', 'area', 'peaksize', 'validity']:
            filtr_results[f] = self.eval_filter(f)
        filtr_results['ubiquity'] = self.eval_filter('ubiquity',
        param = {'only_valid': False},
        runtime = True, repeat = 1, number = 1,
        hit = lambda x: x < ubiquity_treshold)
        filtr_results['val_ubi'] = eval_filter('val_ubi',
            param = {'ubiquity': ubiquity_treshold})
        filtr_results['profile025'] = eval_filter('profile',
            param = {'prfx': ''},
            runtime = True, repeat = 1, number = 1, 
            hit = lambda x: x <= profile_treshold)
        filtr_results['cprofile025'] = self.eval_filter('cprofile',
            runtime = True, repeat = 1, number = 1, 
            hit = lambda x: x <= profile_treshold)
        filtr_results['rprofile'] = self.eval_filter('rprofile',
            runtime = True)
        filtr_results['val_prf'] = self.eval_filter('val_prf',
            param = {'treshold': profile_treshold})
        filtr_results['val_rpr'] = self.eval_filter('val_rpr')
        filtr_results['val_ubi_prf'] = self.eval_filter('val_ubi_prf',
            param = {'treshold': profile_treshold,
                     'ubiquity': ubiquity_treshold})
        filtr_results['val_ubi_prf_rprf'] = \
            self.eval_filter('val_ubi_prf_rprf',
            param = {'treshold': profile_treshold,
                     'ubiquity': ubiquity_treshold})
        self.filter_results = filtr_results

    def positive_negative_runtime(self, ):
        return timeit.timeit('negative_positive(lipids)', 
            setup = 'from __main__ import negative_positive, lipids',
            number = 1)

    def valid_features(self, cache = True):
        '''
        Creates new dict of arrays with only valid features.
        Keys:
            'fe': features
            'mz': m/z values
            'i': original index
        '''
        if cache and os.path.exists(self.validscache):
            self.valids = pickle.load(open(self.validscache, 'rb'))
            return None
        self.apply_filters()
        self.validity_filter()
        self.valids = dict((ltp.upper(), {'pos': {}, 'neg': {}}) \
            for ltp in self.data.keys())
        for ltp, d in self.data.iteritems():
            for pn, tbl in d.iteritems():
                self.valids[ltp.upper()][pn]['fe'] = np.array(tbl['smp'][tbl['vld']])
                self.valids[ltp.upper()][pn]['mz'] = \
                    np.array(tbl['raw'][tbl['vld'], 1])
                self.valids[ltp.upper()][pn]['rt'] = \
                    np.array(tbl['raw'][tbl['vld'], 2:4])
                self.valids[ltp.upper()][pn]['i'] = np.where(tbl['vld'])[0]
        self.norm_all()
        pickle.dump(self.valids, open(self.validscache, 'wb'))

    def norm_all(self):
        '''
        Creates table with all the profiles normalized
        Keys:
            'no': normalized profiles
        '''
        for ltp, d in self.valids.iteritems():
            for pn, tbl in d.iteritems():
                tbl['no'] = self.norm_profiles(tbl['fe'])

    '''
    END: pipeline elements
    '''

    '''
    Distance metrics
    '''

    def profiles_corr(self, metric, prfx):
        '''
        Calculates custom correlation metric
        between each feature and the protein profile.
        '''
        frs = ['c0', 'a9', 'a10', 'a11', 'a12', 'b1']
        for ltp, d in self.valids.iteritems():
            ppr = np.array([self.pprofs[ltp.upper()][frs[i]] \
                for i, fr in enumerate(self.samples_upper[ltp]) \
                    if fr is not None and i != 0])
            ppr = self.norm_profile(ppr).astype(np.float64)
            for pn, tbl in d.iteritems():
                if 'no' not in tbl:
                    self.norm_all()
                tbl['%sv'%prfx] = np.zeros((tbl['no'].shape[0],),
                    dtype = np.float64)
                tbl['%sp'%prfx] = np.zeros((tbl['no'].shape[0],),
                    dtype = np.float64)
                for i, fe in enumerate(tbl['no']):
                    # if one feature is not detected
                    # in any fraction of the sample,
                    # it will have a nan value:
                    if np.any(np.isnan(\
                            fe[np.where(\
                                [fr == 1 \
                                    for fr in self.samples_upper[ltp][1:] \
                                    if fr is not None]
                            )]
                        )):
                        vp = (np.nan, 0.0)
                    else:
                        vp = metric(fe, ppr)
                    tbl['%sv'%prfx][i] = vp[0]
                    tbl['%sp'%prfx][i] = vp[1]

    def gkgamma(self, x, y):
        '''
        Calls Goodman-Kruskal's gamma from vcdExtra
        R package.
        '''
        gkg = rvcd.GKgamma(rbase.matrix(rbase.c(*(list(x) + list(y))),
            nrow = 2))
        # gamma, 0.0, C, D, sigma, CIlevel, CI
        return tuple([gkg[0][0], 0.0] + [i[0] for i in gkg[1:]])

    def roco(self, x, y):
        '''
        Calls R function robust correlation coefficient
        with test from rococo R package.
        '''
        _x = x[np.where(~np.logical_or(np.isnan(x), np.isnan(y)))]
        _y = y[np.where(~np.logical_or(np.isnan(x), np.isnan(y)))]
        if _x.size == 0 or _y.size == 0 or \
            np.all(_x == _x[0]) or np.all(_y == _y[0]):
            return 0.0, 0.0
        _x = rbase.c(*_x)
        _y = rbase.c(*_y)
        rt = rococo.rococo_test( \
            rbase.as_vector(rstats.na_omit(_x.rx( \
                rbase.c(*list(np.where(\
                    map(lambda i: not i, list(rbase.is_na(_y))))[0] + 1))))), \
            rbase.as_vector(rstats.na_omit(_y.rx( \
                rbase.c(*list(np.where(\
                    map(lambda i: not i, list(rbase.is_na(_x))))[0] + 1))))))
        return rt.slots['sample.gamma'][0], rt.slots['p.value'][0]

    def _diff_profiles(self, x, y):
        '''
        Wrapper for diff_profiles() to return a tuple.
        '''
        return self.diff_profiles(x, y), 0.0

    def euclidean_dist(self, x, y):
        '''
        Calculates simple euclidean distance after removing NaN values.
        '''
        _x = x[np.where(~np.logical_or(np.isnan(x), np.isnan(y)))]
        _y = y[np.where(~np.logical_or(np.isnan(x), np.isnan(y)))]
        return (sp.spatial.distance.euclidean(_x, _y), 0.0) \
            if len(_x) > 0 else (np.inf, 0.0)

    def euclidean_dist_norm(self, x, y):
        '''
        This euclidean distance is normalized by the number of dimensions.
        '''
        _x = x[np.where(~np.logical_or(np.isnan(x), np.isnan(y)))]
        _y = y[np.where(~np.logical_or(np.isnan(x), np.isnan(y)))]
        return (sp.spatial.distance.euclidean(_x, _y) / len(_x), 0.0) \
            if len(_x) > 0 else (1.0, 0.0)

    def profiles_corrs(self):
        '''
        Calculates an array of similarity/correlation
        metrics between each MS intensity profile and 
        the corresponding protein concentration profile.
        '''
        metrics = [
            (stats.spearmanr, 'sp'),
            (stats.kendalltau, 'kt'),
            (self._diff_profiles, 'df'),
            (self.gkgamma, 'gk'),
            (self.roco, 'rc'),
            (self.euclidean_dist, 'eu'),
            (self.euclidean_dist_norm, 'en'),
            (stats.pearsonr, 'pe'),
            (self._comp_profiles, 'cp')
        ]
        for metric, prfx in metrics:
            sys.stdout.write('Calculating %s\n' % metric.__name__)
            self.profiles_corr(metric, prfx)

    def inconsistent(self, valids, metric = 'en'):
        attr = '%sv'%metric
        target = '%si'%metric
        sort_alll(valids, attr)
        for ltp, d in valids.iteritems():
            for pn, tbl in d.iteritems():
                vals = tbl[attr]
                incs = np.diff(vals)
                inco = np.array(
                    [0.0] * 3 + \
                    map(lambda i:
                        incs[i] - np.mean(incs[:i]) / np.std(incs[:i]),
                        xrange(2, len(incs))
                    )
                )
                tbl['%sde'%metric] = np.concatenate((np.array([0.0]), incs),
                    axis = 0)
                tbl['%sdde'%metric] = np.concatenate((np.array([0.0]), 
                    np.diff(tbl['%sde'%metric])), axis = 0)
                tbl[target] = inco

    '''
    END: disctance metrics
    '''

    '''
    Functions for clustering
    '''

    def distance_matrix(self, metrics = ['eu'], with_pprof = False, ltps = None):
        _metrics = {
            'eu': ('euclidean distance', self.euclidean_dist),
            'en': ('normalized euclidean distance', self.euclidean_dist_norm)
        }
        frs = ['c0', 'a9', 'a10', 'a11', 'a12', 'b1']
        t0 = time.time()
        for m in metrics:
            prg = progress.Progress(
                len(self.valids) * 2 if ltps is None else len(ltps) * 2,
                'Calculating %s' % _metrics[m][0],
                1, percent = False)
            for ltp, d in self.valids.iteritems():
                if ltps is None or ltp in ltps:
                    if with_pprof:
                        ppr = np.array([self.pprofs[ltp.upper()][frs[i]] \
                            for i, fr in enumerate(self.samples_upper[ltp]) \
                                if i != 0 and fr is not None])
                        ppr = self.norm_profile(ppr).astype(np.float64)
                    for pn, tbl in d.iteritems():
                        prg.step()
                        fnum = tbl['no'].shape[0]
                        # square shape matrix of all features vs all features
                        if with_pprof:
                            tbl['_%sd'%m] = np.empty((fnum + 1, fnum + 1))
                        else:
                            tbl['_%sd'%m] = np.empty((fnum, fnum))
                        # to keep track the order accross sorting
                        tbl['_%so'%m] = np.copy(tbl['i'])
                        for i in xrange(fnum):
                            for j in xrange(fnum):
                                tbl['_%sd'%m][i,j] = \
                                    _metrics[m][1](tbl['no'][i,:],
                                        tbl['no'][j,:])[0]
                        if with_pprof:
                            for i in xrange(fnum):
                                ppr_dist = _metrics[m][1](tbl['no'][i,:], ppr)[0]
                                tbl['_%sd'%m][i,-1] = ppr_dist
                                tbl['_%sd'%m][-1,i] = ppr_dist
                            tbl['_%sd'%m][-1,-1] = _metrics[m][1](ppr, ppr)[0]
        prg.terminate()
        sys.stdout.write('\t:: Time elapsed: %us\n'%(time.time() - t0))
        sys.stdout.flush()

    def features_clustering(self, dist = 'en', method = 'ward', ltps = None):
        '''
        Using the distance matrices calculated by
        `distance_matrix()`, builds clusters using
        the linkage method given by `method` arg.
        '''
        prg = progress.Progress(len(self.valids)*2 \
            if ltps is None else len(ltps)*2,
            'Calculating clusters', 1, percent = False)
        for ltp, d in self.valids.iteritems():
            if ltps is None or ltp in ltps:
                for pn, tbl in d.iteritems():
                    prg.step()
                    tbl['_%sc'%dist] = fastcluster.linkage(tbl['_%sd'%dist],
                        method = method, metric = 'euclidean',
                            preserve_input = True)
        prg.terminate()

    def distance_corr(self, valids, dist = 'en'):
        for ltp, d in valids.iteritems():
            for pn, tbl in d.iteritems():
                tbl['_%sdc'%dist] = np.array([
                        sp.stats.pearsonr(tbl['_%sd'%dist][:,i],
                            tbl['_%sd'%dist][:,-1])[0] \
                    for i in xrange(tbl['_%sd'%dist].shape[1])])

    def fcluster_with_protein(self, lin, t):
        fc = sp.cluster.hierarchy.fcluster(lin, t, criterion = 'distance')
        return np.where(fc == fc[-1])[0]

    def nodes_in_cluster(self, lin, i):
        n = lin.shape[0]
        nodes = map(int, lin[i,:2])
        singletons = set(filter(lambda x: x <= n, nodes))
        upper_levels = set(nodes) - singletons
        for u in upper_levels:
            singletons = singletons | self.nodes_in_cluster(lin, u - n - 1)
        return singletons

    def _get_link_colors_dist(self, tbl, dist, threshold,
        highlight_color, base_color):
        protein_fc = set(self.fcluster_with_protein(tbl['_%sc' % dist], threshold))
        return dict(zip(
            xrange(tbl['_%sc'%dist].shape[0] + 1,
                   tbl['_%sc'%dist].shape[0] * 2 + 2),
            map(lambda x: 
                highlight_color \
                    if set(self.nodes_in_cluster(tbl['_%sc' % dist], x)) <= \
                        protein_fc \
                    else base_color,
            xrange(tbl['_%sc'%dist].shape[0])) + \
                [highlight_color \
                    if len(protein_fc) == tbl['_%sd'%dist].shape[0] \
                    else base_color]
        ))

    def _get_link_colors_corr(self, tbl, dist, cmap, threshold,
        highlight_color, base_color):
        return dict(zip(
            xrange(tbl['_%sc'%dist].shape[0] + 1,
                tbl['_%sc'%dist].shape[0]*2 + 2), # 180
            map(lambda col:
                '#%s' % ''.join(map(lambda cc: '%02X'%(cc*255), col[:3])) \
                    if type(col) is tuple else col,
                map(lambda c:
                    # color from correlation value:
                    #'#%s%s00' % ('%02X'%(c*255), '%02X'%(c*255)) \
                    cmap(c) if c > 0.0 else base_color, # '#FEFE00'
                    map(lambda x:
                        # minimum of these correlations
                        # for each link:
                        min(map(lambda xx:
                            # correlation of distances for one link:
                            tbl['_%sdc'%dist][xx], # 0.9999928, 0.9999871
                            # nodes for each link:
                            list(nodes_in_cluster(tbl['_%sc' % dist], x))
                        )),
                        # all links (rows in linkage matrix):
                        xrange(tbl['_%sc'%dist].shape[0]) # 45
                    ) + \
                    # this is for the root:
                    [min(tbl['_%sdc'%dist])]
                )
            )
        ))

    def _get_dendrogram_cmap(self, cmap, threshold,
        highlight_color, base_color):
        if cmap is None:
            if thershold is None:
                _cmap = 'inferno'
            else:
                _cmap = lambda x: \
                    highlight_color if x >= threshold else base_color
        elif type(cmap) is str and hasattr(mpl.cm, cmap):
                _cmap = mpl.cm.get_cmap(cmap)
        return _cmap

    def _cluster_size_threshold(self, tbl, dist, threshold):
        dist_values = tbl['_%sc'%dist][:,2]
        dist_values = dist_values[np.where(dist_values > 0.0)]
        if threshold is None:
            threshold = len(dist_values) * 0.1
        _max = np.max(dist_values)
        _min = np.min(dist_values)
        _lower = _min
        _upper = _max
        _conv = 9999
        _dconvs = []
        while True:
            _threshold = (_lower + _upper) / 2.0
            cls = sp.cluster.hierarchy.fcluster(tbl['_%sc'%dist], _threshold)
            mean_size = np.mean(collections.Counter(cls).values())
            _over = mean_size > threshold
            if mean_size > threshold:
                _upper = max(_upper - (_upper - _lower) / 2.0, 0.0)
            elif mean_size < threshold:
                _lower = max(_lower + (_upper - _lower) / 2.0, 0.0)
            if _lower >= _upper:
                _upper += (_upper - _lower) / 4.0
            _dconvs.append(_conv - abs(mean_size - threshold))
            _conv = abs(threshold - mean_size)
            if len(_dconvs) > 10 and \
                abs(np.mean(_dconvs[-10:]) - _dconvs[-1]) < 0.005:
                return _threshold

    def _inconsistency_threshold(self, tbl, dist, threshold):
        n = tbl['_%sd'%dist].shape[0]
        if threshold is None:
            threshold = n/20
        clustering = tbl['_%sc'%dist]
        incons = sp.cluster.hierarchy.inconsistent(clustering)
        maxincons = sp.cluster.hierarchy.maxinconsts(clustering, incons)
        fc = sp.cluster.hierarchy.fcluster(clustering, threshold, 
            criterion = 'maxclust_monocrit', monocrit = maxincons)
        nodes = set(np.where(fc == fc[-1])[0])
        print 'nodes in same cluster as protein: %u' % len(nodes)
        _threshold = max(
            map(lambda i: 
                clustering[i,2],
                filter(lambda i: 
                    nodes_in_cluster(clustering, i) <= nodes,
                    xrange(n - 1)
                )
            )
        )
        return _threshold

    def _dendrogram_get_threshold(self, tbl, dist, threshold, threshold_type):
        dist_values = tbl['_%sc'%dist][:,2]
        if threshold_type == 'percent':
            # _threshold = dist_values.max() * threshold / tbl['%sc'%dist].shape[0]
            _threshold = 10**(np.log10(dist_values.max()) * threshold)
        elif threshold_type == 'quantile':
            _threshold = 10**(np.percentile(
                np.log10(dist_values)[np.where(dist_values > 0.0)], threshold))
        elif threshold_type == 'clsize':
            _threshold = _cluster_size_threshold(tbl, dist, threshold)
        elif threshold_type == 'incons':
            _threshold = _inconsistency_threshold(tbl, dist, threshold)
        return _threshold

    def plot_heatmaps_dendrograms_gradient(self, *args, **kwargs):
        pass

    def plot_heatmaps_dendrograms(self, dist = 'en', 
        fname = None, ltps = None, cmap = None,
        highlight_color = '#FFAA00', base_color = '#000000',
        coloring = 'corr', threshold = None,
        threshold_type = 'percent',
        save_selection = None, pca = False):
        '''
        For each LTP plots heatmaps and dendrograms.
        Thanks to http://stackoverflow.com/a/3011894/854988
        '''
        all_hgs = set()
        for ltp, d in self.valids.iteritems():
            for pn, tbl in d.iteritems():
                for ids in tbl['identity'].values():
                    for hg in ids.keys():
                        all_hgs.add(hg)
        hg_cols = dict(map(lambda (i, hg): 
            (hg, self.colors[i]),
            enumerate(sorted(list(all_hgs)))
        ))
        t0 = time.time()
        fname = 'features_clustering-%s%s%s.pdf' % \
                (coloring, 
                '-%s'%threshold_type if threshold is not None else '',
                '-%02f'%threshold if threshold is not None else '') \
            if fname is None else fname
        if coloring == 'corr':
            _cmap = self._get_dendrogram_cmap(cmap, threshold,
                highlight_color, base_color)
        frs = ['c0', 'a9', 'a10', 'a11', 'a12', 'b1']
        with mpl.backends.backend_pdf.PdfPages(fname) as pdf:
            prg = progress.Progress(len(self.valids) * 2 \
                if ltps is None else len(ltps)*2,
                'Plotting heatmaps with dendrograms', 1, percent = False)
            for ltp, d in self.valids.iteritems():
                if ltps is None or ltp in ltps:
                    ppr = np.array([self.pprofs[ltp.upper()][frs[i]] \
                        for i, fr in enumerate(self.samples_upper[ltp]) \
                            if i != 0 and fr is not None])
                    ppr = self.norm_profile(ppr).astype(np.float64)
                    for pn, tbl in d.iteritems():
                        prg.step()
                        if coloring == 'dist':
                            _threshold = self._dendrogram_get_threshold(tbl, dist,
                                threshold, threshold_type)
                        labels = ['%u'%(f) for f in tbl['_%so'%dist]] + [ltp]
                        names = map(lambda oi:
                            ', '.join(sorted(map(lambda (hg, m):
                                hg,
                                filter(lambda (hg, m):
                                    m['ms1_%s'%pn] and m['ms2_%s'%pn],
                                    tbl['identity'][int(oi)].iteritems()
                                )
                            ))) or oi,
                            labels[:-1]
                        ) + [ltp]
                        
                        
                        if coloring == 'corr':
                            _link_colors = \
                                self._get_link_colors_corr(tbl, dist, _cmap,
                                threshold, highlight_color, base_color)
                        elif coloring == 'dist':
                            _link_colors = self._get_link_colors_dist(tbl, dist,
                                _threshold, highlight_color, base_color)
                        protein_fc = \
                            set(self.fcluster_with_protein(tbl['_%sc' % dist],
                            _threshold))
                        if save_selection is not None:
                            oi2i = \
                                dict(zip(labels[:-1], xrange(len(labels) - 1)))
                            tbl[save_selection] = np.array(
                                map(lambda oi:
                                    oi2i['%u'%oi] in protein_fc,
                                    tbl['i']
                                )
                            )
                        
                        mpl.rcParams['lines.linewidth'] = 0.1
                        mpl.rcParams['font.family'] = 'Helvetica Neue LT Std'
                        fig = mpl.figure.Figure(figsize = (8, 8))
                        cvs = mpl.backends.backend_pdf.FigureCanvasPdf(fig)
                        gs = mpl.gridspec.GridSpec(2, 2, 
                            height_ratios=[2, 8], width_ratios = [2, 8])
                        
                        # First dendrogram
                        ax1 = fig.add_subplot(gs[1,0])
                        Z1 = hc.dendrogram(tbl['_%sc'%dist],
                            orientation = 'left',
                            labels = names,
                            leaf_rotation = 0, ax = ax1,
                            link_color_func = lambda i: _link_colors[i])
                        ax1.yaxis.grid(False)
                        ax1.set_xscale('symlog')
                        null = [tl.set_fontsize(1.5) \
                            for tl in ax1.get_yticklabels()]
                        null = [(tl.set_color(hg_cols[tl._text]),
                                tl.set_fontweight('bold')) \
                            for tl in ax1.get_yticklabels() \
                                if tl._text in hg_cols]
                        null = [(tl.set_color(highlight_color),
                                tl.set_fontweight('bold'),
                                tl.set_fontsize(4)) \
                            for tl in ax1.get_yticklabels() if tl._text == ltp]
                        if _threshold is not None:
                            ax1.axvline(x = _threshold,
                                c = '#FFAA00', alpha = 0.5)
                        
                        # Compute and plot second dendrogram.
                        ax2 = fig.add_subplot(gs[0,1])
                        Z2 = hc.dendrogram(tbl['_%sc'%dist], 
                            labels = names,
                            leaf_rotation = 90, ax = ax2,
                            #color_threshold = _threshold)
                            link_color_func = lambda i: _link_colors[i])
                        ax2.xaxis.grid(False)
                        ax2.set_yscale('symlog')
                        null = [tl.set_fontsize(1.5) \
                            for tl in ax2.get_xticklabels()]
                        null = [(tl.set_color(hg_cols[tl._text]),
                                tl.set_fontweight('bold')) \
                            for tl in ax2.get_xticklabels() \
                                if tl._text in hg_cols]
                        null = [(tl.set_color(highlight_color), 
                                tl.set_fontweight('bold'),
                                tl.set_fontsize(4)) \
                            for tl in ax2.get_xticklabels() if tl._text == ltp]
                        if _threshold is not None:
                            ax2.axhline(y = _threshold, c = '#FFAA00',
                                alpha = 0.5)
                        
                        # Plot distance matrix.
                        ax3 = fig.add_subplot(gs[1,1])
                        idx1 = Z1['leaves']
                        idx2 = Z2['leaves']
                        D = tbl['_%sd'%dist][idx1,:]
                        D = D[:,idx2]
                        im = ax3.matshow(D, aspect = 'auto', origin = 'lower',
                            cmap = mpl.cm.get_cmap('Blues'))
                        ax3.xaxis.grid(False)
                        ax3.yaxis.grid(False)
                        ax3.set_xticklabels([])
                        ax3.set_yticklabels([])
                        
                        fig.suptitle('%s :: %s mode\n'\
                            'clustering valid features; '\
                            'features in highlighted cluster: %u' % \
                            (ltp, pn, len(protein_fc) - 1),
                            color = '#AA0000' if ltp in self.singles else '#000000')
                        
                        cvs.draw()
                        fig.tight_layout()
                        fig.subplots_adjust(top = 0.90)
                        
                        cvs.print_figure(pdf)
                        fig.clf()
                        
                        if pca:
                            oi2i = dict(zip(tbl['i'],
                                xrange(tbl['no'].shape[0])))
                            pca = sklearn.decomposition.PCA(n_components = 2)
                            fe = np.vstack((tbl['no'][map(lambda oi:
                                    oi2i[int(oi)],
                                    labels[:-1]
                                ),:], ppr))
                            col = map(lambda i:
                                highlight_color \
                                    if i in protein_fc else base_color,
                                xrange(len(labels) - 1)
                            ) + ['#3383BE']
                            fe[np.where(np.isnan(fe))] = 0.0
                            pca = pca.fit(fe)
                            coo = pca.transform(fe)
                            fig = mpl.figure.Figure(figsize = (8, 8))
                            cvs = mpl.backends.backend_pdf.FigureCanvasPdf(fig)
                            ax = fig.gca()
                            ax.scatter(coo[:,0], coo[:,1], c = col, 
                                linewidth = 0.0, alpha = 0.7)
                            ax.set_title('%s :: %s mode :: PCA' % (ltp, pn),
                                color = '#AA0000' \
                                    if ltp in singles else '#000000')
                            cvs.draw()
                            fig.tight_layout()
                            cvs.print_figure(pdf)
                            fig.clf()
                        
            pdfinf = pdf.infodict()
            pdfinf['Title'] = 'Features clustering'
            pdfinf['Author'] = 'Dénes Türei'.decode('utf-8')
            pdfinf['Subject'] = 'Clustering MS1 features '\
                'based on Euclidean distances'
            pdfinf['Keywords'] = 'lipid transfer protein, LTP,'\
                ' lipidomics, mass spectrometry'
            pdfinf['CreationDate'] = datetime.datetime(2016, 02, 22)
            pdfinf['ModDate'] = datetime.datetime.today()
        
        prg.terminate()
        sys.stdout.write('\t:: Time elapsed: %us\n'%(time.time() - t0))
        sys.stdout.write('\t:: Plots saved to %s\n'%fname)
        sys.stdout.flush()

    '''
    END: Clustering
    '''

    def plot_increment(self, valids, singles, metric = 'en',
        fname = 'increments.pdf'):
        with mpl.backends.backend_pdf.PdfPages(fname) as pdf:
            for ltp, d in valids.iteritems():
                for pn, tbl in d.iteritems():
                    fig = mpl.figure.Figure(figsize = (8,8))
                    cvs = mpl.backends.backend_pdf.FigureCanvasPdf(fig)
                    ax = fig.gca()
                    ax.plot(xrange(tbl['%si'%metric].shape[0]),
                        tbl['%si'%metric], c = '#FCCC06',
                        label = 'Inconsistency')
                    ax.plot(xrange(tbl['%sde'%metric].shape[0]),
                        tbl['%sde'%metric] * 100, c = '#007B7F',
                        label = '1st derivate',
                        lw = 2, alpha = 0.5)
                    ax.plot(xrange(tbl['%sde'%metric].shape[0]),
                        tbl['%sdde'%metric] * 100, c = '#6EA945',
                        label = '2nd derivate',
                        lw = 2, alpha = 0.5)
                    handles, labels = ax.get_legend_handles_labels()
                    ax.legend(handles, labels)
                    ax.set_title('%s :: %s mode :: distance increments' % \
                        (ltp, pn),
                        color = '#AA0000' if ltp in singles else '#000000')
                    ax.set_xlim([3.0, 50.0])
                    cvs.print_figure(pdf)
                    fig.clf()
            
            pdfinf = pdf.infodict()
            pdfinf['Title'] = 'Features distance increment'
            pdfinf['Author'] = 'Dénes Türei'.decode('utf-8')
            pdfinf['Subject'] = 'Features distance increment'
            pdfinf['Keywords'] = 'lipid transfer protein, LTP,'\
                ' lipidomics, mass spectrometry'
            pdfinf['CreationDate'] = datetime.datetime(2016, 02, 22)
            pdfinf['ModDate'] = datetime.datetime.today()

    def kmeans(self, valids, pprofs, samples):
        frs = ['c0', 'a9', 'a10', 'a11', 'a12', 'b1']
        prg = progress.Progress(len(valids) * 2, 
            'Calculating k-means', 1, percent = False)
        for ltp, d in valids.iteritems():
            ppr = np.array([pprofs[ltp.upper()][frs[i]] \
                for i, fr in enumerate(samples[ltp]) if fr == 1 and i != 0])
            ppr = norm_profile(ppr).astype(np.float64)
            for pn, tbl in d.iteritems():
                prg.step()
                with_protein = np.vstack(tbl['no'], ppr)
                whitened = sp.cluster.vq.whiten(with_protein)
                code_book = sp.cluster.vq.kmeans(whitened, 2)
                # tbl['km'] = 
        prg.terminate()

    def read_positives(self, basedir, fname = 'manual_positive.csv'):
        '''
        Reads manually annotated positive hits
        from file.
        '''
        with open(os.path.join(basedir, fname), 'r') as f:
            _result = [[c.strip() for c in l.split('\t')] \
                for l in f.read().split('\n')]
        result = dict((ltp, []) for ltp in uniqList([x[0] \
            for x in _result if x[0] != '']))
        for l in _result:
            if l[0] != '':
                result[l[0]].append(l)
        return result

    def spec_sens(self, valids, ltp, pos, metric, asc):
        '''
        Calculates specificity, sensitivity, precision,
        false discovery rate as a function of critical
        values of a score, for one LTP for one mode.
        Returns dict of lists.
        '''
        result = {'spec': [], 'sens': [], 'prec': [],
            'fdr': [], 'cutoff': [], 'n': 0}
        tbl = valids[ltp][pos]
        ioffset = 0 if pos == 'pos' else 6
        if len(tbl['std']) > 0:
            _sort_all(tbl, metric, asc)
            p = len(tbl['std'])
            for i, cutoff in enumerate(tbl[metric]):
                tp = sum([1 for oi in tbl['i'][:i + 1] if oi in tbl['std']])
                fp = sum([1 for oi in tbl['i'][:i + 1] \
                    if oi not in tbl['std']])
                fn = sum([1 for oi in tbl['i'][i + 1:] if oi in tbl['std']])
                tn = sum([1 for oi in tbl['i'][i + 1:] \
                    if oi not in tbl['std']])
                result['cutoff'].append(cutoff)
                result['spec'].append(tn / float(tn + fp))
                result['sens'].append(tp / float(tp + fn))
                result['prec'].append(tp / float(tp + fp))
                result['fdr'].append(fp / float(tp + fp))
            result['n'] = p
        return result

    def evaluate_scores(self, valids, stdltps):
        '''
        Calculates specificity, sensitivity, precision,
        false discovery rate as a function of critical
        values of each scores, for all LTPs with manual
        positive annotations, for all modes.
        Returns 4x embedded dicts of
        LTPs/modes/metrics/performance metrics.
        E.g. result['STARD10']['pos']['ktv']['sens']
        '''
        result = dict((ltp, {'pos': {}, 'neg': {}}) for ltp in stdltps)
        metrics = [
            ('Kendall\'s tau', 'ktv', False),
            ('Spearman corr.', 'spv', False),
            ('Pearson corr.', 'pev', False),
            ('Euclidean dist.', 'euv', True),
            ('Robust corr.', 'rcv', False),
            ('Goodman-Kruskal\'s gamma', 'gkv', False),
            ('Difference', 'dfv', True)
        ]
        for ltp in stdltps:
            for pos in ['pos', 'neg']:
                tbl = valids[ltp][pos]
                for m in metrics:
                    result[ltp][pos][m[1]] = \
                        spec_sens(valids, ltp, pos, m[1], m[2])
        return result

    def count_threshold_filter(self, valids, score, threshold, count = 10,
        threshold_type = 'fix', asc = True):
        '''
        Builds a boolean array whether the values of a score fall below
        or above certain critical value. The critical value can be defined
        as a fix value,
        or as a fraction of the minimum or maximum value of the score.
        Threshold type is either `fix`, `relative` or `best_fraction`.
        fix: values below or equal this number if lower is the better, otherwise
            values above or equal will be selected
        relative: threshold will be set the minimum value (if ordered ascending)
            or maximum value multiplied by the threshold
        best_fraction: opposite way as at the `relative`, at ascending order
            the maximum, at descending the minimum will be multiplied by the
            threshold
        count: the absolute maximum number of selected instances
        '''
        sort_alll(valids, score, asc = True)
        for ltp, d in valids.iteritems():
            for pn, tbl in d.iteritems():
                limScore = np.nanmin(tbl[score]) \
                    if asc and threshold_type =='relative' or \
                        not asc and threshold_type == 'best_fraction' \
                    else np.nanmean(tbl[score]) \
                    if threshold_type == 'mean_relative' \
                    else np.nanmedian(tbl[score]) \
                    if threshold_type == 'median_relative' \
                    else np.nanmax(tbl[score])
                _threshold = limScore * threshold \
                    if threshold_type == 'relative' or \
                        threshold_type == 'best_fraction' or \
                        threshold_type == 'mean_relative' or \
                        threshold_type == 'median_relative' \
                    else threshold
                cnt = 0
                boolArray = []
                for i in (xrange(len(tbl[score])) if asc \
                    else xrange(len(tbl[score])-1,-1,-1)):
                    if np.isnan(tbl[score][i]) or np.isinf(tbl[score][i]):
                        boolArray.append(False)
                        continue
                    if cnt > count or asc and tbl[score][i] > _threshold or \
                        not asc and tbl[score][i] < _threshold:
                        boolArray.append(False)
                    else:
                        cnt += 1
                        boolArray.append(True)
                tbl['bool_%s'%score] = np.array(boolArray)

    def scores_plot(self, valids, score = 'env', asc = True, 
        score_name = 'Euclidean distance', pdfname = None, singles = None,
        hlines = [3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 10.0],
        derivates = True):
        pdfname = 'scores_%s.pdf' % score if pdfname is None else pdfname
        font_family = 'Helvetica Neue LT Std'
        sns.set(font = font_family)
        sort_alll(valids, score, asc = asc)
        fig, axs = plt.subplots(8, 8, figsize = (20, 20))
        ltps = sorted(valids.keys())
        bool_score = 'bool_%s' % score
        prg = progress.Progress(len(ltps), 'Plotting scores', 1,
            percent = False)
        for i in xrange(len(ltps)):
            prg.step()
            ax = axs[i / 8, i % 8]
            ltpname = ltps[i]
            for pn in ['pos', 'neg']:
                col = '#CC0000' if pn == 'pos' else '#0000CC'
                scoreMin = np.nanmin(valids[ltpname][pn][score]) if asc \
                    else np.nanmax(valids[ltpname][pn][score])
                ax.plot(np.arange(len(valids[ltpname][pn][score])), 
                    valids[ltpname][pn][score], color = col,
                    alpha = 0.7, ls = '-',
                    linewidth = 0.3)
                if derivates:
                    xd = np.arange(1, len(valids[ltpname][pn][score]))
                    yd = np.array([(valids[ltpname][pn][score][j] - \
                            valids[ltpname][pn][score][j-1]) for j in xd])
                    yd = yd / np.nanmax(yd)
                    if not asc:
                        yd = -1 * yd
                    ax.plot(xd, yd, color = col, alpha = 0.2, ls = '-',
                        linewidth = 0.3)
                _xlim = ax.get_xlim()
                _ylim = ax.get_ylim()
                plt.setp(ax.xaxis.get_majorticklabels(), rotation = 90)
                best_ones = [j for j, b in \
                    enumerate(valids[ltpname][pn][bool_score]) if b]
                xbreak = np.nanmax(best_ones) if len(best_ones) > 0 else 0.0
                ybreak = valids[ltpname][pn][score][int(xbreak)]
                ax.plot([xbreak], [ybreak], marker = 'o', markersize = 2.0,
                    markerfacecolor = col, alpha = 0.7)
                ax.annotate(
                        '%u'%(xbreak), 
                        xy = (xbreak, ybreak), 
                        #xytext = (xbreak + (20.0 if pn == 'pos' else 0.0), 
                        #ybreak + (0.5 if pn == 'neg' else 0.0)), 
                        #xycoords = 'data',
                        #textcoords = 'offset points',
                        #ha = 'center', va = 'bottom', 
                        color = col,
                        alpha = 0.7,
                        fontsize = 'xx-small',
                        arrowprops = dict(arrowstyle = '-',
                            connectionstyle = 'arc,rad=.0',
                            color = col, edgecolor = col, alpha = 0.7, 
                            visible = True, linewidth = 0.2), 
                    )
                if type(hlines) is list:
                    for c in hlines:
                        ax.axhline(y = scoreMin*c, 
                            linewidth = 0.2, alpha = 0.5, color = col)
                ax.set_xlim(_xlim)
                ax.set_ylim(_ylim)
            ax.set_title(ltpname, color = '#CC0000' \
                if singles is not None and ltpname in singles else '#000000')
            ax.set_ylabel(score_name)
            ax.set_xlabel('Features (ordered %s)' % \
                'ascending' if asc else 'descending')
        fig.tight_layout()
        fig.savefig(pdfname)
        prg.terminate()
        plt.close()

    def fractions_barplot(self, samples, pprofs,
        pprofs_original, features = False,
        valids = None, highlight = False, highlight2 = False,
        all_features = True, pdfname = None):
        if pdfname is None:
            pdfname = 'protein_profiles%s.pdf' % \
                ('' if features is None else '_features')
        fracs = ['a9', 'a10', 'a11', 'a12', 'b1']
        font_family = 'Helvetica Neue LT Std'
        sns.set(font = font_family)
        fig, axs = plt.subplots(8, 8, figsize = (20, 20))
        ltps = sorted(samples.keys())
        prg = progress.Progress(len(ltps), 'Plotting profiles', 1,
            percent = False)
        for i in xrange(len(ltps)):
            prg.step()
            ax = axs[i / 8, i % 8]
            ltpname = ltps[i]
            ppr = np.array([pprofs[ltpname][fr] for fr in fracs])
            ppr_o = np.array([pprofs_original[ltpname][fr] for fr in fracs])
            if features:
                ppmax = np.nanmax(ppr_o)
                ppmin = np.nanmin(ppr_o)
                ppr = (ppr - ppmin) / (ppmax - ppmin)
                ppr[ppr < 0.0] = 0.0
                ppr_o = (ppr_o - ppmin) / (ppmax - ppmin)
            #B6B7B9 gray (not measured)
            #6EA945 mantis (protein sample)
            #007B7F teal (control/void)
            col = ['#6EA945' if s == 1 else \
                    '#B6B7B9' if s is None else \
                    '#007B7F' \
                for s in samples[ltpname][1:]]
            ax.bar(np.arange(len(ppr)), ppr_o, color = col, alpha = 0.1,
                edgecolor = 'none')
            ax.bar(np.arange(len(ppr)), ppr, color = col, edgecolor = 'none')
            if features and valids is not None:
                for pn in ['pos', 'neg']:
                    for fi, fe in enumerate(valids[ltpname][pn]['no']):
                        alpha = 0.20
                        lwd = 0.1
                        lst = '-'
                        plot_this = False
                        color = '#FFCCCC' if pn == 'pos' else '#CCCCFF'
                        try:
                            def _feg(self, fe):
                                for fei in fe:
                                    yield fei
                            feg = _feg(fe)
                            if highlight2 and \
                                valids[ltpname][pn][highlight2][fi]:
                                color = '#FF0000' if pn == 'pos' else '#0000FF'
                                alpha = 0.15
                                lwd = 0.4
                                lst = ':'
                                plot_this = True
                            if highlight and \
                                valids[ltpname][pn][highlight][fi]:
                                color = '#FF0000' if pn == 'pos' else '#0000FF'
                                alpha = 0.75
                                lwd = 0.4
                                lst = '-'
                                plot_this = True
                            if all_features or plot_this:
                                ax.plot(np.arange(len(ppr)) + 0.4,
                                    np.array([feg.next() \
                                        if s is not None else 0.0 \
                                        for s in samples[ltpname][1:]]),
                                    linewidth = lwd, markersize = 0.07,
                                    linestyle = lst, 
                                    color = color, alpha = alpha, marker = 'o')
                        except ValueError:
                            print 'Unequal length dimensions: %s, %s' % \
                                (ltpname, pn)
            ax.set_xticks(np.arange(len(ppr)) + 0.4)
            ax.set_xticklabels(fracs)
            ax.set_title('%s protein conc.'%ltpname)
        fig.tight_layout()
        fig.savefig(pdfname)
        prg.terminate()
        plt.close()

    # ltp.fractions_barplot(samples_upper, pprofs)

    def plot_score_performance(self, perf):
        metrics = [
            ('Kendall\'s tau', 'ktv', False),
            ('Spearman corr.', 'spv', False),
            ('Pearson corr.', 'pev', False),
            ('Euclidean dist.', 'euv', True),
            ('Robust corr.', 'rcv', False),
            ('Goodman-Kruskal\'s gamma', 'gkv', False),
            ('Difference', 'dfv', True)
        ]
        perfmetrics = [
            ('sens', '#6EA945', 'Sensitivity'),
            ('spec', '#FCCC06', 'Specificity'),
            ('prec', '#DA0025', 'Precision'),
            ('fdr', '#007B7F', 'FDR')
        ]
        font_family = 'Helvetica Neue LT Std'
        sns.set(font = font_family)
        # plt.gca().set_color_cycle(['red', 'green', 'blue', 'yellow'])
        for ltp, d1 in perf.iteritems():
            for pos, d2 in d1.iteritems():
                fig, axs = plt.subplots(len(metrics),
                    figsize = (10, 20), sharex = False)
                plt.suptitle('Performance of scores :: %s, %s' % \
                    (ltp, pos), size = 16)
                for i, m in enumerate(metrics):
                    for pm in perfmetrics:
                        axs[i].plot(d2[m[1]]['cutoff'], d2[m[1]][pm[0]],
                            '-', linewidth = 0.33,
                            color = pm[1], label = pm[2])
                    leg = axs[i].legend()
                    axs[i].set_ylabel('Performance\nmetrics')
                    axs[i].set_xlabel('%s value'%m[0])
                    axs[i].set_title(m[0])
                    if m[2]:
                        axs[i].set_xlim(axs[i].get_xlim()[::-1])
                fig.tight_layout()
                plt.subplots_adjust(top = 0.95)
                fig.savefig('score_performance_%s-%s.pdf'%(ltp, pos))
                plt.close(fig)

    def plot_roc(self, perf):
        metrics = [
            ('Kendall\'s tau', 'ktv', False),
            ('Spearman corr.', 'spv', False),
            ('Pearson corr.', 'pev', False),
            ('Euclidean dist.', 'euv', True),
            ('Robust corr.', 'rcv', False),
            ('Goodman-Kruskal\'s gamma', 'gkv', False),
            ('Difference', 'dfv', True)
        ]
        def colors(self, ):
            _colors = ['#6EA945', '#FCCC06', '#DA0025', 
                '#007B7F', '#454447', '#996A44']
            for c in _colors:
                yield c
        font_family = 'Helvetica Neue LT Std'
        sns.set(font = font_family)
        # plt.gca().set_color_cycle(['red', 'green', 'blue', 'yellow'])
        fig, axs = plt.subplots((len(metrics) + len(metrics) % 2) / 2, 2,
            figsize = (10, 20), sharex = False, sharey = False)
        plt.suptitle('Reciever operating characteristics', size = 16)
        for i, m in enumerate(metrics):
            c = colors()
            ax = axs[i/2][i%2]
            for ltp, d1 in perf.iteritems():
                for pos, d2 in d1.iteritems():
                    ax.plot(1 - np.array(d2[m[1]]['spec']),
                        np.array(d2[m[1]]['sens']),
                        '-', linewidth = 0.33, color = c.next(),
                        label = '%s-%s n = %u' % (ltp, pos, d2[m[1]]['n']))
            ax.plot([0,1], [0,1], '--', linewidth = 0.33, color = '#777777')
            leg = ax.legend(loc = 4)
            ax.set_ylabel('Sensitivity')
            ax.set_xlabel('1 - Specificity')
            ax.set_title('ROC :: %s'%m[0])
        fig.tight_layout()
        plt.subplots_adjust(top = 0.95)
        fig.savefig('roc.pdf')
        plt.close(fig)

    def best_combined(self, valids, scores, best = 10, ubiquity_treshold = 5):
        for ltp, d in valids.iteritems():
            for pn, tbl in d.iteritems():
                sorted_scores = []
                for sc in scores:
                    oi_sorted = np.copy(tbl['i'][tbl[sc[0]].argsort()])
                    if sc[1]:
                        sorted_scores.append(oi_sorted)
                    else:
                        sorted_scores.append(oi_sorted[::-1])
                all_ranks = np.empty((len(scores), len(tbl['i'])))
                for j, oi in enumerate(tbl['i']):
                    for i, scs in enumerate(sorted_scores):
                        for si, sc in enumerate(scs):
                            if oi == sc:
                                if tbl['cpv'][j] < 1.0 or \
                                    np.isnan(
                                        np.sum([tbl[s[0]][j] for s in scores])
                                    ) or (\
                                        'ubi' in tbl and \
                                        tbl['ubi'][j] > ubiquity_treshold
                                    ):
                                    all_ranks[i,j] = np.inf
                                else:
                                    all_ranks[i,j] = si
                combined_ranks = np.sum(all_ranks, axis = 0)
                best_oi = tbl['i'][np.argsort(combined_ranks)][:best]
                tbl['best%u'%best] = best_oi

    def best_gk_eu(self, valids, best = 10):
        best_combined(valids, [('gkv', False), ('euv', True)], best = best)

    def best_table(self, valids, fname, pos, best = 10):
        hdr = ['LTP', 'm/z', 'Database_ID', 'Level', 'Full_name',
               'Formula', 'Adduct', 'Database_m/z', 'Adduct_m/z']
        sort_alll(valids, 'mz')
        with open(fname, 'w') as f:
            f.write('\t'.join(hdr) + '\n')
            for ltp, d in valids.iteritems():
                tbl = d[pos]
                for oi in tbl['best%u'%best]:
                    i = np.where(tbl['i'] == oi)[0]
                    mz = tbl['mz'][i]
                    if tbl['lip'][oi] is not None:
                        for lip in tbl['lip'][oi]:
                            f.write('\t'.join([ltp, '%.08f'%mz] + \
                                [str(x) for x in list(lip)]) + '\n')
                    else:
                        f.write('\t'.join([ltp, '%.08f'%mz] + \
                            ['unknown']*6) + '\n')

    def true_positives(self, valids, stdpos, ltp, pos = 'pos',
        tolerance = 0.01):
        '''
        For one LTP having known binders looks up these
        among the valid features either in positive OR
        negative mode. (So you need to run twice, once
        for +, once for -.)
        '''
        ioffset = 0 if pos == 'pos' else 6
        valids[ltp][pos]['std'] = {}
        for fe in stdpos[ltp]:
            if fe[4 + ioffset] != '':
                mz = float(fe[ioffset + 4])
                iu = valids[ltp][pos]['mz'].searchsorted(mz)
                idx = []
                if iu < len(valids[ltp][pos]['mz']):
                    u = 0
                    while valids[ltp][pos]['mz'][iu + u] - mz <= tolerance:
                        idx.append(iu + u)
                        u += 1
                if iu > 0:
                    l = 1
                    while mz - valids[ltp][pos]['mz'][iu -l] <= tolerance:
                        idx.append(iu - l)
                        l += 1
                for i in idx:
                    identify_feature(valids[ltp][pos], fe, i, ioffset)

    def identify_feature(self, tbl, fe, i, ioffset):
        '''
        For one feature looks up if the lipid database ID
        and the adduct do match.
        Adds the matching record from the gold standard
        to lists in the `std` dict. There can be accessed
        by the original index.
        '''
        result = []
        oi = tbl['i'][i]
        if len(np.logical_and(tbl['lip'][oi][:,0] == fe[1 + ioffset], \
            tbl['lip'][oi][:,4] == fe[5 + ioffset])) > 0:
            if oi not in tbl['std']:
                tbl['std'][oi] = []
            tbl['std'][oi].append(fe)
        else:
            sys.stdout.write('Feature %.05f matched, lipid '\
                'ID hasn\'t been found: %s' % \
                (fe[4 + ioffset], fe[1 + ioffset]))
            sys.stdout.flush()

    def sortby_score(self, tbl, attr, asc = True):
        '''
        Returns sorted views of normalized features,
        m/z's and the attribute the sorting carried
        out by.
        '''
        sorted_no = tbl['no'][np.argsort(tbl[attr]),:]
        sorted_mz = tbl['mz'][np.argsort(tbl[attr])]
        sorted_attr = np.sort(tbl[attr])
        if asc:
            return sorted_no, sorted_mz, sorted_attr
        else:
            return sorted_no[::-1], sorted_mz[::-1], sorted_attr[::-1]

    def _sort_all(self, tbl, attr, asc = True):
        '''
        Sorts all arrays in one table by one specified 
        attribute.
        Either ascending or descending.
        '''
        ind = tbl[attr].argsort()
        dim = tbl[attr].shape[0]
        for k, a in tbl.iteritems():
            if k != attr and type(tbl[k]) == np.ndarray and \
                tbl[k].shape[0] == dim and not k.startswith('_'):
                if len(a.shape) == 1:
                    if asc:
                        tbl[k] = tbl[k][ind]
                    else:
                        tbl[k] = tbl[k][ind][::-1]
                else:
                    if asc:
                        tbl[k] = tbl[k][ind,:]
                    else:
                        tbl[k] = tbl[k][ind,:][::-1]
        tbl[attr].sort()
        if not asc:
            tbl[attr] = tbl[attr][::-1]

    def sort_alll(self, attr, asc = True):
        '''
        Sorts all arrays in all tables by one specified 
        attribute.
        Either ascending or descending.
        '''
        for d in self.valids.values():
            for tbl in d.values():
                self._sort_all(tbl, attr, asc)

    def get_scored_hits(self, data):
        hits = val_ubi_prf_rpr_hits(data, ubiquity = 70, profile_best = 50000)
        [v[0].shape[0] if v is not None else None \
            for vv in hits.values() for v in vv.values()]
        hits_upper = dict((l.upper(), {'pos': None, 'neg': None}) \
            for l in hits.keys())
        for l, d in hits.iteritems():
            for pn, tbl in d.iteritems():
                if hits_upper[l.upper()][pn] is None:
                    hits_upper[l.upper()][pn] = tbl
        return hits_upper

    '''
    MS1 lipid identification
    '''

    def lipid_lookup(self, stage0, runtime = None):
        '''
        Obtains full SwissLipids data
        '''
        pAdducts, nAdducts = get_swisslipids(
            adducts = ['[M+H]+', '[M+NH4]+', '[M-H]-'],
            formiate = True)
        lipids = find_lipids(stage0, pAdducts, nAdducts)
        if runtime:
            runtime = timeit.timeit(
                'find_lipids(final_hits, pAdducts, nAdducts)',
                setup = 'from __main__ import find_lipids, final_hits,'\
                'pAdducts, nAdducts', number = 1)
        return pAdducts, nAdducts, lipids, runtime

    def lipid_lookup_exact(self, verbose = False, outfile = None, charge = 1):
        '''
        Fetches data from SwissLipids and LipidMaps
        if not given.
        Looks up lipids based on exact masses (but
        taking into account the adducts).
        Writes hits into arrays in `lip` dict, having
        original indices as keys.
        Returns an array of lipid databases with exact
        masses.
        '''
        if self.exacts is None:
            self.get_swisslipids_exact()
            self.lipidmaps_exact()
        self.find_lipids_exact(verbose = verbose,
            outfile = outfile, charge = charge)

    def evaluate_results(self, stage0, stage2, lipids, samples_upper,
        letter = 'e'):
        '''
        Tracks the number of features/lipids along stages.
        Does some plotting.
        '''
        stage3 = [(
                l.upper(), 
                [len(i['neg']), len(i['pos'])], 
                len(stage2[l.upper()]), 
                [len(stage0[l]['neg'][0]), len(stage0[l]['pos'][0])], 
                np.nansum([x for x in samples_upper[l] if x is not None])
            ) if len(i.values()) > 0 \
            else (
                l.upper(), 0, 0, 0,
                np.nansum([x for x in samples_upper[l] if x is not None])
            ) \
            for l, i in lipids.iteritems()]
        #
        [sys.stdout.write(str(i) + '\n') for i in stage3]
        sum([i[2] > 0 for i in stage3])
        #
        font_family = 'Helvetica Neue LT Std'
        sns.set(font = font_family)
        fig, ax = plt.subplots()
        color = ['#FCCC06', '#6EA945', '#007B7F']
        for f in [1, 2, 3]:
            x1 = [i[1][0] for i in stage3 if i[-1] == f]
            x2 = [i[1][1] for i in stage3 if i[-1] == f]
            y = [i[2] for i in stage3 if i[-1] == f]
            p = plt.scatter(x1, y, marker = 's', c = color[f - 1],
                edgecolors = 'none', alpha = .5)
            p = plt.scatter(x2, y, marker = 'o', c = color[f - 1],
                edgecolors = 'none', alpha = .5)
            for xa, xb, yy in zip(x1, x2, y):
                p = plt.plot([xa, xb], [yy, yy], lw = .3, c = '#CCCCCC')
        plt.xlabel('Lipids (square: --; dot: +)')
        plt.ylabel('Lipids matching (square: -; dot: +)')
        fig.tight_layout()
        fig.savefig('hits-1-1000-%s.pdf'%letter)
        plt.close()
        #
        fig, ax = plt.subplots()
        color = ['#FCCC06', '#6EA945', '#007B7F']
        for f in [1, 2, 3]:
            x1 = [i[3][0] for i in stage3 if i[-1] == f]
            y1 = [i[1][0] for i in stage3 if i[-1] == f]
            x2 = [i[3][1] for i in stage3 if i[-1] == f]
            y2 = [i[1][1] for i in stage3 if i[-1] == f]
            p = plt.scatter(x1, y1, marker = 's', c = color[f - 1],
                edgecolors = 'none', alpha = .5)
            p = plt.scatter(x2, y2, marker = 'o', c = color[f - 1],
                edgecolors = 'none', alpha = .5)
            for xa, xb, ya, yb in zip(x1, x2, y1, y2):
                p = plt.plot([xa, xb], [ya, yb], lw = .3, c = '#CCCCCC')
        #
        plt.xlabel('Num of m/z`s (green: --; yellow: +)')
        plt.ylabel('Lipids (green: --; yellow: +)')
        fig.tight_layout()
        fig.savefig('hits-2-1000-%s.pdf'%letter)
        plt.close()
        # 
        fig, ax = plt.subplots()
        color = ['#FCCC06', '#6EA945', '#007B7F']
        for f in [1, 2, 3]:
            x1 = [i[3][0] for i in stage3 if i[-1] == f]
            x2 = [i[3][1] for i in stage3 if i[-1] == f]
            y = [i[2] for i in stage3 if i[-1] == f]
            p = plt.scatter(x1, y, marker = 's', c = color[f - 1],
                edgecolors = 'none', alpha = .5)
            p = plt.scatter(x2, y, marker = 'o', c = color[f - 1],
                edgecolors = 'none', alpha = .5)
            for xa, xb, yy in zip(x1, x2, y):
                p = plt.plot([xa, xb], [yy, yy], lw = .3, c = '#CCCCCC')
        plt.xlabel('Num of m/z`s (square: --; dot: +)')
        plt.ylabel('Lipids matching (square: -; dot: +)')
        fig.tight_layout()
        fig.savefig('hits-5-1000-%s.pdf'%letter)
        plt.close()
        #
        fig, ax = plt.subplots()
        y = flatList([i[3] for i in stage3 if type(i[3]) is not int])
        x = flatList([[i[4]]*2 for i in stage3 if type(i[3]) is not int])
        p = plt.scatter(x, y, c = '#6ea945', edgecolors = 'none')
        plt.xlabel('Fractions in sample')
        plt.ylabel('M/z min(pos, neg)')
        fig.tight_layout()
        fig.savefig('hits-3-150-%s.pdf'%letter)
        plt.close()
        #
        fig, ax = plt.subplots()
        y = [i[2] for i in stage3]
        x = [i[4] for i in stage3]
        p = plt.scatter(x, y, c = '#6ea945', edgecolors = 'none', alpha = .2)
        plt.xlabel('Fractions in sample')
        plt.ylabel('Lipids matched (pos, neg)')
        fig.tight_layout()
        fig.savefig('hits-4-1000-%s.pdf' % letter)
        plt.close()
        #
        fig, ax = plt.subplots()
        p = plt.bar(xrange(len(stage3)),
            [i[2] for i in sorted(stage3,
                key = lambda x: x[2], reverse = True)],
            0.5, color = '#6EA945', edgecolor = 'none')
        plt.xlabel('Lipid transfer protein')
        plt.ylabel('Number of valid lipid matches\n'\
            '(with no filtering based on profiles)')
        ax.set_xticks(np.arange(len(stage3)) + .25)
        ax.set_xticklabels([i[0] for i in \
                sorted(stage3, key = lambda x: x[2], reverse = True)],
            rotation = 90, fontsize = 5)
        fig.tight_layout()
        fig.savefig('lipid-matches-3000-%s.pdf'%letter)
        plt.close()

    def ms1_headgroups(self, verbose = False):
        '''
        Identifies headgroups by keywords and fatty acids
        from database record names.
        '''
        for ltp, d in self.valids.iteritems():
            for pn, tbl in d.iteritems():
                tbl['ms1hg'] = {}
                tbl['ms1fa'] = {}
                for oi, lips in tbl['lip'].iteritems():
                    tbl['ms1hg'][oi] = set([])
                    tbl['ms1fa'][oi] = {}
                    if lips is not None:
                        for lip in lips:
                            if lip[7] is not None:
                                if verbose:
                                    sys.stdout.write('\t:: %s-%s: '\
                                        'found %s\n' % \
                                        (ltp, pn, lip[7]))
                                    sys.stdout.flush()
                                posAdd = self.lipnames[lip[7]]['pos_adduct']
                                negAdd = self.lipnames[lip[7]]['neg_adduct']
                                thisModeAdd = \
                                    self.lipnames[lip[7]]['%s_adduct'%pn]
                                hg = lip[7]
                                fa = '%u:%u' % (lip[8][0], lip[8][1]) \
                                    if lip[8] is not None else None
                                if (posAdd is None and negAdd is None) or \
                                    thisModeAdd == lip[4]:
                                    if verbose:
                                        sys.stdout.write('\t\taccepting '\
                                            '%s-%s for'\
                                            ' %s-%s\n' % \
                                            (hg, lip[4], ltp, pn))
                                        sys.stdout.flush()
                                    tbl['ms1hg'][oi].add(hg)
                                    if fa is not None:
                                        if hg  not in tbl['ms1fa']:
                                            tbl['ms1fa'][oi][hg] = set([])
                                        tbl['ms1fa'][oi][hg].add(fa)
                                else:
                                    if verbose:
                                        sys.stdout.write(
                                            '\t\tdiscarding %s-%s'\
                                            ' for %s, %s\n'\
                                            ' in %s mode' % (
                                                lip[7],
                                                lip[4],
                                                ltp,
                                                    '%s is the main adduct' % \
                                                    thisModeAdd \
                                                    if thisModeAdd \
                                                        is not None \
                                                    else \
                                                        '%s does not'\
                                                        ' ionize' % \
                                                        lip[4],
                                                pn
                                            )
                                        )
                                        sys.stdout.flush()

    def headgroups_by_fattya(self, verbose = False):
        '''
        Limits the number of possible headgroups based on detected
        MS2 fatty acids.
        Creates dict `hgfa`.
        '''
        for ltp, d in self.valids.iteritems():
            for mod, tbl in d.iteritems():
                tbl['hgfa'] = {}
                for oi in tbl['i']:
                    tbl['hgfa'][oi] = set([])
                    if oi in tbl['ms2fas'] and tbl['ms2fas'][oi] is not None:
                        ms2fa = tbl['ms2fas'][oi]
                        for hg, ms1fa in tbl['ms1fa'][oi].iteritems():
                            if ms2fa in ms1fa:
                                tbl['hgfa'][oi].add(hg)
                                if verbose:
                                    sys.stdout.write('\t%s is among '\
                                        'possible fat'\
                                        'ty acids (%s) for %s based on MS1, '\
                                        'for feature %u at %s-%s \n' % \
                                        (ms2fa, ', '.join(list(ms1fa)), 
                                            hg, oi, ltp, mod))
                            else:
                                if verbose:
                                    sys.stdout.write('\t%s not among '\
                                        'possible fat'\
                                        'ty acids (%s) for %s based on MS1, '\
                                        'for feature %u at %s-%s \n' % \
                                        (ms2fa, ', '.join(list(ms1fa)), 
                                            hg, oi, ltp, mod))

    def identity_combined(self):
        '''
        Combined identification based on MS1 database lookup,
        MS2 headgroup fragments and MS2 fatty acids.
        Creates dicts `combined_hg` and `combined_fa`.
        '''
        for ltp, d in self.valids.iteritems():
            for mod, tbl in d.iteritems():
                tbl['combined_hg'] = {}
                tbl['combined_fa'] = {}
                for oi in tbl['i']:
                    hgs = set([])
                    tbl['combined_fa'][oi] = {}
                    if oi not in tbl['ms2hg'] or tbl['ms2hg'][oi] is None:
                        hgs = tbl['ms1hg'][oi]
                    else:
                        if len(tbl['ms1hg'][oi] & tbl['ms2hg'][oi]):
                            hgs = tbl['ms1hg'][oi] & tbl['ms2hg'][oi]
                        else:
                            hgs = tbl['ms1hg'][oi] | tbl['ms2hg'][oi]
                    if len(tbl['hgfa'][oi]) and len(hgs & tbl['hgfa'][oi]):
                        hgs = hgs & tbl['hgfa'][oi]
                    else:
                        hgs = tbl['hgfa'][oi]
                    tbl['combined_hg'][oi] = hgs
                    for hg in hgs:
                        if hg in tbl['ms1fa'][oi] and \
                            tbl['ms2fas'][oi] is not None and \
                            tbl['ms2fas'][oi] in tbl['ms1fa'][oi][hg]:
                            if hg not in tbl['combined_fa'][oi]:
                                tbl['combined_fa'][oi][hg] = set([])
                            # currently only the most certain cases:
                            tbl['combined_fa'][oi][hg].add(tbl['ms2fas'][oi])
                            # maybe later we need those with less evidence

    def identity_combined_ms2(self):
        '''
        Combined identification based on MS1 database lookup,
        MS2 headgroup fragments and MS2 fatty acids.
        Creates dicts `combined_hg` and `combined_fa`.
        '''
        for ltp, d in self.valids.iteritems():
            for mod, tbl in d.iteritems():
                tbl['combined_hg_ms2'] = {}
                for oi in tbl['i']:
                    tbl['combined_hg_ms2'][oi] = set([])
                    if oi in tbl['ms2hg']:
                        hgs = tbl['ms1hg'][oi] & tbl['ms2hg'][oi]
                    if len(tbl['hgfa'][oi]) and len(hgs & tbl['hgfa'][oi]):
                        hgs = hgs & tbl['hgfa'][oi]
                        tbl['combined_hg_ms2'][oi] = hgs

    def ms1_table(self, include = 'cl70pct'):
        ltps = sorted(self.valids.keys())
        # collecting primary and secondary column names
        # and cell values
        result = dict((ltp, {}) for ltp in ltps)
        colnames = {}
        for ltp in ltps:
            for pn, tbl in self.valids[ltp].iteritems():
                _np = 'pos' if pn == 'neg' else 'neg'
                nptbl = tbl[_np]
                for i, oi in enumerate(tbl['i']):
                    if tbl[include][i]:
                        if tbl['lip'][oi] is not None:
                            for lip in tbl['lip'][oi]:
                                hg = lip[7]
                                if hg is not None:
                                    fa = '%u:%u' % tuple(lip[8]) \
                                        if lip[8] is not None else 'unknown'
                                    # colnames
                                    if hg not in colnames:
                                        colnames[hg] = set([])
                                    colnames[hg].add(fa)
                                    # cells
                                    if hg not in result[ltp]:
                                        result[ltp][hg] = {}
                                    if fa not in result[ltp][hg]:
                                        # 3 bool values: +, -,
                                        # both + & - at same exact mass
                                        result[ltp][hg][fa] = \
                                            [False, False, False]
                                    # cell values
                                    if pn == 'pos':
                                        result[ltp][hg][fa][0] = True
                                    elif pn == 'neg':
                                        result[ltp][hg][fa][1] = True
                                    for _oi, pnlips in \
                                        tbl['%s_lip'%_np][oi].iteritems():
                                        if not result[ltp][hg][fa][2] and \
                                            pnlips is not None:
                                            for pnlip in pnlips:
                                                pfa = '%u:%u' % \
                                                    tuple(pnlip[12]) \
                                                    if pnlip[12] is not None \
                                                    else 'unknown'
                                                nfa = '%u:%u' % \
                                                    tuple(pnlip[13]) \
                                                    if pnlip[13] is not None \
                                                    else 'unknown'
                                                if pfa == fa and nfa == fa and \
                                                    pnlip[10] == hg and \
                                                    pnlip[11] == hg:
                                                    result[ltp][hg][fa][2] = \
                                                        True
                                                    break
        return colnames, result
    
    def today(self):
        return datetime.date.today().strftime(r'%Y-%m-%d')
    
    def ms1_table_html(self, filename = None, include = 'cl70pct'):
        filename = 'results_%s_ms1.html' % self.today() \
            if filename is None \
            else filename
        colnames, ms1tab = self.ms1_table(include = include)
        title = 'Binding specificities of LTPs detected in MS1'
        table = ''
        tablerow = '\t\t<tr>\n%s\t\t</tr>\n'
        tablehcell = '\t\t\t<th>\n\t\t\t\t%s\n\t\t\t</th>\n'
        tableshcell = '\t\t\t<th colspan="%u">\n\t\t\t\t%s\n\t\t\t</th>\n'
        tablecell = '\t\t\t<td class="%s" title="%s">\n\t\t\t\t%s'\
            '\n\t\t\t</td>\n'
        th1 = tablehcell % ''
        th2 = tablehcell % 'LTP'
        for hg in sorted(colnames.keys()):
            th1 += tableshcell % (len(colnames[hg]), hg)
            for fa in sorted(colnames[hg]):
                th2 += tablehcell % fa
        table += tablerow % th1
        table += tablerow % th2
        for ltp in sorted(self.valids.keys()):
            row = tablecell % ('rowname', ltp, ltp)
            for hg in sorted(colnames.keys()):
                for fa in sorted(colnames[hg]):
                    if hg in ms1tab[ltp] and fa in ms1tab[ltp][hg]:
                        if ms1tab[ltp][hg][fa][2]:
                            row += tablecell % ('matching', 
                                '%s (%s) detected in '\
                                'Positive & Negative modes at %s' % \
                                    (hg, fa, ltp),
                                '')
                        elif ms1tab[ltp][hg][fa][0] and \
                            ms1tab[ltp][hg][fa][1]:
                            row += tablecell % ('both', '%s (%s) detected in '\
                                'Positive & Negative modes at %s' % \
                                    (hg, fa, ltp),
                                '')
                        elif ms1tab[ltp][hg][fa][0]:
                            row += tablecell % ('positive',
                                '%s (%s) detected '\
                                'in Positive mode at %s' % (hg, fa, ltp), '')
                        elif ms1tab[ltp][hg][fa][1]:
                            row += tablecell % ('negative',
                                '%s (%s) detected '\
                                'in Negative mode at %s' % (hg, fa, ltp), '')
                    else:
                        row += tablecell % ('empty',
                            '%s (%s) not detected at %s' % (hg, fa, ltp), '')
            table += tablerow % row
        with open(filename, 'w') as f:
            f.write(self.html_table_template % (title, title, table))

    def ms1_table_html_simple(self, filename = None, include = 'cl70pct'):
        '''
        Outputs a HTML table LTPs vs lipid classes (headgroups)
        based on MS1 identifications.
        '''
        filename = 'results_%s_ms1hg.html' % self.today() \
            if filename is None \
            else filename
        colnames, ms1tab = self.ms1_table(include = include)
        title = 'Binding specificities of LTPs by headgroups detected in MS1'
        table = ''
        tablerow = '\t\t<tr>\n%s\t\t</tr>\n'
        tablehcell = '\t\t\t<th>\n\t\t\t\t%s\n\t\t\t</th>\n'
        tablecell = '\t\t\t<td class="%s" title="%s">\n\t\t\t\t%s'\
            '\n\t\t\t</td>\n'
        th1 = tablehcell % 'LTP'
        for hg in sorted(colnames.keys()):
            th1 += tablehcell % hg
        table += tablerow % th1
        for ltp in sorted(self.valids.keys()):
            row = tablecell % ('rowname', ltp, ltp)
            for hg in sorted(colnames.keys()):
                pos_neg = False
                pos = False
                neg = False
                for fa in sorted(colnames[hg]):
                    if hg in ms1tab[ltp] and fa in ms1tab[ltp][hg]:
                        if ms1tab[ltp][hg][fa][2]:
                            pos_neg = True
                        elif ms1tab[ltp][hg][fa][0] and \
                            ms1tab[ltp][hg][fa][1]:
                            pos_neg = True
                        elif ms1tab[ltp][hg][fa][0]:
                            pos = True
                        elif ms1tab[ltp][hg][fa][1]:
                            neg = True
                if pos_neg:
                    row += tablecell % ('matching', 
                        '%s detected in Positive & '\
                        'Negative modes' % hg, '')
                elif pos:
                    row += tablecell % ('positive', '%s detected in Positive '\
                        'mode' % hg, '')
                elif neg:
                    row += tablecell % ('negative', '%s detected in Negative '\
                        'mode' % hg, '')
                else:
                    row += tablecell % ('empty', '%s not detected' % hg, '')
            table += tablerow % row
        with open(filename, 'w') as f:
            f.write(self.html_table_template % (title, title, table))
    
    def ms1_table_latex_simple(self, filename = 'ms1headgroups.tex', include = 'cl70pct', break_half = True):
        '''
        Outputs a LaTeX table LTPs vs lipid classes (headgroups)
        based on MS1 identifications.
        '''
        colnames, ms1tab = self.ms1_table(include = include)
        table = '\\begin{tabular}{l%s}\n' % ('l' * len(colnames))
        tablerow = '\t%s\\\\\n'
        hdr = '& '
        hdr += ' & '.join(
            sorted(
                map(
                    lambda hg:
                        '\\rot{%s}' % hg,
                    colnames.keys()
                )
            )
        )
        table += tablerow % hdr
        for i, ltp in enumerate(sorted(self.valids.keys())):
            row = [ltp]
            for hg in sorted(colnames.keys()):
                pos_neg = False
                pos = False
                neg = False
                for fa in sorted(colnames[hg]):
                    if hg in ms1tab[ltp] and fa in ms1tab[ltp][hg]:
                        if ms1tab[ltp][hg][fa][2]:
                            pos_neg = True
                        elif ms1tab[ltp][hg][fa][0] and \
                            ms1tab[ltp][hg][fa][1]:
                            pos_neg = True
                        elif ms1tab[ltp][hg][fa][0]:
                            pos = True
                        elif ms1tab[ltp][hg][fa][1]:
                            neg = True
                if pos_neg:
                    row.append('\\cellcolor{emblyellow!75}')
                elif pos:
                    row.append('\\cellcolor{emblgreen!75}')
                elif neg:
                    row.append('\\cellcolor{emblpetrol!75}')
                else:
                    row.append('')
            table += tablerow % ' & '.join(row)
            if break_half and i == np.ceil(len(self.valids) / 2.0):
                table += '\\end{tabular}\n\quad\n'
                table += '\\begin{tabular}{l%s}\n' % ('l' * len(colnames))
                table += tablerow % hdr
        table += '\\end{tabular}\n'
        with open(filename, 'w') as f:
            f.write(table)

    '''
    END: MS1 lipid identification
    '''

    def ms2_table_html_simple(self, filename = None, include = 'cl70pct'):
        '''
        Outputs a HTML table LTPs vs lipid classes (headgroups)
        based on MS2 identifications.
        '''
        filename = 'results_%s_ms2hg.html' % self.today() \
            if filename is None \
            else filename
        title = 'Binding specificities of LTPs by headgroups detected in MS2'
        table = ''
        tablerow = '\t\t<tr>\n%s\t\t</tr>\n'
        tablehcell = '\t\t\t<th>\n\t\t\t\t%s\n\t\t\t</th>\n'
        tablecell = '\t\t\t<td class="%s" title="%s">\n\t\t\t\t%s'\
            '\n\t\t\t</td>\n'
        th1 = tablehcell % 'LTP'
        colnames = set([])
        for ltp, d in self.valids.iteritems():
            for pn, tbl in d.iteritems():
                for hg in tbl['ms2hg'].values():
                    if hg is not None:
                        colnames = colnames | hg
        colnames = sorted(list(colnames))
        for hg in colnames:
            th1 += tablehcell % hg
        table += tablerow % th1
        for ltp in sorted(self.valids.keys()):
            row = tablecell % ('rowname', ltp, ltp)
            for hg in colnames:
                pos_neg_same = False
                pos_neg = False
                pos = False
                neg = False
                pos_unambig = False
                neg_unambig = False
                pos_neg_same_unambig = False
                pos_neg_unambig = False
                for pn, tbl in self.valids[ltp].iteritems():
                    for ms2hg in tbl['ms2hg'].values():
                        if ms2hg is not None and hg in ms2hg:
                            if pn == 'pos':
                                pos = True
                                if len(ms2hg) == 1:
                                    pos_unambig = True
                            elif pn == 'neg':
                                neg = True
                                if len(ms2hg) == 1:
                                    neg_unambig = True
                if pos and neg:
                    pos_neg = True
                if pos_unambig or neg_unambig:
                    pos_neg_unambig = True
                for ms2hgs in self.valids[ltp]['pos']['ms2hg_neg'].values():
                    for ms2hg in ms2hgs.values():
                        if hg in ms2hg:
                            pos_neg_same = True
                            if len(ms2hg) == 1:
                                pos_neg_unambig = True
                if pos_neg_same:
                    unambig = 'UA' if pos_neg_same_unambig else 'A'
                    unambig2 = '\nUnambiguous at least once' \
                        if pos_neg_same_unambig \
                        else '\nOnly ambiguous'
                    row += tablecell % ('matching', 'Detected in Positive &'\
                        ' Negative modes,\nat same exact mass%s'%\
                        unambig2, unambig)
                elif pos_neg:
                    unambig = 'UA' if pos_neg_unambig else 'A'
                    unambig2 = '\nUnambiguous at least once' \
                        if pos_neg_unambig \
                        else '\nOnly ambiguous'
                    row += tablecell % ('both', 'Detected in Positive &'\
                        ' Negative modes,\nat different exact mass%s'%\
                        unambig2, unambig)
                elif pos:
                    unambig2 = '\nUnambiguous at least once' if pos_unambig \
                        else '\nOnly ambiguous'
                    unambig = 'UA' if pos_unambig else 'A'
                    row += tablecell % ('positive',
                        'Detected in Positive mode%s' % \
                        unambig2, unambig)
                elif neg:
                    unambig2 = '\nUnambiguous at least once' if neg_unambig \
                        else '\nOnly ambiguous'
                    unambig = 'UA' if neg_unambig else 'A'
                    row += tablecell % ('negative',
                        'Detected in Negative mode%s'%\
                        unambig2, unambig)
                else:
                    row += tablecell % ('empty', 'Not detected', '')
            table += tablerow % row
        with open(filename, 'w') as f:
            f.write(self.html_table_template % (title, title, table))

    def ms2_table_latex_simple(self, filename = 'ms2headgroups.tex', include = 'cl70pct', break_half = True):
        '''
        Outputs a LaTeX table LTPs vs lipid classes (headgroups)
        based on MS2 identifications.
        '''
        colnames = set([])
        for ltp, d in self.valids.iteritems():
            for pn, tbl in d.iteritems():
                for hg in tbl['ms2hg'].values():
                    if hg is not None:
                        colnames = colnames | hg
        tablerow = '\t%s\\\\\n'
        colnames.remove('')
        colnames = sorted(list(colnames))
        table = '\\begin{tabular}{l%s}\n' % ('l' * len(colnames))
        hdr = '\t& '
        hdr += ' & '.join(
            map(
                lambda hg:
                    '\\rot{%s}' % hg,
                colnames
            )
        )
        table += tablerow % hdr
        for i, ltp in enumerate(sorted(self.valids.keys())):
            row = [ltp]
            for hg in colnames:
                pos_neg_same = False
                pos_neg = False
                pos = False
                neg = False
                pos_unambig = False
                neg_unambig = False
                pos_neg_same_unambig = False
                pos_neg_unambig = False
                for pn, tbl in self.valids[ltp].iteritems():
                    for ms2hg in tbl['ms2hg'].values():
                        if ms2hg is not None and hg in ms2hg:
                            if pn == 'pos':
                                pos = True
                                if len(ms2hg) == 1:
                                    pos_unambig = True
                            elif pn == 'neg':
                                neg = True
                                if len(ms2hg) == 1:
                                    neg_unambig = True
                if pos and neg:
                    pos_neg = True
                if pos_unambig or neg_unambig:
                    pos_neg_unambig = True
                for ms2hgs in self.valids[ltp]['pos']['ms2hg_neg'].values():
                    for ms2hg in ms2hgs.values():
                        if hg in ms2hg:
                            pos_neg_same = True
                            if len(ms2hg) == 1:
                                pos_neg_unambig = True
                if pos_neg_same or pos_neg:
                    row.append('\\cellcolor{emblyellow!75}')
                elif pos:
                    row.append('\\cellcolor{emblgreen!75}')
                elif neg:
                    row.append('\\cellcolor{emblpetrol!75}')
                else:
                    row.append('')
            table += tablerow % ' & '.join(row)
            if break_half and i == np.ceil(len(self.valids) / 2.0):
                table += '\\end{tabular}\n\quad\n'
                table += '\\begin{tabular}{l%s}\n' % ('l' * len(colnames))
                table += tablerow % hdr
        table += '\\end{tabular}\n'
        with open(filename, 'w') as f:
            f.write(table)

    def ms1_ms2_table_html_simple(self, filename = None, include = 'cl70pct'):
        '''
        Outputs a HTML table LTPs vs lipid classes (headgroups)
        based on MS1 and MS2 identifications.
        '''
        filename = 'results_%s_ms1ms2hg.html' % self.today() \
            if filename is None \
            else filename
        title = 'Binding specificities of LTPs by headgroups'\
            ' detected in MS1 and MS2'
        table = ''
        tablerow = '\t\t<tr>\n%s\t\t</tr>\n'
        tablehcell = '\t\t\t<th>\n\t\t\t\t%s\n\t\t\t</th>\n'
        tablecell = '\t\t\t<td class="%s" title="%s">\n\t\t\t\t%s'\
            '\n\t\t\t</td>\n'
        th1 = tablehcell % 'LTP'
        colnames = set([])
        for ltp, d in self.valids.iteritems():
            for pn, tbl in d.iteritems():
                for hgs in tbl['identity'].values():
                    for hg, ids in hgs.iteritems():
                        if hg is not None and \
                            (ids['ms2_pos'] or ids['ms2_neg']):
                            colnames.add(hg)
        colnames = sorted(list(colnames))
        # header row (lipid species)
        for hg in colnames:
            th1 += tablehcell % hg
        table += tablerow % th1
        # rows by LTP
        for ltp in sorted(self.valids.keys()):
            row = tablecell % ('rowname', ltp, ltp)
            for hg in colnames:
                pos = False
                neg = False
                pos_neg = False
                pos_unambig = False
                neg_unambig = False
                pos_neg_same_unambig = False
                pos_neg_unambig = False
                pos_neg_same = False
                tbl = self.valids[ltp]['pos']
                for oi in tbl['i'][np.where(tbl[include])[0]]:
                    if hg in tbl['identity'][oi]:
                        this_hg = tbl['identity'][oi][hg]
                        if this_hg['ms1_pos'] and this_hg['ms2_pos']:
                            pos = True
                            if this_hg['ms1_neg'] and this_hg['ms2_neg']:
                                pos_neg_same = True
                                for noi in tbl['neg'][oi].keys():
                                    if hg in valids[ltp]['neg']\
                                        ['identity'][noi]:
                                        if self.valids[ltp]['neg']['identity']\
                                            [noi][hg]['ms1_neg'] \
                                                and \
                                            self.valids[ltp]['neg']['identity']\
                                            [noi][hg]['ms2_neg']:
                                            pos_neg = True
                                            if sum(map(lambda (_hg, this_hg):
                                                    _hg != hg \
                                                            and \
                                                        this_hg['ms1_neg'] \
                                                            and \
                                                        this_hg['ms2_neg'],
                                                    valids[ltp]['neg']\
                                                        ['identity']\
                                                        [noi].iteritems()
                                                )) == 0:
                                                pos_neg_same_unambig = True
                            if sum(map(lambda (_hg, this_hg):
                                    _hg != hg and this_hg['ms1_pos'] \
                                        and this_hg['ms2_pos'],
                                    tbl['identity'][oi].iteritems()
                                )) == 0:
                                pos_unambig = True
                tbl = self.valids[ltp]['neg']
                for oi in tbl['i'][np.where(tbl[include])[0]]:
                    if hg in tbl['identity'][oi]:
                        this_hg = tbl['identity'][oi][hg]
                        if this_hg['ms1_neg'] and this_hg['ms2_neg']:
                            neg = True
                            if this_hg['ms1_pos'] and this_hg['ms2_pos']:
                                pos_neg_same = True
                                for noi in tbl['pos'][oi].keys():
                                    if hg in self.valids[ltp]['pos']\
                                        ['identity'][noi]:
                                        if valids[ltp]['pos']['identity']\
                                            [noi][hg]['ms1_pos'] and \
                                            self.valids[ltp]['pos']['identity']\
                                                [noi][hg]['ms2_pos']:
                                            pos_neg = True
                                            if sum(map(lambda (_hg, this_hg):
                                                    _hg != hg and \
                                                        this_hg['ms1_pos'] \
                                                            and \
                                                        this_hg['ms2_pos'],
                                                    valids[ltp]['pos']\
                                                        ['identity']\
                                                        [noi].iteritems()
                                                )) == 0:
                                                pos_neg_same_unambig = True
                            if sum(map(lambda (_hg, this_hg):
                                    _hg != hg and this_hg['ms1_neg'] \
                                        and this_hg['ms2_neg'],
                                    tbl['identity'][oi].iteritems()
                                )) == 0:
                                neg_unambig = True
                if pos_unambig and neg_unambig:
                    pos_neg_unambig = True
                if pos_neg_same:
                    unambig = 'UA' if pos_neg_same_unambig else 'A'
                    unambig2 = '\nUnambiguous at least once' \
                        if pos_neg_same_unambig \
                        else '\nOnly ambiguous'
                    row += tablecell % ('matching', '%s detected in Positive &'\
                        ' Negative modes,\nat same exact mass%s'%\
                        (hg, unambig2), unambig)
                elif pos_neg:
                    unambig = 'UA' if pos_neg_unambig else 'A'
                    unambig2 = '\nUnambiguous at least once' \
                        if pos_neg_unambig \
                        else '\nOnly ambiguous'
                    row += tablecell % ('both', '%s detected in Positive &'\
                        ' Negative modes,\nat different exact mass%s' % \
                        (hg, unambig2), unambig)
                elif pos:
                    unambig2 = '\nUnambiguous at least once' if pos_unambig \
                        else '\nOnly ambiguous'
                    unambig = 'UA' if pos_unambig else 'A'
                    row += tablecell % ('positive',
                        '%s detected in Positive mode%s' % \
                        (hg, unambig2), unambig)
                elif neg:
                    unambig2 = '\nUnambiguous at least once' if neg_unambig \
                        else '\nOnly ambiguous'
                    unambig = 'UA' if neg_unambig else 'A'
                    row += tablecell % ('negative',
                        '%s detected in Negative mode%s' % \
                        (hg, unambig2), unambig)
                else:
                    row += tablecell % ('empty', '%s not detected' % hg, '')
            table += tablerow % row
        with open(filename, 'w') as f:
            f.write(self.html_table_template % (title, title, table))
    
    def ms1_ms2_table_latex_simple(self, filename = 'ms1ms2headgroups.tex',
        include = 'cl70pct', break_half = True):
        '''
        Outputs a LaTeX table LTPs vs lipid classes (headgroups)
        based on MS1 and MS2 identifications.
        '''
        colnames = set([])
        for ltp, d in self.valids.iteritems():
            for pn, tbl in d.iteritems():
                for hgs in tbl['identity'].values():
                    for hg, ids in hgs.iteritems():
                        if hg is not None and \
                            (ids['ms2_pos'] or ids['ms2_neg']):
                            colnames.add(hg)
        colnames.remove('')
        colnames = sorted(list(colnames))
        tablerow = '\t%s\\\\\n'
        table = '\\begin{tabular}{l%s}\n' % ('l' * len(colnames))
        hdr = '\t& '
        hdr += ' & '.join(
            map(
                lambda hg:
                    '\\rot{%s}' % hg,
                colnames
            )
        )
        table += tablerow % hdr
        # rows by LTP
        for i, ltp in enumerate(sorted(self.valids.keys())):
            row = [ltp]
            for hg in colnames:
                pos = False
                neg = False
                pos_neg = False
                pos_unambig = False
                neg_unambig = False
                pos_neg_same_unambig = False
                pos_neg_unambig = False
                pos_neg_same = False
                tbl = self.valids[ltp]['pos']
                for oi in tbl['i'][np.where(tbl[include])[0]]:
                    if hg in tbl['identity'][oi]:
                        this_hg = tbl['identity'][oi][hg]
                        if this_hg['ms1_pos'] and this_hg['ms2_pos']:
                            pos = True
                            if this_hg['ms1_neg'] and this_hg['ms2_neg']:
                                pos_neg_same = True
                                for noi in tbl['neg'][oi].keys():
                                    if hg in valids[ltp]['neg']\
                                        ['identity'][noi]:
                                        if self.valids[ltp]['neg']['identity']\
                                            [noi][hg]['ms1_neg'] \
                                                and \
                                            self.valids[ltp]['neg']['identity']\
                                            [noi][hg]['ms2_neg']:
                                            pos_neg = True
                                            if sum(map(lambda (_hg, this_hg):
                                                    _hg != hg \
                                                            and \
                                                        this_hg['ms1_neg'] \
                                                            and \
                                                        this_hg['ms2_neg'],
                                                    valids[ltp]['neg']\
                                                        ['identity']\
                                                        [noi].iteritems()
                                                )) == 0:
                                                pos_neg_same_unambig = True
                            if sum(map(lambda (_hg, this_hg):
                                    _hg != hg and this_hg['ms1_pos'] \
                                        and this_hg['ms2_pos'],
                                    tbl['identity'][oi].iteritems()
                                )) == 0:
                                pos_unambig = True
                tbl = self.valids[ltp]['neg']
                for oi in tbl['i'][np.where(tbl[include])[0]]:
                    if hg in tbl['identity'][oi]:
                        this_hg = tbl['identity'][oi][hg]
                        if this_hg['ms1_neg'] and this_hg['ms2_neg']:
                            neg = True
                            if this_hg['ms1_pos'] and this_hg['ms2_pos']:
                                pos_neg_same = True
                                for noi in tbl['pos'][oi].keys():
                                    if hg in self.valids[ltp]['pos']\
                                        ['identity'][noi]:
                                        if valids[ltp]['pos']['identity']\
                                            [noi][hg]['ms1_pos'] and \
                                            self.valids[ltp]['pos']['identity']\
                                                [noi][hg]['ms2_pos']:
                                            pos_neg = True
                                            if sum(map(lambda (_hg, this_hg):
                                                    _hg != hg and \
                                                        this_hg['ms1_pos'] \
                                                            and \
                                                        this_hg['ms2_pos'],
                                                    valids[ltp]['pos']\
                                                        ['identity']\
                                                        [noi].iteritems()
                                                )) == 0:
                                                pos_neg_same_unambig = True
                            if sum(map(lambda (_hg, this_hg):
                                    _hg != hg and this_hg['ms1_neg'] \
                                        and this_hg['ms2_neg'],
                                    tbl['identity'][oi].iteritems()
                                )) == 0:
                                neg_unambig = True
                if pos_unambig and neg_unambig:
                    pos_neg_unambig = True
                if pos_neg_unambig:
                    row.append('\\cellcolor{emblyellow!75}')
                elif pos_neg:
                    row.append('\\cellcolor{emblyellow!75}')
                elif pos:
                    row.append('\\cellcolor{emblgreen!75}')
                elif neg:
                    row.append('\\cellcolor{emblpetrol!75}')
                else:
                    row.append('')
            table += tablerow % ' & '.join(row)
            if break_half and i == np.ceil(len(self.valids) / 2.0):
                table += '\\end{tabular}\n\quad\n'
                table += '\\begin{tabular}{l%s}\n' % ('l' * len(colnames))
                table += tablerow % hdr
        table += '\\end{tabular}\n'
        with open(filename, 'w') as f:
            f.write(table)
    
    def identities_latex_table(self, data,
        filename = 'ms1ms2hcheadgroups.tex', break_half = True):
        allhgs = set([])
        for protein, d in data.iteritems():
            for hgs in d.values():
                allhgs = allhgs | hgs
        colnames = sorted(list(allhgs))
        tablerow = '\t%s\\\\\n'
        table = '\\begin{tabular}{l%s}\n' % ('l' * len(colnames))
        hdr = '\t& '
        hdr += ' & '.join(
            map(
                lambda hg:
                    '\\rot{%s}' % hg,
                colnames
            )
        )
        table += tablerow % hdr
        for i, protein in enumerate(sorted(data.keys())):
            row = [protein]
            for hg in colnames:
                if hg in data[protein]['pos'] and \
                    hg in data[protein]['neg']:
                    row.append('\\cellcolor{emblyellow!75}')
                elif hg in data[protein]['neg']:
                    row.append('\\cellcolor{emblpetrol!75}')
                elif hg in data[protein]['pos']:
                    row.append('\\cellcolor{emblgreen!75}')
                else:
                    row.append('')
            table += tablerow % ' & '.join(row)
            if break_half and i == np.ceil(len(self.valids) / 2.0):
                table += '\\end{tabular}\n\quad\n'
                table += '\\begin{tabular}{l%s}\n' % ('l' * len(colnames))
                table += tablerow % hdr
        table += '\\end{tabular}\n'
        with open(filename, 'w') as f:
            f.write(table)
    
    def feature_identity_table(self):
        '''
        Creates dictionaries named `identity`, having
        original IDs as keys and 4 element dictionaries
        as values with keys ms1_pos, ms2_pos, ms1_neg, ms2_neg,
        each having a boolean value.
        '''
        self.sort_alll('mz')
        for ltp, d in self.valids.iteritems():
            for pn, tbl in d.iteritems():
                opp_mode = 'neg' if pn == 'pos' else 'pos'
                tbl['identity'] = {}
                for oi in tbl['i']:
                    tbl['identity'][oi] = {}
                    ms1 = set([]) if tbl['ms1hg'][oi] is None \
                        else tbl['ms1hg'][oi]
                    ms2 = set([]) \
                        if oi not in tbl['ms2hg'] or tbl['ms2hg'][oi] is None \
                        else tbl['ms2hg'][oi]
                    ms1_opp = set([])
                    ms2_opp = set([])
                    for opp_oi in tbl[opp_mode][oi].keys():
                        if opp_oi in d[opp_mode]['ms1hg']:
                            ms1_opp = ms1_opp | d[opp_mode]['ms1hg'][opp_oi]
                        if opp_oi in d[opp_mode]['ms2hg'] and \
                            d[opp_mode]['ms2hg'][opp_oi] is not None:
                            ms2_opp = ms2_opp | d[opp_mode]['ms2hg'][opp_oi]
                    hg_all = ms1 | ms2 | ms1_opp | ms2_opp
                    for hg in hg_all:
                        tbl['identity'][oi][hg] = {
                            'ms1_%s'%pn: hg in ms1,
                            'ms2_%s'%pn: hg in ms2,
                            'ms1_%s'%opp_mode: hg in ms1_opp,
                            'ms2_%s'%opp_mode: hg in ms2_opp
                        }

    def feature_identity_html_table(self, valids, bindprop,
        fits_pprop = 'cl5pct',
        outf = 'identities.html'):
        hdr = ['LTP', 'm/z', 'mode', 'fits protein +', 'fits protein -', 
            'headgroup', 'MS1+', 'MS2+', 'MS1-', 'MS2-']
        title = 'Identities for all features'
        table = ''
        tablerow = '\t\t<tr>\n%s\t\t</tr>\n'
        tablehcell = '\t\t\t<th>\n\t\t\t\t%s\n\t\t\t</th>\n'
        tablecell = '\t\t\t<td class="%s" title="%s">\n\t\t\t\t%s'\
            '\n\t\t\t</td>\n'
        hrow = ''
        for coln in hdr:
            hrow += tablehcell % coln
        hrow = tablerow % hrow
        table = hrow
        empty_ids = {'': {'ms1_pos': False, 'ms2_pos': False, 
            'ms1_neg': False, 'ms2_neg': False}}
        for ltp, d in valids.iteritems():
            for pn, tbl in d.iteritems():
                _np = 'pos' if pn == 'neg' else 'neg'
                for i, oi in enumerate(tbl['i']):
                    for hg, ids in (tbl['identity'][oi].iteritems() \
                        if len(tbl['identity'][oi]) > 0 \
                        else empty_ids.iteritems()):
                        this_row = ''
                        this_row += tablecell % ('rowname', '', ltp)
                        this_row += tablecell % \
                            ('nothing', '', '%.05f'%tbl['mz'][i])
                        this_row += tablecell % ('nothing', '', 
                            '+' if pn == 'pos' else '-')
                        fits_pos = tbl[fits_pprop][i] if pn == 'pos' else \
                            bool(sum(map(lambda (i, v): v,
                                    filter(lambda (i, v): 
                                        d['pos']['i'][i] in tbl['pos'][oi], 
                                        enumerate(d['pos'][fits_pprop])
                                    )
                                )
                            ))
                        fits_neg = tbl[fits_pprop][i] if pn == 'neg' else \
                            bool(sum(map(lambda (i, v): v,
                                    filter(lambda (i, v): 
                                        d['neg']['i'][i] in tbl['neg'][oi], 
                                        enumerate(d['neg'][fits_pprop])
                                    )
                                )
                            ))
                        this_row += tablecell % (
                            'positive' if fits_pos else 'nothing',
                            'Positive mode fits well on protein profile' \
                                if fits_pos \
                                else 'Positive mode does not fit well on'\
                                    ' protein profile or not available',
                            '')
                        this_row += tablecell % (
                            'positive' if fits_neg else 'nothing',
                            'Negative mode fits well on protein profile' \
                                if fits_neg \
                                else 'Negative mode does not fit well on'\
                                    ' protein profile or not available',
                            '')
                        this_row += tablecell % (
                            'positive' if hg in bindprop[ltp] else 'nothing',
                            '%s is known binder of %s' % (hg, ltp) \
                                if hg in bindprop[ltp] \
                                else 'No literature data '\
                                    'about %s binding %s' % \
                                    (ltp, hg),
                            hg)
                        this_row += tablecell % (
                            'positive' if ids['ms1_pos'] else 'nothing',
                            'Identified in MS1 positive mode' \
                                if ids['ms1_pos'] \
                                else 'Not identified in MS1 positive mode',
                            '')
                        this_row += tablecell % (
                            'positive' if ids['ms2_pos'] else 'nothing',
                            'Identified in MS2 positive mode' \
                                if ids['ms2_pos'] \
                                else 'Not identified in MS2 positive mode',
                            '')
                        this_row += tablecell % (
                            'positive' if ids['ms1_neg'] else 'nothing',
                            'Identified in MS1 negative mode' \
                                if ids['ms1_neg'] \
                                else 'Not identified in MS1 negative mode',
                            '')
                        this_row += tablecell % (
                            'positive' if ids['ms2_neg'] else 'nothing',
                            'Identified in MS2 negative mode' \
                                if ids['ms2_neg'] \
                                else 'Not identified in MS2 negative mode',
                            '')
                        this_row = tablerow % this_row
                        table += this_row
        with open(outf, 'w') as f:
            f.write(html_table_template % (title, title, table))

    def known_binders_enrichment(self, valids, bindprop, classif = 'cl5pct'):
        for ltp, d in valids.iteritems():
            for pn, tbl in d.iteritems():
                if 'enr' not in tbl:
                    tbl['enr'] = {}
                if classif not in tbl['enr']:
                    tbl['enr'][classif] = {}
                tbl['enr'][classif]['tp'] = len(
                    filter(lambda (i, oi):
                        tbl[classif][i] and \
                            len(bindprop[ltp] & \
                                set(tbl['identity'][oi].keys())) > 0,
                        enumerate(tbl['i'])
                    )
                )
                tbl['enr'][classif]['fp'] = len(
                    filter(lambda (i, oi):
                        tbl[classif][i] and \
                            len(bindprop[ltp] & \
                                set(tbl['identity'][oi].keys())) == 0,
                        enumerate(tbl['i'])
                    )
                )
                tbl['enr'][classif]['tn'] = len(
                    filter(lambda (i, oi):
                        not tbl[classif][i] and \
                            len(bindprop[ltp] & \
                                set(tbl['identity'][oi].keys())) == 0,
                        enumerate(tbl['i'])
                    )
                )
                tbl['enr'][classif]['fn'] = len(
                    filter(lambda (i, oi):
                        not tbl[classif][i] and \
                            len(bindprop[ltp] & \
                                set(tbl['identity'][oi].keys())) > 0,
                        enumerate(tbl['i'])
                    )
                )
                try:
                    tbl['enr'][classif]['or'] = \
                        float(tbl['enr'][classif]['tp'] * \
                        tbl['enr'][classif]['tn']) / \
                        float(tbl['enr'][classif]['fp'] * \
                        tbl['enr'][classif]['fn'])
                except ZeroDivisionError:
                    tbl['enr'][classif]['or'] = np.inf
                tbl['enr'][classif]['contab'] = np.array([
                    [tbl['enr'][classif]['tp'], tbl['enr'][classif]['fp']],
                    [tbl['enr'][classif]['fn'], tbl['enr'][classif]['tn']]])
                tbl['enr'][classif]['fisher'] = \
                    sp.stats.fisher_exact(tbl['enr'][classif]['contab'])

    def enrichment_barplot(self, valids, classif = 'cl5pct', 
        outf = 'known_binders_enrichment.pdf'):
        w = 0.45
        labels = sorted(filter(lambda ltp:
            valids[ltp]['pos']['enr'][classif]['tp'] + \
                valids[ltp]['pos']['enr'][classif]['fn'] > 0 or \
            valids[ltp]['neg']['enr'][classif]['tp'] + \
                valids[ltp]['neg']['enr'][classif]['fn'] > 0,
            valids.keys()
        ))
        ppvals = map(lambda ltp:
            '%.03f' % valids[ltp]['pos']['enr'][classif]['fisher'][1],
            labels
        )
        npvals = map(lambda ltp:
            '%.03f' % valids[ltp]['neg']['enr'][classif]['fisher'][1],
            labels
        )
        pors = map(lambda ltp:
            '%.03f' % valids[ltp]['pos']['enr'][classif]['fisher'][0],
            labels
        )
        nors = map(lambda ltp:
            '%.03f' % valids[ltp]['neg']['enr'][classif]['fisher'][0],
            labels
        )
        fig = mpl.figure.Figure(figsize = (8, 4))
        cvs = mpl.backends.backend_pdf.FigureCanvasPdf(fig)
        ax = fig.gca()
        ax.bar(np.arange(len(labels)), pors, w, color = '#97BE73', lw = 0.0)
        ax.bar(np.arange(len(labels)) + w, nors, w, color = '#49969A', lw = 0.0)
        ax.set_yscale('symlog')
        ax.set_xticks(np.arange(len(labels)) + w)
        ax.set_xticklabels(labels, rotation = 90)
        ax.set_xlabel('LTPs')
        ax.set_ylabel('Enrichment (odds ratio)')
        cvs.draw()
        fig.tight_layout()
        cvs.print_figure(outf)
        fig.clf()

    def identification_levels(self, ltp = 'STARD10',
        hg = 'PC', classif = None):
        visited = {'neg': set([]), 'pos': set([])}
        labels = ['MS1 & MS2 both +/-', 
            'MS1 & MS2 only +', 'MS1 & MS2 only -',
            'MS1 both +/-, no MS2', 'MS2 both +/-, no MS1',
            'MS1 + and MS2 -', 'MS1 - and MS2 +',
            'Only MS1 +', 'Only MS1 -', 'Only MS2 +', 'Only MS2 -',
            'MS1 both +/-, MS2 +', 'MS1 both +/-, MS2 -',
            'MS1 +, MS2 both +/-', 'MS1 -, MS2 both +/-',
            'Nothing', 'Non %s' % hg,
            'Total', '+/- Total', 'Only + Total', 'Only - Total']
        values = dict(zip(labels, [0] * len(labels)))
        modes = {'pos': '+', 'neg': '-'}
        keys = ['ms1_pos', 'ms2_pos', 'ms1_neg', 'ms2_neg']
        combinations = {
            (True, True, True, True): 'MS1 & MS2 both +/-',
            (True, True, False, False): 'MS1 & MS2 only +',
            (False, False, True, True): 'MS1 & MS2 only -',
            (True, False, True, False): 'MS1 both +/-, no MS2',
            (False, True, False, True): 'MS2 both +/-, no MS1',
            (True, False, False, True): 'MS1 + and MS2 -',
            (False, True, True, False): 'MS1 - and MS2 +',
            (True, False, False, False): 'Only MS1 +',
            (False, False, True, False): 'Only MS1 -',
            (False, True, False, False): 'Only MS2 +',
            (False, False, False, True): 'Only MS2 -',
            (True, True, True, False): 'MS1 both +/-, MS2 +',
            (True, False, True, True): 'MS1 both +/-, MS2 -',
            (True, True, False, True): 'MS1 +, MS2 both +/-',
            (False, True, True, True): 'MS1 -, MS2 both +/-'
        }
        for mode, opp_mod in [('pos', 'neg'), ('neg', 'pos')]:
            tbl = valids[ltp][mode]
            opp_tbl = self.valids[ltp][opp_mod]
            for i, oi in enumerate(tbl['i']):
                if oi not in visited[mode] and \
                    (classif is None or tbl[classif][i]):
                    visited[mode].add(oi)
                    values['Total'] += 1
                    opp_indices = map(lambda opp_oi:
                        np.where(opp_tbl['i'] == opp_oi)[0][0],
                        tbl[opp_mod][oi].keys()
                    )
                    opp_class = classif is None or len(opp_indices) == 0 or \
                        sum(opp_tbl[classif][opp_indices]) > 0
                    if len(tbl[opp_mod][oi]) > 0 and opp_class:
                        values['+/- Total'] += 1
                        for opp_oi in tbl[opp_mod][oi].keys():
                            if classif is None or \
                            any(opp_tbl[classif][opp_indices]):
                                visited[opp_mod].add(opp_oi)
                    else:
                        values['Only %s Total'%modes[mode]] += 1
                    if hg not in tbl['identity'][oi] or (\
                        hg in tbl['identity'][oi] and \
                        not tbl['identity'][oi][hg]['ms1_%s'%mode] and \
                        not tbl['identity'][oi][hg]['ms2_%s'%mode]):
                        if len(set(tbl['identity'][oi]) - set([hg])) > 0:
                            values['Non %s'%hg] += 1
                        else:
                            values['Nothing'] += 1
                    else:
                        comb = map(lambda k: 
                            tbl['identity'][oi][hg][k], 
                            keys
                        )
                        if not opp_class:
                            if opp_mod == 'neg':
                                comb[2] = False
                                comb[3] = False
                            else:
                                comb[0] = False
                                comb[1] = False
                        comb = tuple(comb)
                        values[combinations[comb]] += 1
        return values

    '''
    {'MS1 both +/-, MS2 +': 7, 'Only MS2 -': 0, 'MS1 - and MS2 +': 2,
     'MS1 both +/-, MS2 -': 1, 'MS1 +, MS2 both +/-': 0, '+/- Total': 46,
     'MS2 both +/-, no MS1': 0, 'Non PC': 79, 'Nothing': 221,
     'MS1 & MS2 only +': 12, 'MS1 & MS2 both +/-': 8,
     'MS1 & MS2 only -': 0, 'Only + Total': 298, 'Only MS1 -': 9,
     'MS1 + and MS2 -': 0, 'MS1 both +/-, no MS2': 17,
     'MS1 -, MS2 both +/-': 1, 'Total': 448, 'Only - Total': 104,
     'Only MS1 +': 68, 'Only MS2 +': 23}
    '''

    def plot_identification_levels(self, idlevels, ltp, hg,
        fname = '%s-%s-classes.pdf'):
        fname = fname % (ltp, hg)
        w = 0.8 / len(idlevels)
        cols = ['#97BE73', '#49969A', '#996A44', '#FDD73F']
        labels = ['MS1 & MS2 both +/-', 
            'MS1 & MS2 only +', 'MS1 & MS2 only -',
            'MS1 both +/-, no MS2', 'MS2 both +/-, no MS1', 
            'MS1 + and MS2 -', 'MS1 - and MS2 +',
            'Only MS1 +', 'Only MS1 -', 'Only MS2 +', 'Only MS2 -', 
            'MS1 both +/-, MS2 +', 'MS1 both +/-, MS2 -',
            'MS1 +, MS2 both +/-', 'MS1 -, MS2 both +/-',
            'Nothing', 'Non %s'%hg,
            'Total', '+/- Total', 'Only + Total', 'Only - Total']
        fig = mpl.figure.Figure(figsize = (8, 4))
        cvs = mpl.backends.backend_pdf.FigureCanvasPdf(fig)
        ax = fig.gca()
        i = 0
        for lab in sorted(idlevels.keys()):
            ax.bar(np.arange(len(labels)) + i * w, 
                map(lambda l: idlevels[lab][l], labels), 
                w, color = cols[i], lw = 0.0)
            i += 1
        lhandles = map(lambda (i, lab):
            mpl.patches.Patch(color = cols[i], label = lab),
            enumerate(sorted(idlevels.keys())))
        leg = ax.legend(handles = lhandles)
        ax.set_xticks(np.arange(len(labels)) + w * i / 2.0)
        ax.set_xticklabels(labels, rotation = 90)
        ax.set_xlabel('Identification levels')
        ax.set_ylabel('Number of features')
        cvs.draw()
        fig.tight_layout()
        cvs.print_figure(fname)
        fig.clf()

    def mz_report(self, ltp, mode, mz):
        self.sort_alll('mz')
        tbl = self.valids[ltp][mode]
        ui = tbl['mz'].searchsorted(mz)
        i = ui if tbl['mz'][ui] - mz < mz - tbl['mz'][ui - 1] else ui - 1
        oi = tbl['i'][i]
        sys.stdout.write('\n\t:: Looking up %.08f\n'\
            '\t -- The closest m/z found is %.08f\n'\
            '\t -- Current index is %u\n'\
            '\t -- Original index is %u\n'\
            '\t -- MS1 headgroups identified: %s\n'\
            '\t -- MS2 headgroups identified: %s\n'\
            '\t -- MS1 fatty acids identified: %s\n'\
            '\t -- MS2 fatty acids identified: %s\n'\
            '\t -- Headgroups based on fatty acids: %s\n'\
            '\t -- Combined identity: %s\n'\
            '\n' % \
            (mz, tbl['mz'][i], i, oi, 
            ', '.join(sorted(list(tbl['ms1hg'][oi]))),
            ', '.join(sorted(list(tbl['ms2hg'][oi]))) \
                if oi in tbl['ms2hg'] and \
                    type(tbl['ms2hg'][oi]) is set else '',
            ', '.join(map(lambda (hg, fa): 
                    '%s: %s' % (hg, ', '.join(list(fa))), 
                    tbl['ms1fa'][oi].iteritems()
                )
            ),
            ', '.join(list(tbl['ms2fa'][oi])) \
                if oi in tbl['ms2fa'] else '',
            ', '.join(list(tbl['hgfa'][oi])),
            ', '.join(list(tbl['combined_hg'][oi]))
            )
        )

    def _database_details_list(self, lip):
        if lip is None: lip = []
        return '\n'.join(map(lambda l:
            '⚫ %s\t%.04f\t%s' % (l[4], l[5], l[2]),
            lip
        ))

    def _fragment_details_list(self, ms2r, ms2files, path):
        fractions = {9: 'A09', 10: 'A10', 11: 'A11', 12: 'A12', 1: 'B01'}
        if ms2r is not None:
            try:
                sort = ms2r[:,3].argsort()[::-1]
            except:
                print ms2r.shape
        else:
            sort = []
        return '\n'.join(map(lambda i:
            '⚫ %s (%.02f)\n    @ file: %s\n    @ scan no. %u' % \
                (ms2r[i,0], ms2r[i,3],
                    ms2files[fractions[int(ms2r[i,4])]].replace(path, '')[1:],
                    int(ms2r[i,5])),
            filter(lambda i:
                ms2r[i,0] != 'unknown',
                sort
            )
        ))

    def _features_table_row(self, ltp, mod, tbl, oi, i, fits_profile, drift):
        tablecell = '\t\t\t<td class="%s" title="%s">\n\t\t\t\t%s'\
            '\n\t\t\t</td>\n'
        tableccell = '\t\t\t<td class="%s" title="%s"'\
            ' onclick="showTooltip(event);">'\
            '\n\t\t\t\t%s\n\t\t\t</td>\n'
        row = []
        original_mz = tbl['mz'][i] / drift
        row.append(tablecell % \
            ('nothing', 'm/z measured; %s, %s mode; raw: %.07f, '\
                'recalibrated: %.07f' % \
                (ltp, 'negative' if mod == 'neg' else 'positive',
                    original_mz, tbl['mz'][i]),
            '%.04f (%.04f)' % (tbl['mz'][i], original_mz)))
        row.append(tableccell % (
            ('nothing clickable',
                self._database_details_list(tbl['lip'][oi]),
                'see DB records') \
                    if tbl['lip'][oi] is not None and \
                        len(tbl['lip'][oi]) > 0 \
                    else \
            ('nothing',
                'Lookup of this m/z resulted no records',
                'No DB records')
        ))
        row.append(tablecell % \
            ('nothing', 'Possible headgroups based on database records',
            ', '.join(tbl['ms1hg'][oi])))
        row.append(tableccell % \
            (('nothing clickable',
                    self._fragment_details_list(tbl['ms2r'][oi],
                        self.ms2files[ltp][mod], self.ltpdirs[0]),
                    'see MS2 frags.') \
                if oi in tbl['ms2r'] else \
            ('nothing', 'No MS2 results for this feature', 'No MS2'))
        )
        row.append(tablecell % \
            ('nothing', 'Possible MS2 headgroups based on fragmets lookup',
            ', '.join(tbl['ms2hg'][oi] \
                if oi in tbl['ms2hg'] and tbl['ms2hg'][oi] is not None else '')
        ))
        row.append(tablecell % \
            ('nothing', 'MS2 headroups and fatty acids identified based on '\
                'MS1 and MS2 new rule sets.',
            '' if oi not in self.valids[ltp][mod]['ms2f'] \
                else str(self.valids[ltp][mod]['ms2f'][oi])
        ))
        row.append(tablecell % \
            ('nothing',
            'Possible fatty acids based on database records',
            '; '.join(map(lambda (hg, fa):
                '%s: %s' % (hg, ', '.join(list(fa))),
                tbl['ms1fa'][oi].iteritems()
            )) if len(tbl['ms1fa'][oi]) else '')
        )
        row.append(tablecell % \
            ('nothing',
            'Fatty acids identified in MS2',
            ', '.join(list(tbl['ms2fa'][oi])) \
                if oi in tbl['ms2fa'] and len(tbl['ms2fa'][oi]) else '')
        )
        row.append(tablecell % \
            ('nothing',
            'Combined identity (MS1, MS2, fatty acids)',
            ', '.join(list(tbl['combined_hg'][oi])) \
                if len(tbl['combined_hg'][oi]) else '')
        )
        row.append(tablecell % (
            'positive' if tbl[fits_profile][i] else 'nothing',
            'Fits protein profile' if tbl[fits_profile][i] \
                else 'Does not fit protein profile',
            'Yes' if tbl[fits_profile][i] else 'No'
        ))
        return row

    def features_table(self, filename = None,
        fits_profile = 'cl70pct'):
        hdr = ['+m/z', '+Database', '+MS1 HGs',
            '+MS2 frags', '+MS2 HGs', '+ID',
            '+MS1 FAs', '+MS2 FAs',
            '+MS1&2',
            '+Fits protein',
            '-m/z', '-Database', '-MS1 HGs',
            '-MS2 frags', '-MS2 HGs', '-ID',
            '-MS1 FAs', '-MS2 FAs',
            '-MS1&2',
            '-Fits protein']
        
        filename = 'results_%s_identities_details' % self.today() \
            if filename is None \
            else filename
        tablerow = '\t\t<tr>\n%s\t\t</tr>\n'
        tablehcell = '\t\t\t<th>\n\t\t\t\t%s\n\t\t\t</th>\n'
        tablecell = '\t\t\t<td class="%s" title="%s">\n\t\t\t\t%s'\
            '\n\t\t\t</td>\n'
        
        # navigation html
        if not os.path.isdir(filename):
            os.mkdir(filename)
        navigation = '%s.html' % filename
        title = 'Identification details of all features'
        navhtml = ''
        navcnum = 7
        navrnum = len(self.valids) / navcnum + 1
        ltps = sorted(self.valids.keys())
        for i in xrange(navrnum):
            thisRow = []
            for j in xrange(navcnum):
                thisRow.append(tablecell % \
                    (('nothing',
                    'See identifications for all features of %s' % \
                        (ltps[j*navrnum + i]),
                    '<a href="%s/%s_%s.html">%s</a>' % \
                        (filename, filename, 
                        ltps[j*navrnum + i], ltps[j*navrnum + i])
                    ) if j*navrnum + i < len(ltps) else ('nothing', '', ''))
                )
            navhtml += tablerow % '\n'.join(thisRow)
        with open(navigation, 'w') as f:
            f.write(self.html_table_template % (title, title, navhtml))
        
        for ltp, d in self.valids.iteritems():
            title = '%s: identities for all features, detailed' % ltp
            thisFilename = '%s/%s_%s.html' % (filename, filename, ltp)
            table = ''
            hrow = ''
            for coln in hdr:
                hrow += tablehcell % coln
            hrow = tablerow % hrow
            table = hrow
            visited = {'pos': set([]), 'neg': set([])}
            thisRow = []
            thisRow.append(tablecell % ('rowname', ltp, ltp))
            thisRow += map(lambda i:
                tablecell % ('nothing', '', ''),
                xrange(len(hdr) - 1)
            )
            table += (tablerow % '\n'.join(thisRow))
            for mod, tbl in d.iteritems():
                opp_mod = 'neg' if mod == 'pos' else 'pos'
                drift = 1.0 if not self.ltps_drifts \
                    or 'recalibrated' not in tbl \
                    or not tbl['recalibrated'] \
                    else self.ppm2ratio(np.nanmedian(
                        np.array(self.ltps_drifts[ltp][mod].values())
                    ))
                opp_drift = 1.0 if not self.ltps_drifts \
                    or 'recalibrated' not in tbl \
                    or not tbl['recalibrated'] \
                    else self.ppm2ratio(np.nanmedian(
                        np.array(self.ltps_drifts[ltp][opp_mod].values())
                    ))
                for i, oi in enumerate(tbl['i']):
                    if oi not in visited[mod]:
                        thisRow = {'pos': [], 'neg': []}
                        visited[mod].add(oi)
                        thisRow[mod].append(
                            self._features_table_row(ltp, mod, tbl,
                                oi, i, fits_profile, drift))
                        try:
                            if len(tbl[opp_mod][oi]) == 0:
                                thisRow[opp_mod] = [map(lambda i:
                                    tablecell % ('nothing', '', ''),
                                    xrange(len(hdr) / 2)
                                )]
                            else:
                                for opp_oi in tbl[opp_mod][oi].keys():
                                    opp_row = []
                                    visited[opp_mod].add(opp_oi)
                                    opp_i = np.where(d[opp_mod]['i'] == \
                                        opp_oi)[0][0]
                                    thisRow[opp_mod].append(
                                        self._features_table_row(
                                            ltp, opp_mod, d[opp_mod],
                                            opp_oi, opp_i, fits_profile,
                                            opp_drift
                                        )
                                    )
                        except KeyError:
                            print ltp, mod, opp_mod
                        for pos_row in thisRow['pos']:
                            for neg_row in thisRow['neg']:
                                #try:
                                table += tablerow % ('\n%s\n%s\n' % \
                                    ('\n'.join(pos_row), '\n'.join(neg_row)))
                                #except TypeError:
                                    #print pos_row
                                    #print neg_row
            with open(thisFilename, 'w') as f:
                f.write(self.html_table_template % (title, title, table))

    def combined_table(self, valids, filename = 'headgroups_combined.html',
        include = 'cl70pct', identity = 'combined_hg'):
        tablecell = '\t\t\t<td class="%s" title="%s">\n\t\t\t\t%s'\
            '\n\t\t\t</td>\n'
        tableccell = '\t\t\t<td class="%s" title="%s"'\
            ' onclick="showTooltip(event);">'\
            '\n\t\t\t\t%s\n\t\t\t</td>\n'
        tablerow = '\t\t<tr>\n%s\t\t</tr>\n'
        tablehcell = '\t\t\t<th>\n\t\t\t\t%s\n\t\t\t</th>\n'
        title = 'Headgroups, combined identification'
        sort_alll(valids, 'mz')
        all_hgs = set([])
        for ltp, d in valids.iteritems():
            for mod, tbl in d.iteritems():
                for i, oi in enumerate(tbl['i']):
                    if tbl[include][i]:
                        all_hgs = all_hgs | tbl[identity][oi]
        all_hgs = sorted(list(all_hgs))
        data = dict(map(lambda ltp:
                (ltp, dict(map(lambda hg:
                    (hg, []), 
                    all_hgs
                ))),
                valids.keys()
            ))
        for ltp, d in valids.iteritems():
            for mod, tbl in d.iteritems():
                for i, oi in enumerate(tbl['i']):
                    if tbl[include][i]:
                        for hg in tbl[identity][oi]:
                            adds = ';'.join(uniqList(\
                                list(tbl['lip'][oi][\
                                    np.where(tbl['lip'][oi][:,7] == hg)[0], 4\
                                ])\
                            ))
                            fa = ','.join(list(tbl['combined_fa'][oi][hg])) \
                                if hg in tbl['combined_fa'][oi] else \
                                tbl['ms2fas'][oi] if oi in tbl['ms2fas'] else \
                                ','.join(list(tbl['ms1fa'][oi][hg]))
                            data[ltp][hg].append(
                                ['%.04f' % tbl['mz'][i], adds, fa]
                            )
        hdr = map(lambda hg:
            tablehcell % hg,
            all_hgs
        )
        table = tablerow % '\n'.join([tablehcell % ('')] + hdr)
        for ltp in sorted(valids.keys()):
            thisRow = [tablecell % ('rowname', '', ltp)]
            for hg in all_hgs:
                c = data[ltp][hg]
                if not c.__len__():
                    thisRow.append(tablecell % ('nothing', '', ''))
                else:
                    thisRow.append(tableccell % (
                        'positive clickable' \
                            if 'neg' not in map(lambda i: i[1], c) else \
                        'negative clickable' \
                            if 'pos' not in map(lambda i: i[1], c) else \
                        'both clickable',
                        '%s\n%s' % (hg, '\n'.join(map(lambda l:
                            '⚫ %s' % '     '.join(l),
                            c
                        ))),
                        ''
                    ))
            table += tablerow % '\n'.join(thisRow)
        with open(filename, 'w') as f:
            f.write(html_table_template % (title, title, table))
    
    def i2oi(self, protein, mode, i):
        return self.valids[protein][mode]['i'][i] \
            if (0 <= i < len(self.valids[protein][mode]['i'])) \
            else None
    
    def oi2i(self, protein, mode, oi):
        i = np.where(self.valids[protein][mode]['i'] == oi)[0]
        return i[0] if len(i) > 0 else None
    
    def ms2identities_summary(self, string = False):
        result = \
            dict(
                map(
                    lambda (protein, d):
                        (protein,
                        dict(
                            map(
                                lambda (mode, tbl):
                                    (mode,
                                    filter(
                                        lambda i:
                                            i is not None,
                                        map(
                                            lambda f:
                                                None \
                                                if len(f.identities) == 0 \
                                                else reduce(
                                                    lambda (a, b):
                                                        a | b,
                                                    f.identities
                                                ),
                                            tbl['ms2f'].values()
                                        ),
                                    )
                                    ),
                                d.iteritems()
                            )
                            )
                        ),
                    self.valids.iteritems()
                )
            )
        
        if not string:
            return result
        
        result_str = \
            ''.join(
                map(
                    lambda (protein, d):
                        'Protein: %s\n%s\n' % (
                            protein,
                            ''.join(
                                map(
                                    lambda (mode, res):
                                        '\tMode: %s\n\t   %s\n' % (
                                            mode,
                                            '\n\t   '.join(
                                                map(
                                                    lambda hg:
                                                        '%s (detected %u '\
                                                            'times)' % \
                                                                (hg,
                                                                collections.\
                                                                Counter(res)\
                                                                    [hg]
                                                                ),
                                                    uniqList(res)
                                                ),
                                            )
                                        ),
                                    d.iteritems()
                                )
                            )
                        ),
                    result.iteritems()
                )
            )
        return result_str
    
    def bindprop_latex_table(self, only_with_ms_data = False, outfile = 'binding_properties_table.tex'):
        allhgs = \
        sorted(
            list(
                reduce(
                    lambda lst, lips:
                        lst | lips,
                    map(
                        lambda (protein, lips):
                            lips,
                        filter(
                            lambda (protein, lips):
                                not only_with_ms_data or protein in self.samples_upper,
                            self.bindprop.iteritems()
                        )
                    )
                )
            )
        )
        tbl = r'''\begin{tabular}{l%s}
            & %s \\
            %s \\
        \end{tabular}
        '''
        colalign = 'l' * len(allhgs)
        rownames = \
        sorted(
            filter(
                lambda protein:
                    not only_with_ms_data or protein in self.samples_upper,
                self.bindprop.keys()
            )
        )
        hdr = ' & '.join(
            map(
                lambda lip:
                    '\\rot{%s}' % lip,
                allhgs
            )
        )
        cells = \
        '\\\\\n'.join(
            map(
                lambda protein:
                    '%s & %s' % (
                        protein,
                        ' & '.join(
                            map(
                                lambda lip:
                                    r'\cellcolor{emblgreen}' \
                                    if lip in self.bindprop[protein] \
                                    else '',
                                allhgs
                            )
                        )
                    ),
                rownames
            )
        )
        with open(outfile, 'w') as f:
            f.write(tbl % (colalign, hdr, cells))
